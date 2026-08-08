
import streamlit as st
import json
import streamlit.components.v1 as components
import pandas as pd
import requests
import base64
from collections import Counter, defaultdict
from itertools import product, combinations
from pathlib import Path
from io import BytesIO
# Selection Engine V1 disatukan dalam app.py untuk Streamlit Cloud.
from collections import Counter
import math


def _pad4(value):
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return "".join(char for char in text if char.isdigit()).zfill(4)[-4:]


def _key4(value):
    return "".join(sorted(_pad4(value)))


def _pair_key(value):
    return "".join(sorted(str(value).zfill(2)[-2:]))


def _pairs(numbers):
    rows, seen = [], set()
    for source, number in zip(("1st", "2nd", "3rd"), numbers):
        for position, pair in zip(
            ("Front", "Middle", "Back"),
            (number[:2], number[1:3], number[2:]),
        ):
            key = _pair_key(pair)
            if key in seen:
                continue
            seen.add(key)
            rows.append((source, position, pair))
    return rows






def _candidates(numbers):
    existing = sorted(set("".join(numbers)))
    missing = sorted(set("0123456789") - set(existing))
    pairs = _pairs(numbers)
    rows = {}

    def add(number, route, source, position):
        key = _key4(number)
        row = rows.setdefault(
            key,
            {
                "key": key,
                "No": number,
                "routes": set(),
                "slots": set(),
            },
        )
        row["routes"].add(route)
        row["slots"].add(f"{source}-{position}")

    for source, position, pair in pairs:
        for missing_digit in missing:
            for existing_digit in existing:
                add(
                    f"{pair}{missing_digit}{existing_digit}",
                    "V1",
                    source,
                    position,
                )
        for pool, route in (
            (missing, "V2-Missing"),
            (existing, "V2-Existing"),
        ):
            for first_digit in pool:
                for second_digit in pool:
                    if first_digit != second_digit:
                        add(
                            f"{pair}{first_digit}{second_digit}",
                            route,
                            source,
                            position,
                        )

    for row in rows.values():
        row["tokens"] = {
            *(f"route:{route}" for route in row["routes"]),
            *(f"slot:{slot}" for slot in row["slots"]),
            *(
                f"route_slot:{route}|{slot}"
                for route in row["routes"]
                for slot in row["slots"]
            ),
        }
    return list(rows.values())


class _Model:
    def __init__(self):
        self.exposure = Counter()
        self.wins = Counter()
        self.total_exposure = 0
        self.total_wins = 0

    def update(self, candidates, targets):
        for row in candidates:
            won = int(row["key"] in targets)
            self.total_exposure += 1
            self.total_wins += won
            for token in row["tokens"]:
                self.exposure[token] += 1
                self.wins[token] += won

    def score(self, row, prefixes):
        global_rate = (self.total_wins + 1) / (self.total_exposure + 100)
        values = []
        for token in row["tokens"]:
            if not token.startswith(prefixes):
                continue
            exposure = self.exposure[token]
            if exposure < 20:
                continue
            rate = (self.wins[token] + 8 * global_rate) / (exposure + 8)
            reliability = min(1.0, exposure / 250)
            values.append(
                math.log(max(rate, 1e-9) / global_rate) * reliability
            )
        return sum(values) / math.sqrt(max(1, len(values)))


def build_selection_engine(history, first, second, third, lookback=100):
    """Selection V1: Pair Slot sahaja berdasarkan audit walk-forward."""
    if history is None or len(history) < 3:
        return {"combined": []}

    frame = history.reset_index(drop=True)
    start = max(0, len(frame) - int(lookback) - 1)
    model = _Model()
    for index in range(start, len(frame) - 1):
        source = [_pad4(frame.iloc[index][col]) for col in ("first", "second", "third")]
        target = {
            _key4(frame.iloc[index + 1][col])
            for col in ("first", "second", "third")
        }
        model.update(_candidates(source), target)

    current = _candidates([_pad4(first), _pad4(second), _pad4(third)])
    pair_ranked = sorted(
        current,
        key=lambda row: (-model.score(row, ("slot:",)), row["No"]),
    )
    pair = [row["No"] for row in pair_ranked[:10]]
    return {"combined": pair}


def build_second_prize_2d_carry_engine(
    history, second, bridge_v1_df, bridge_v2_df, lookback=100
):
    """Pilih kedudukan 2D paling kerap dibawa, kemudian tapis Bridge."""
    second_no = _pad4(second)
    positions = list(combinations(range(4), 2))
    audit_rows = []
    frame = history.reset_index(drop=True) if history is not None else pd.DataFrame()
    transition_start = max(0, len(frame) - int(lookback) - 1)
    transition_end = max(0, len(frame) - 1)

    for left, right in positions:
        hits = 0
        transitions = 0
        for index in range(transition_start, transition_end):
            source_second = _pad4(frame.iloc[index]["second"])
            target_numbers = [
                _pad4(frame.iloc[index + 1][column])
                for column in ("first", "second", "third")
            ]
            duo_counts = Counter((source_second[left], source_second[right]))
            carried = any(
                all(Counter(number)[digit] >= count for digit, count in duo_counts.items())
                for number in target_numbers
            )
            transitions += 1
            hits += int(carried)

        current_duo = second_no[left] + second_no[right]
        audit_rows.append({
            "Position": f"{left + 1}+{right + 1}",
            "2D Semasa": current_duo,
            "Historical Hit": hits,
            "Transitions": transitions,
            "Hit Rate %": round(hits / transitions * 100, 1) if transitions else 0,
        })

    audit_df = pd.DataFrame(audit_rows).sort_values(
        ["Historical Hit", "Position"],
        ascending=[False, True],
    ).reset_index(drop=True)
    audit_df.insert(0, "Priority", range(1, len(audit_df) + 1))
    selected_position = str(audit_df.iloc[0]["Position"]) if not audit_df.empty else ""
    selected_duo = str(audit_df.iloc[0]["2D Semasa"]) if not audit_df.empty else ""
    selected_key = "".join(sorted(selected_duo))
    selected_counts = Counter(selected_key)
    all_duo_rows = []
    seen_duo_keys = set()
    for left, right in positions:
        display = second_no[left] + second_no[right]
        key = "".join(sorted(display))
        if key in seen_duo_keys:
            continue
        seen_duo_keys.add(key)
        all_duo_rows.append((display, Counter(key)))

    def filter_bridge(bridge_frame, route, selected_only=False):
        if bridge_frame is None or bridge_frame.empty or "No" not in bridge_frame.columns:
            return pd.DataFrame(columns=["No", "2D dari 2nd", "Bridge"])
        rows = []
        seen = set()
        for number in bridge_frame["No"].astype(str):
            key = _key4(number)
            if key in seen:
                continue
            number_counts = Counter(_pad4(number))
            matches = [
                display for display, duo_counts in all_duo_rows
                if all(
                    number_counts[digit] >= count
                    for digit, count in duo_counts.items()
                )
            ]
            if selected_only:
                matches = [
                    selected_duo
                    for _ in (0,)
                    if all(
                        number_counts[digit] >= count
                        for digit, count in selected_counts.items()
                    )
                ]
            if not matches:
                continue
            seen.add(key)
            rows.append({
                "No": _pad4(number),
                "2D dari 2nd": " / ".join(matches),
                "Bridge": route,
            })
        return pd.DataFrame(rows)

    all_v1 = filter_bridge(bridge_v1_df, "V1")
    all_v2 = filter_bridge(bridge_v2_df, "V2")
    selected_v1 = filter_bridge(bridge_v1_df, "V1", selected_only=True)
    selected_v2 = filter_bridge(bridge_v2_df, "V2", selected_only=True)
    all_duos = [display for display, _ in all_duo_rows]
    return (
        audit_df, all_duos, all_v1, all_v2,
        selected_position, selected_duo, selected_v1, selected_v2,
    )


@st.cache_data(show_spinner=False)
def build_2d_missing_first_digit_engine(
    history, first, second, third, bridge_v1_df, bridge_v2_df, lookback=100
):
    """2D hadiah kedua + missing + satu digit hadiah pertama."""
    second_positions = list(combinations(range(4), 2))
    first_positions = range(4)
    frame = history.reset_index(drop=True) if history is not None else pd.DataFrame()
    start = max(0, len(frame) - int(lookback) - 1)
    stop = max(0, len(frame) - 1)
    route_hits = Counter()
    second_draw_hits = Counter()
    first_draw_hits = Counter()
    transitions = 0

    for index in range(start, stop):
        source_first = _pad4(frame.iloc[index]["first"])
        source_second = _pad4(frame.iloc[index]["second"])
        source_top3 = [
            _pad4(frame.iloc[index][column])
            for column in ("first", "second", "third")
        ]
        missing_digits = sorted(set("0123456789") - set("".join(source_top3)))
        targets = {
            _key4(frame.iloc[index + 1][column])
            for column in ("first", "second", "third")
        }
        second_hit_now = set()
        first_hit_now = set()
        for left, right in second_positions:
            second_label = f"{left + 1}+{right + 1}"
            duo = source_second[left] + source_second[right]
            for first_index in first_positions:
                hit = any(
                    _key4(f"{duo}{missing}{source_first[first_index]}") in targets
                    for missing in missing_digits
                )
                if hit:
                    route_hits[(second_label, first_index + 1)] += 1
                    second_hit_now.add(second_label)
                    first_hit_now.add(first_index + 1)
        for label in second_hit_now:
            second_draw_hits[label] += 1
        for position in first_hit_now:
            first_draw_hits[position] += 1
        transitions += 1

    second_audit = pd.DataFrame([
        {
            "Kedudukan 2D": f"{left + 1}+{right + 1}",
            "Hit Draw": second_draw_hits[f"{left + 1}+{right + 1}"],
            "Draw Diuji": transitions,
            "Hit Rate %": round(
                second_draw_hits[f"{left + 1}+{right + 1}"] / transitions * 100, 1
            ) if transitions else 0,
        }
        for left, right in second_positions
    ]).sort_values(["Hit Draw", "Kedudukan 2D"], ascending=[False, True]).reset_index(drop=True)

    first_audit = pd.DataFrame([
        {
            "Kedudukan Digit 1st": position,
            "Hit Draw": first_draw_hits[position],
            "Draw Diuji": transitions,
            "Hit Rate %": round(first_draw_hits[position] / transitions * 100, 1)
            if transitions else 0,
        }
        for position in range(1, 5)
    ]).sort_values(
        ["Hit Draw", "Kedudukan Digit 1st"], ascending=[False, True]
    ).reset_index(drop=True)

    joint_audit = pd.DataFrame([
        {
            "Kedudukan 2D": f"{left + 1}+{right + 1}",
            "Kedudukan Digit 1st": first_position,
            "Hit Draw": route_hits[(f"{left + 1}+{right + 1}", first_position)],
            "Draw Diuji": transitions,
            "Hit Rate %": round(
                route_hits[(f"{left + 1}+{right + 1}", first_position)]
                / transitions * 100, 1
            ) if transitions else 0,
        }
        for left, right in second_positions
        for first_position in range(1, 5)
    ]).sort_values(
        ["Hit Draw", "Kedudukan 2D", "Kedudukan Digit 1st"],
        ascending=[False, True, True],
    ).reset_index(drop=True)

    best_second_hits = int(second_audit["Hit Draw"].max()) if not second_audit.empty else 0
    best_first_hits = int(first_audit["Hit Draw"].max()) if not first_audit.empty else 0
    selected_second = second_audit.loc[
        second_audit["Hit Draw"].eq(best_second_hits), "Kedudukan 2D"
    ].astype(str).tolist()
    selected_first = first_audit.loc[
        first_audit["Hit Draw"].eq(best_first_hits), "Kedudukan Digit 1st"
    ].astype(int).tolist()

    current_first = _pad4(first)
    current_second = _pad4(second)
    current_top3 = [_pad4(first), _pad4(second), _pad4(third)]
    current_missing = sorted(set("0123456789") - set("".join(current_top3)))
    bridge_v1 = {
        _key4(number): _pad4(number)
        for number in bridge_v1_df.get("No", pd.Series(dtype=str)).astype(str)
    } if bridge_v1_df is not None and not bridge_v1_df.empty else {}
    bridge_v2 = {
        _key4(number): _pad4(number)
        for number in bridge_v2_df.get("No", pd.Series(dtype=str)).astype(str)
    } if bridge_v2_df is not None and not bridge_v2_df.empty else {}

    pair_groups = []
    pair_group_lookup = {}
    for left, right in second_positions:
        second_label = f"{left + 1}+{right + 1}"
        duo = current_second[left] + current_second[right]
        duo_key = _pair_key(duo)
        if duo_key not in pair_group_lookup:
            pair_group_lookup[duo_key] = {
                "2D": duo, "Kedudukan": [],
            }
            pair_groups.append(pair_group_lookup[duo_key])
        pair_group_lookup[duo_key]["Kedudukan"].append(second_label)

    all_rows = []
    for pair_group in pair_groups:
        duo = pair_group["2D"]
        second_label = " / ".join(dict.fromkeys(pair_group["Kedudukan"]))
        for first_index in first_positions:
            first_digit = current_first[first_index]
            for missing in current_missing:
                generated = f"{duo}{missing}{first_digit}"
                key = _key4(generated)
                all_rows.append({
                    "Kedudukan 2D": second_label,
                    "2D": duo,
                    "Missing": missing,
                    "Kedudukan Digit 1st": first_index + 1,
                    "Digit 1st": first_digit,
                    "No Terhasil": generated,
                    "Bridge V1": bridge_v1.get(key, ""),
                    "Bridge V2": bridge_v2.get(key, ""),
                })
    all_candidates = pd.DataFrame(all_rows)
    selected_candidates = all_candidates[
        all_candidates["Kedudukan 2D"].apply(
            lambda value: any(
                position in [part.strip() for part in str(value).split("/")]
                for position in selected_second
            )
        )
        & all_candidates["Kedudukan Digit 1st"].isin(selected_first)
    ].copy() if not all_candidates.empty else all_candidates.copy()
    return {
        "second_audit": second_audit,
        "first_audit": first_audit,
        "joint_audit": joint_audit,
        "selected_second": selected_second,
        "selected_first": selected_first,
        "missing": current_missing,
        "selected": selected_candidates,
        "all": all_candidates,
    }


@st.cache_data(show_spinner=False)
def build_2d_route_signal(history, first, second, third, lookback=100):
    """Pilih laluan 2D semasa melalui padanan keadaan sejarah."""
    frame = history.reset_index(drop=True) if history is not None else pd.DataFrame()
    start = max(0, len(frame) - int(lookback) - 1)
    stop = max(0, len(frame) - 1)
    rows = []

    def occurrence_pairs(first_no, third_no):
        pool = _pad4(first_no) + _pad4(third_no)
        return sorted({
            "".join(sorted((pool[left], pool[right])))
            for left, right in combinations(range(len(pool)), 2)
        })

    def state_values(first_no, second_no, third_no):
        numbers = [_pad4(first_no), _pad4(second_no), _pad4(third_no)]
        return {
            "missing": 10 - len(set("".join(numbers))),
            "second_unique": len(set(numbers[1])),
            "first_unique": len(set(numbers[0])),
            "third_unique": len(set(numbers[2])),
            "repeat_prizes": sum(len(set(number)) < 4 for number in numbers),
            "second_double": int(len(set(numbers[1])) < 4),
        }

    for index in range(start, stop):
        source_first = _pad4(frame.iloc[index]["first"])
        source_second = _pad4(frame.iloc[index]["second"])
        source_third = _pad4(frame.iloc[index]["third"])
        source_numbers = [source_first, source_second, source_third]
        targets = {
            _key4(frame.iloc[index + 1][column])
            for column in ("first", "second", "third")
        }
        missing_digits = sorted(
            set("0123456789") - set("".join(source_numbers))
        )
        suffixes = occurrence_pairs(source_first, source_third)
        missing_hit = False
        first_third_hit = False
        for left, right in combinations(range(4), 2):
            duo = source_second[left] + source_second[right]
            if not missing_hit:
                missing_hit = any(
                    _key4(f"{duo}{missing}{first_digit}") in targets
                    for missing in missing_digits
                    for first_digit in source_first
                )
            if not first_third_hit:
                first_third_hit = any(
                    _key4(f"{duo}{suffix}") in targets
                    for suffix in suffixes
                )
        row = state_values(source_first, source_second, source_third)
        row.update({
            "missing_hit": missing_hit,
            "first_third_hit": first_third_hit,
        })
        rows.append(row)

    current = state_values(first, second, third)
    state_groups = [
        ("missing", "second_unique"),
        ("missing", "repeat_prizes"),
        ("missing", "second_double"),
        ("missing", "first_unique", "third_unique"),
    ]
    votes = Counter()
    evidence = []
    for features in state_groups:
        matches = [
            row for row in rows
            if all(row[feature] == current[feature] for feature in features)
        ]
        if len(matches) < 3:
            continue
        missing_rate = sum(row["missing_hit"] for row in matches) / len(matches)
        first_third_rate = sum(
            row["first_third_hit"] for row in matches
        ) / len(matches)
        if missing_rate > first_third_rate:
            vote = "2D + Missing"
        elif first_third_rate > missing_rate:
            vote = "2D + 1st & 3rd"
        else:
            vote = "Seimbang"
        votes[vote] += 1
        evidence.append({
            "Keadaan": " + ".join(features),
            "Padanan": len(matches),
            "2D + Missing %": round(missing_rate * 100, 1),
            "2D + 1st & 3rd %": round(first_third_rate * 100, 1),
            "Signal": vote,
        })

    decisive = {
        name: count for name, count in votes.items() if name != "Seimbang"
    }
    if not decisive:
        signal = "Seimbang"
        support = 0
    else:
        best = max(decisive.values())
        leaders = [name for name, count in decisive.items() if count == best]
        signal = leaders[0] if len(leaders) == 1 else "Seimbang"
        support = best if len(leaders) == 1 else 0
    return {
        "signal": signal,
        "support": support,
        "tested_states": len(evidence),
        "evidence": pd.DataFrame(evidence),
    }


@st.cache_data(show_spinner=False)
def build_2d_first_third_pair_engine(history, first, second, third, lookback=100):
    """2D daripada 2nd + dua digit berdasarkan kemunculan sebenar 1st dan 3rd."""
    second_positions = list(combinations(range(4), 2))
    frame = history.reset_index(drop=True) if history is not None else pd.DataFrame()
    start = max(0, len(frame) - int(lookback) - 1)
    stop = max(0, len(frame) - 1)
    position_hits = Counter()
    transitions = 0

    def occurrence_pairs(first_no, third_no):
        pool = _pad4(first_no) + _pad4(third_no)
        return sorted({
            "".join(sorted((pool[left], pool[right])))
            for left, right in combinations(range(len(pool)), 2)
        })

    for index in range(start, stop):
        source_first = _pad4(frame.iloc[index]["first"])
        source_second = _pad4(frame.iloc[index]["second"])
        source_third = _pad4(frame.iloc[index]["third"])
        suffixes = occurrence_pairs(source_first, source_third)
        target_keys = {
            _key4(frame.iloc[index + 1][column])
            for column in ("first", "second", "third")
        }
        for left, right in second_positions:
            label = f"{left + 1}+{right + 1}"
            duo = source_second[left] + source_second[right]
            if any(_key4(f"{duo}{suffix}") in target_keys for suffix in suffixes):
                position_hits[label] += 1
        transitions += 1

    audit_df = pd.DataFrame([
        {
            "Kedudukan 2D": f"{left + 1}+{right + 1}",
            "Hit Draw": position_hits[f"{left + 1}+{right + 1}"],
            "Draw Diuji": transitions,
            "Hit Rate %": round(
                position_hits[f"{left + 1}+{right + 1}"] / transitions * 100, 1
            ) if transitions else 0,
        }
        for left, right in second_positions
    ]).sort_values(["Hit Draw", "Kedudukan 2D"], ascending=[False, True]).reset_index(drop=True)
    best_hits = int(audit_df["Hit Draw"].max()) if not audit_df.empty else 0
    selected_positions = audit_df.loc[
        audit_df["Hit Draw"].eq(best_hits), "Kedudukan 2D"
    ].astype(str).tolist()

    current_second = _pad4(second)
    suffixes = occurrence_pairs(first, third)
    pair_groups = []
    pair_group_lookup = {}
    for left, right in second_positions:
        label = f"{left + 1}+{right + 1}"
        duo = current_second[left] + current_second[right]
        duo_key = _pair_key(duo)
        if duo_key not in pair_group_lookup:
            pair_group_lookup[duo_key] = {"2D": duo, "Kedudukan": []}
            pair_groups.append(pair_group_lookup[duo_key])
        pair_group_lookup[duo_key]["Kedudukan"].append(label)

    rows = []
    for pair_group in pair_groups:
        duo = pair_group["2D"]
        label = " / ".join(dict.fromkeys(pair_group["Kedudukan"]))
        for suffix in suffixes:
            number = f"{duo}{suffix}"
            rows.append({
                "Kedudukan 2D": label,
                "2D": duo,
                "Pair 1st+3rd": suffix,
                "No Terhasil": number,
            })
    all_df = pd.DataFrame(rows)
    selected_df = all_df[
        all_df["Kedudukan 2D"].apply(
            lambda value: any(
                position in [part.strip() for part in str(value).split("/")]
                for position in selected_positions
            )
        )
    ].copy() if not all_df.empty else all_df.copy()
    return {
        "audit": audit_df,
        "selected_positions": selected_positions,
        "suffixes": suffixes,
        "selected": selected_df,
        "all": all_df,
    }


def _occurrence_pairs_two_prizes(left_prize, right_prize):
    """Pair digit mengikut kemunculan sebenar dalam dua hadiah."""
    pool = _pad4(left_prize) + _pad4(right_prize)
    return sorted({
        "".join(sorted((pool[left], pool[right])))
        for left, right in combinations(range(len(pool)), 2)
    })


def _position_pair_groups(number):
    """Enam kedudukan 2D; pair terbalik/berulang dipaparkan sekali."""
    groups, lookup = [], {}
    number = _pad4(number)
    for left, right in combinations(range(4), 2):
        label = f"{left + 1}+{right + 1}"
        duo = number[left] + number[right]
        duo_key = _pair_key(duo)
        if duo_key not in lookup:
            lookup[duo_key] = {"2D": duo, "Kedudukan": []}
            groups.append(lookup[duo_key])
        lookup[duo_key]["Kedudukan"].append(label)
    return groups


@st.cache_data(show_spinner=False)
def build_1st_missing_digit_engine(
    history, first, second, third, bridge_v1_df, bridge_v2_df, lookback=100
):
    """1st 2D + missing + digit daripada hadiah 2nd atau 3rd."""
    frame = history.reset_index(drop=True) if history is not None else pd.DataFrame()
    start = max(0, len(frame) - int(lookback) - 1)
    stop = max(0, len(frame) - 1)
    positions = list(combinations(range(4), 2))
    position_hits, source_hits, digit_hits, joint_hits = (
        Counter(), Counter(), Counter(), Counter()
    )
    transitions = 0

    for index in range(start, stop):
        source_first = _pad4(frame.iloc[index]["first"])
        source_second = _pad4(frame.iloc[index]["second"])
        source_third = _pad4(frame.iloc[index]["third"])
        source_numbers = [source_first, source_second, source_third]
        missing_digits = sorted(set("0123456789") - set("".join(source_numbers)))
        targets = {
            _key4(frame.iloc[index + 1][column])
            for column in ("first", "second", "third")
        }
        hit_positions, hit_sources, hit_digits = set(), set(), set()
        for left, right in positions:
            position = f"{left + 1}+{right + 1}"
            duo = source_first[left] + source_first[right]
            for source_name, source_number in (
                ("Digit 2nd", source_second), ("Digit 3rd", source_third)
            ):
                for digit_index, digit in enumerate(source_number, start=1):
                    hit = any(
                        _key4(f"{duo}{missing}{digit}") in targets
                        for missing in missing_digits
                    )
                    if hit:
                        joint_hits[(position, source_name, digit_index)] += 1
                        hit_positions.add(position)
                        hit_sources.add(source_name)
                        hit_digits.add((source_name, digit_index))
        for position in hit_positions:
            position_hits[position] += 1
        for source_name in hit_sources:
            source_hits[source_name] += 1
        for source_digit in hit_digits:
            digit_hits[source_digit] += 1
        transitions += 1

    position_audit = pd.DataFrame([{
        "Kedudukan 1st 2D": f"{left + 1}+{right + 1}",
        "Hit Draw": position_hits[f"{left + 1}+{right + 1}"],
        "Draw Diuji": transitions,
        "Hit Rate %": round(
            position_hits[f"{left + 1}+{right + 1}"] / transitions * 100, 1
        ) if transitions else 0,
    } for left, right in positions]).sort_values(
        ["Hit Draw", "Kedudukan 1st 2D"], ascending=[False, True]
    ).reset_index(drop=True)
    source_audit = pd.DataFrame([{
        "Sumber Digit": source_name,
        "Hit Draw": source_hits[source_name],
        "Draw Diuji": transitions,
        "Hit Rate %": round(source_hits[source_name] / transitions * 100, 1)
        if transitions else 0,
    } for source_name in ("Digit 2nd", "Digit 3rd")]).sort_values(
        ["Hit Draw", "Sumber Digit"], ascending=[False, True]
    ).reset_index(drop=True)
    digit_audit = pd.DataFrame([{
        "Sumber Digit": source_name,
        "Kedudukan Digit": digit_index,
        "Hit Draw": digit_hits[(source_name, digit_index)],
        "Draw Diuji": transitions,
        "Hit Rate %": round(
            digit_hits[(source_name, digit_index)] / transitions * 100, 1
        ) if transitions else 0,
    } for source_name in ("Digit 2nd", "Digit 3rd")
      for digit_index in range(1, 5)]).sort_values(
        ["Hit Draw", "Sumber Digit", "Kedudukan Digit"],
        ascending=[False, True, True],
    ).reset_index(drop=True)
    joint_audit = pd.DataFrame([{
        "Kedudukan 1st 2D": f"{left + 1}+{right + 1}",
        "Sumber Digit": source_name,
        "Kedudukan Digit": digit_index,
        "Hit Draw": joint_hits[(f"{left + 1}+{right + 1}", source_name, digit_index)],
        "Draw Diuji": transitions,
        "Hit Rate %": round(
            joint_hits[(f"{left + 1}+{right + 1}", source_name, digit_index)]
            / transitions * 100, 1
        ) if transitions else 0,
    } for left, right in positions
      for source_name in ("Digit 2nd", "Digit 3rd")
      for digit_index in range(1, 5)]).sort_values(
        ["Hit Draw", "Kedudukan 1st 2D", "Sumber Digit", "Kedudukan Digit"],
        ascending=[False, True, True, True],
    ).reset_index(drop=True)

    best_position = int(position_audit["Hit Draw"].max()) if not position_audit.empty else 0
    selected_positions = position_audit.loc[
        position_audit["Hit Draw"].eq(best_position), "Kedudukan 1st 2D"
    ].astype(str).tolist()
    best_source = int(source_audit["Hit Draw"].max()) if not source_audit.empty else 0
    selected_sources = source_audit.loc[
        source_audit["Hit Draw"].eq(best_source), "Sumber Digit"
    ].astype(str).tolist()
    selected_joint = joint_audit[
        joint_audit["Kedudukan 1st 2D"].isin(selected_positions)
        & joint_audit["Sumber Digit"].isin(selected_sources)
    ]
    best_joint = int(selected_joint["Hit Draw"].max()) if not selected_joint.empty else 0
    selected_digits = selected_joint.loc[
        selected_joint["Hit Draw"].eq(best_joint),
        ["Sumber Digit", "Kedudukan Digit"],
    ].drop_duplicates().to_dict("records")

    bridge_v1 = {
        _key4(number): _pad4(number)
        for number in bridge_v1_df.get("No", pd.Series(dtype=str)).astype(str)
    } if bridge_v1_df is not None and not bridge_v1_df.empty else {}
    bridge_v2 = {
        _key4(number): _pad4(number)
        for number in bridge_v2_df.get("No", pd.Series(dtype=str)).astype(str)
    } if bridge_v2_df is not None and not bridge_v2_df.empty else {}
    current_numbers = [_pad4(first), _pad4(second), _pad4(third)]
    current_missing = sorted(set("0123456789") - set("".join(current_numbers)))
    source_lookup = {"Digit 2nd": current_numbers[1], "Digit 3rd": current_numbers[2]}
    rows = []
    for group in _position_pair_groups(first):
        labels = " / ".join(dict.fromkeys(group["Kedudukan"]))
        for source_name, source_number in source_lookup.items():
            for digit_index, digit in enumerate(source_number, start=1):
                for missing in current_missing:
                    generated = f"{group['2D']}{missing}{digit}"
                    family = _key4(generated)
                    rows.append({
                        "Kedudukan 1st 2D": labels, "1st 2D": group["2D"],
                        "Missing": missing, "Sumber Digit": source_name,
                        "Kedudukan Digit": digit_index, "Digit": digit,
                        "No Terhasil": generated,
                        "Bridge V1": bridge_v1.get(family, ""),
                        "Bridge V2": bridge_v2.get(family, ""),
                    })
    all_df = pd.DataFrame(rows)
    selected_df = all_df[
        all_df["Kedudukan 1st 2D"].apply(
            lambda value: any(
                position in [part.strip() for part in str(value).split("/")]
                for position in selected_positions
            )
        )
        & all_df.apply(
            lambda row: any(
                row["Sumber Digit"] == item["Sumber Digit"]
                and int(row["Kedudukan Digit"]) == int(item["Kedudukan Digit"])
                for item in selected_digits
            ), axis=1,
        )
    ].copy() if not all_df.empty else all_df.copy()
    return {
        "position_audit": position_audit, "source_audit": source_audit,
        "digit_audit": digit_audit, "joint_audit": joint_audit,
        "selected_positions": selected_positions,
        "selected_sources": selected_sources, "selected_digits": selected_digits,
        "missing": current_missing, "selected": selected_df, "all": all_df,
    }


@st.cache_data(show_spinner=False)
def build_1st_second_third_pair_engine(history, first, second, third, lookback=100):
    """1st 2D + pair digit yang muncul dalam hadiah 2nd dan 3rd."""
    frame = history.reset_index(drop=True) if history is not None else pd.DataFrame()
    start = max(0, len(frame) - int(lookback) - 1)
    stop = max(0, len(frame) - 1)
    positions = list(combinations(range(4), 2))
    position_hits, transitions = Counter(), 0
    for index in range(start, stop):
        source_first = _pad4(frame.iloc[index]["first"])
        source_second = _pad4(frame.iloc[index]["second"])
        source_third = _pad4(frame.iloc[index]["third"])
        suffixes = _occurrence_pairs_two_prizes(source_second, source_third)
        targets = {
            _key4(frame.iloc[index + 1][column])
            for column in ("first", "second", "third")
        }
        for left, right in positions:
            label = f"{left + 1}+{right + 1}"
            duo = source_first[left] + source_first[right]
            if any(_key4(f"{duo}{suffix}") in targets for suffix in suffixes):
                position_hits[label] += 1
        transitions += 1
    audit_df = pd.DataFrame([{
        "Kedudukan 1st 2D": f"{left + 1}+{right + 1}",
        "Hit Draw": position_hits[f"{left + 1}+{right + 1}"],
        "Draw Diuji": transitions,
        "Hit Rate %": round(
            position_hits[f"{left + 1}+{right + 1}"] / transitions * 100, 1
        ) if transitions else 0,
    } for left, right in positions]).sort_values(
        ["Hit Draw", "Kedudukan 1st 2D"], ascending=[False, True]
    ).reset_index(drop=True)
    best_hits = int(audit_df["Hit Draw"].max()) if not audit_df.empty else 0
    selected_positions = audit_df.loc[
        audit_df["Hit Draw"].eq(best_hits), "Kedudukan 1st 2D"
    ].astype(str).tolist()
    suffixes = _occurrence_pairs_two_prizes(second, third)
    rows = []
    for group in _position_pair_groups(first):
        labels = " / ".join(dict.fromkeys(group["Kedudukan"]))
        for suffix in suffixes:
            rows.append({
                "Kedudukan 1st 2D": labels, "1st 2D": group["2D"],
                "Pair 2nd+3rd": suffix,
                "No Terhasil": f"{group['2D']}{suffix}",
            })
    all_df = pd.DataFrame(rows)
    selected_df = all_df[
        all_df["Kedudukan 1st 2D"].apply(
            lambda value: any(
                position in [part.strip() for part in str(value).split("/")]
                for position in selected_positions
            )
        )
    ].copy() if not all_df.empty else all_df.copy()
    return {
        "audit": audit_df, "selected_positions": selected_positions,
        "suffixes": suffixes, "selected": selected_df, "all": all_df,
    }


@st.cache_data(show_spinner=False)
def build_1st_route_signal(history, first, second, third, lookback=100):
    """Pilih V1 atau V2 bagi laluan 1st 2D mengikut keadaan sejarah semasa."""
    frame = history.reset_index(drop=True) if history is not None else pd.DataFrame()
    start = max(0, len(frame) - int(lookback) - 1)
    stop = max(0, len(frame) - 1)

    def state_values(first_no, second_no, third_no):
        numbers = [_pad4(first_no), _pad4(second_no), _pad4(third_no)]
        return {
            "missing": 10 - len(set("".join(numbers))),
            "first_unique": len(set(numbers[0])),
            "second_unique": len(set(numbers[1])),
            "third_unique": len(set(numbers[2])),
            "repeat_prizes": sum(len(set(number)) < 4 for number in numbers),
            "first_double": int(len(set(numbers[0])) < 4),
        }

    rows = []
    for index in range(start, stop):
        source_first = _pad4(frame.iloc[index]["first"])
        source_second = _pad4(frame.iloc[index]["second"])
        source_third = _pad4(frame.iloc[index]["third"])
        numbers = [source_first, source_second, source_third]
        missing = sorted(set("0123456789") - set("".join(numbers)))
        suffixes = _occurrence_pairs_two_prizes(source_second, source_third)
        targets = {
            _key4(frame.iloc[index + 1][column])
            for column in ("first", "second", "third")
        }
        missing_hit = False
        second_third_hit = False
        for left, right in combinations(range(4), 2):
            duo = source_first[left] + source_first[right]
            missing_hit |= any(
                _key4(f"{duo}{digit}{source_digit}") in targets
                for digit in missing for source_digit in source_third
            )
            second_third_hit |= any(
                _key4(f"{duo}{suffix}") in targets for suffix in suffixes
            )
        row = state_values(*numbers)
        row.update({"missing_hit": missing_hit, "second_third_hit": second_third_hit})
        rows.append(row)

    current = state_values(first, second, third)
    groups = [
        ("missing", "first_unique"), ("missing", "repeat_prizes"),
        ("missing", "first_double"),
        ("missing", "second_unique", "third_unique"),
    ]
    votes, evidence = Counter(), []
    for features in groups:
        matches = [
            row for row in rows
            if all(row[feature] == current[feature] for feature in features)
        ]
        if len(matches) < 3:
            continue
        v1_rate = sum(row["missing_hit"] for row in matches) / len(matches)
        v2_rate = sum(row["second_third_hit"] for row in matches) / len(matches)
        vote = "1st 2D + Missing" if v1_rate > v2_rate else (
            "1st 2D + 2nd & 3rd" if v2_rate > v1_rate else "Seimbang"
        )
        votes[vote] += 1
        evidence.append({
            "Keadaan": " + ".join(features), "Padanan": len(matches),
            "1st Missing %": round(v1_rate * 100, 1),
            "1st 2nd&3rd %": round(v2_rate * 100, 1), "Signal": vote,
        })
    decisive = {name: count for name, count in votes.items() if name != "Seimbang"}
    if not decisive:
        signal, support = "Seimbang", 0
    else:
        best = max(decisive.values())
        leaders = [name for name, count in decisive.items() if count == best]
        signal = leaders[0] if len(leaders) == 1 else "Seimbang"
        support = best if len(leaders) == 1 else 0
    return {
        "signal": signal, "support": support,
        "tested_states": len(evidence), "evidence": pd.DataFrame(evidence),
    }


@st.cache_data(show_spinner=False)
def build_3rd_missing_first_digit_engine(
    history, first, second, third, bridge_v1_df, bridge_v2_df, lookback=100
):
    """2D hadiah ketiga + missing + satu digit hadiah pertama."""
    frame = history.reset_index(drop=True) if history is not None else pd.DataFrame()
    start = max(0, len(frame) - int(lookback) - 1)
    stop = max(0, len(frame) - 1)
    position_hits, first_hits, joint_hits = Counter(), Counter(), Counter()
    transitions = 0
    positions = list(combinations(range(4), 2))

    for index in range(start, stop):
        source_first = _pad4(frame.iloc[index]["first"])
        source_third = _pad4(frame.iloc[index]["third"])
        source_numbers = [
            _pad4(frame.iloc[index][column])
            for column in ("first", "second", "third")
        ]
        missing_digits = sorted(set("0123456789") - set("".join(source_numbers)))
        targets = {
            _key4(frame.iloc[index + 1][column])
            for column in ("first", "second", "third")
        }
        hit_positions, hit_first = set(), set()
        for left, right in positions:
            label = f"{left + 1}+{right + 1}"
            duo = source_third[left] + source_third[right]
            for first_index, first_digit in enumerate(source_first, start=1):
                hit = any(
                    _key4(f"{duo}{missing}{first_digit}") in targets
                    for missing in missing_digits
                )
                if hit:
                    joint_hits[(label, first_index)] += 1
                    hit_positions.add(label)
                    hit_first.add(first_index)
        for label in hit_positions:
            position_hits[label] += 1
        for first_index in hit_first:
            first_hits[first_index] += 1
        transitions += 1

    position_audit = pd.DataFrame([{
        "Kedudukan 3rd 2D": f"{left + 1}+{right + 1}",
        "Hit Draw": position_hits[f"{left + 1}+{right + 1}"],
        "Draw Diuji": transitions,
        "Hit Rate %": round(position_hits[f"{left + 1}+{right + 1}"] / transitions * 100, 1)
        if transitions else 0,
    } for left, right in positions]).sort_values(
        ["Hit Draw", "Kedudukan 3rd 2D"], ascending=[False, True]
    ).reset_index(drop=True)
    first_audit = pd.DataFrame([{
        "Kedudukan Digit 1st": position,
        "Hit Draw": first_hits[position],
        "Draw Diuji": transitions,
        "Hit Rate %": round(first_hits[position] / transitions * 100, 1)
        if transitions else 0,
    } for position in range(1, 5)]).sort_values(
        ["Hit Draw", "Kedudukan Digit 1st"], ascending=[False, True]
    ).reset_index(drop=True)
    joint_audit = pd.DataFrame([{
        "Kedudukan 3rd 2D": f"{left + 1}+{right + 1}",
        "Kedudukan Digit 1st": first_position,
        "Hit Draw": joint_hits[(f"{left + 1}+{right + 1}", first_position)],
        "Draw Diuji": transitions,
        "Hit Rate %": round(
            joint_hits[(f"{left + 1}+{right + 1}", first_position)] / transitions * 100, 1
        ) if transitions else 0,
    } for left, right in positions for first_position in range(1, 5)]).sort_values(
        ["Hit Draw", "Kedudukan 3rd 2D", "Kedudukan Digit 1st"],
        ascending=[False, True, True],
    ).reset_index(drop=True)

    best_position = int(position_audit["Hit Draw"].max()) if not position_audit.empty else 0
    best_first = int(first_audit["Hit Draw"].max()) if not first_audit.empty else 0
    selected_positions = position_audit.loc[
        position_audit["Hit Draw"].eq(best_position), "Kedudukan 3rd 2D"
    ].astype(str).tolist()
    selected_first = first_audit.loc[
        first_audit["Hit Draw"].eq(best_first), "Kedudukan Digit 1st"
    ].astype(int).tolist()

    current_first, current_third = _pad4(first), _pad4(third)
    current_missing = sorted(
        set("0123456789") - set(_pad4(first) + _pad4(second) + current_third)
    )
    v1 = {
        _key4(number): _pad4(number)
        for number in bridge_v1_df.get("No", pd.Series(dtype=str)).astype(str)
    } if bridge_v1_df is not None and not bridge_v1_df.empty else {}
    v2 = {
        _key4(number): _pad4(number)
        for number in bridge_v2_df.get("No", pd.Series(dtype=str)).astype(str)
    } if bridge_v2_df is not None and not bridge_v2_df.empty else {}
    rows = []
    for group in _position_pair_groups(current_third):
        label = " / ".join(group["Kedudukan"])
        for first_index, first_digit in enumerate(current_first, start=1):
            for missing in current_missing:
                generated = f"{group['2D']}{missing}{first_digit}"
                key = _key4(generated)
                rows.append({
                    "Kedudukan 3rd 2D": label, "2D": group["2D"],
                    "Missing": missing, "Kedudukan Digit 1st": first_index,
                    "Digit 1st": first_digit, "No Terhasil": generated,
                    "Bridge V1": v1.get(key, ""), "Bridge V2": v2.get(key, ""),
                })
    all_df = pd.DataFrame(rows)
    selected_df = all_df[
        all_df["Kedudukan 3rd 2D"].apply(
            lambda value: any(
                position in [part.strip() for part in str(value).split("/")]
                for position in selected_positions
            )
        ) & all_df["Kedudukan Digit 1st"].isin(selected_first)
    ].copy() if not all_df.empty else all_df.copy()
    return {
        "position_audit": position_audit, "first_audit": first_audit,
        "joint_audit": joint_audit, "selected_positions": selected_positions,
        "selected_first": selected_first, "missing": current_missing,
        "selected": selected_df, "all": all_df,
    }


@st.cache_data(show_spinner=False)
def build_3rd_first_second_pair_engine(history, first, second, third, lookback=100):
    """2D hadiah ketiga + dua digit daripada hadiah pertama dan kedua."""
    frame = history.reset_index(drop=True) if history is not None else pd.DataFrame()
    start = max(0, len(frame) - int(lookback) - 1)
    stop = max(0, len(frame) - 1)
    position_hits, transitions = Counter(), 0
    positions = list(combinations(range(4), 2))
    for index in range(start, stop):
        source_first = _pad4(frame.iloc[index]["first"])
        source_second = _pad4(frame.iloc[index]["second"])
        source_third = _pad4(frame.iloc[index]["third"])
        suffixes = _occurrence_pairs_two_prizes(source_first, source_second)
        targets = {
            _key4(frame.iloc[index + 1][column])
            for column in ("first", "second", "third")
        }
        for left, right in positions:
            label = f"{left + 1}+{right + 1}"
            duo = source_third[left] + source_third[right]
            if any(_key4(f"{duo}{suffix}") in targets for suffix in suffixes):
                position_hits[label] += 1
        transitions += 1
    audit_df = pd.DataFrame([{
        "Kedudukan 3rd 2D": f"{left + 1}+{right + 1}",
        "Hit Draw": position_hits[f"{left + 1}+{right + 1}"],
        "Draw Diuji": transitions,
        "Hit Rate %": round(position_hits[f"{left + 1}+{right + 1}"] / transitions * 100, 1)
        if transitions else 0,
    } for left, right in positions]).sort_values(
        ["Hit Draw", "Kedudukan 3rd 2D"], ascending=[False, True]
    ).reset_index(drop=True)
    best_hits = int(audit_df["Hit Draw"].max()) if not audit_df.empty else 0
    selected_positions = audit_df.loc[
        audit_df["Hit Draw"].eq(best_hits), "Kedudukan 3rd 2D"
    ].astype(str).tolist()
    suffixes = _occurrence_pairs_two_prizes(first, second)
    rows = []
    for group in _position_pair_groups(third):
        label = " / ".join(group["Kedudukan"])
        for suffix in suffixes:
            rows.append({
                "Kedudukan 3rd 2D": label, "2D": group["2D"],
                "Pair 1st+2nd": suffix,
                "No Terhasil": f"{group['2D']}{suffix}",
            })
    all_df = pd.DataFrame(rows)
    selected_df = all_df[
        all_df["Kedudukan 3rd 2D"].apply(
            lambda value: any(
                position in [part.strip() for part in str(value).split("/")]
                for position in selected_positions
            )
        )
    ].copy() if not all_df.empty else all_df.copy()
    return {
        "audit": audit_df, "selected_positions": selected_positions,
        "suffixes": suffixes, "selected": selected_df, "all": all_df,
    }


@st.cache_data(show_spinner=False)
def build_4_route_signal(history, first, second, third, lookback=100):
    """Pilih satu daripada empat laluan 2D tanpa menggabungkan outputnya."""
    frame = history.reset_index(drop=True) if history is not None else pd.DataFrame()
    start = max(0, len(frame) - int(lookback) - 1)
    stop = max(0, len(frame) - 1)
    route_names = (
        "2nd 2D + Missing",
        "2nd 2D + 1st & 3rd",
        "3rd 2D + Missing",
        "3rd 2D + 1st & 2nd",
    )

    def state_values(first_no, second_no, third_no):
        numbers = [_pad4(first_no), _pad4(second_no), _pad4(third_no)]
        return {
            "missing": 10 - len(set("".join(numbers))),
            "first_unique": len(set(numbers[0])),
            "second_unique": len(set(numbers[1])),
            "third_unique": len(set(numbers[2])),
            "repeat_prizes": sum(len(set(number)) < 4 for number in numbers),
            "second_double": int(len(set(numbers[1])) < 4),
            "third_double": int(len(set(numbers[2])) < 4),
        }

    rows = []
    for index in range(start, stop):
        source_first = _pad4(frame.iloc[index]["first"])
        source_second = _pad4(frame.iloc[index]["second"])
        source_third = _pad4(frame.iloc[index]["third"])
        numbers = [source_first, source_second, source_third]
        targets = {
            _key4(frame.iloc[index + 1][column])
            for column in ("first", "second", "third")
        }
        missing = sorted(set("0123456789") - set("".join(numbers)))
        suffix_13 = _occurrence_pairs_two_prizes(source_first, source_third)
        suffix_12 = _occurrence_pairs_two_prizes(source_first, source_second)
        route_hits = {name: False for name in route_names}
        for left, right in combinations(range(4), 2):
            duo_second = source_second[left] + source_second[right]
            duo_third = source_third[left] + source_third[right]
            route_hits[route_names[0]] |= any(
                _key4(f"{duo_second}{digit}{first_digit}") in targets
                for digit in missing for first_digit in source_first
            )
            route_hits[route_names[1]] |= any(
                _key4(f"{duo_second}{suffix}") in targets for suffix in suffix_13
            )
            route_hits[route_names[2]] |= any(
                _key4(f"{duo_third}{digit}{first_digit}") in targets
                for digit in missing for first_digit in source_first
            )
            route_hits[route_names[3]] |= any(
                _key4(f"{duo_third}{suffix}") in targets for suffix in suffix_12
            )
        row = state_values(*numbers)
        row.update(route_hits)
        rows.append(row)

    current = state_values(first, second, third)
    state_groups = [
        ("missing", "repeat_prizes"),
        ("missing", "second_unique", "third_unique"),
        ("missing", "first_unique"),
        ("missing", "second_double", "third_double"),
    ]
    votes, evidence = Counter(), []
    for features in state_groups:
        matches = [
            row for row in rows
            if all(row[feature] == current[feature] for feature in features)
        ]
        if len(matches) < 3:
            continue
        rates = {
            name: sum(bool(row[name]) for row in matches) / len(matches)
            for name in route_names
        }
        best_rate = max(rates.values()) if rates else 0
        leaders = [name for name, rate in rates.items() if rate == best_rate]
        vote = leaders[0] if len(leaders) == 1 else "Seimbang"
        votes[vote] += 1
        evidence.append({
            "Keadaan": " + ".join(features), "Padanan": len(matches),
            **{f"{name} %": round(rate * 100, 1) for name, rate in rates.items()},
            "Signal": vote,
        })
    decisive = {name: count for name, count in votes.items() if name != "Seimbang"}
    if not decisive:
        signal, support = "Seimbang", 0
    else:
        best = max(decisive.values())
        leaders = [name for name, count in decisive.items() if count == best]
        signal = leaders[0] if len(leaders) == 1 else "Seimbang"
        support = best if len(leaders) == 1 else 0
    return {
        "signal": signal, "support": support,
        "tested_states": len(evidence), "evidence": pd.DataFrame(evidence),
    }


@st.cache_data(show_spinner=False)
def build_3rd_route_signal(history, first, second, third, lookback=100):
    """Pilih antara dua laluan 3rd sahaja; laluan 2nd tidak disentuh."""
    frame = history.reset_index(drop=True) if history is not None else pd.DataFrame()
    start = max(0, len(frame) - int(lookback) - 1)
    stop = max(0, len(frame) - 1)
    rows = []

    def state_values(first_no, second_no, third_no):
        numbers = [_pad4(first_no), _pad4(second_no), _pad4(third_no)]
        return {
            "missing": 10 - len(set("".join(numbers))),
            "first_unique": len(set(numbers[0])),
            "second_unique": len(set(numbers[1])),
            "third_unique": len(set(numbers[2])),
            "repeat_prizes": sum(len(set(number)) < 4 for number in numbers),
            "third_double": int(len(set(numbers[2])) < 4),
        }

    for index in range(start, stop):
        source_first = _pad4(frame.iloc[index]["first"])
        source_second = _pad4(frame.iloc[index]["second"])
        source_third = _pad4(frame.iloc[index]["third"])
        numbers = [source_first, source_second, source_third]
        targets = {
            _key4(frame.iloc[index + 1][column])
            for column in ("first", "second", "third")
        }
        missing = sorted(set("0123456789") - set("".join(numbers)))
        suffixes = _occurrence_pairs_two_prizes(source_first, source_second)
        missing_hit = False
        first_second_hit = False
        for left, right in combinations(range(4), 2):
            duo = source_third[left] + source_third[right]
            missing_hit |= any(
                _key4(f"{duo}{digit}{first_digit}") in targets
                for digit in missing for first_digit in source_first
            )
            first_second_hit |= any(
                _key4(f"{duo}{suffix}") in targets for suffix in suffixes
            )
        row = state_values(*numbers)
        row.update({"missing_hit": missing_hit, "first_second_hit": first_second_hit})
        rows.append(row)

    current = state_values(first, second, third)
    groups = [
        ("missing", "third_unique"),
        ("missing", "repeat_prizes"),
        ("missing", "third_double"),
        ("missing", "first_unique", "second_unique"),
    ]
    votes, evidence = Counter(), []
    for features in groups:
        matches = [
            row for row in rows
            if all(row[feature] == current[feature] for feature in features)
        ]
        if len(matches) < 3:
            continue
        missing_rate = sum(row["missing_hit"] for row in matches) / len(matches)
        first_second_rate = sum(row["first_second_hit"] for row in matches) / len(matches)
        if missing_rate > first_second_rate:
            vote = "3rd 2D + Missing"
        elif first_second_rate > missing_rate:
            vote = "3rd 2D + 1st & 2nd"
        else:
            vote = "Seimbang"
        votes[vote] += 1
        evidence.append({
            "Keadaan": " + ".join(features), "Padanan": len(matches),
            "3rd 2D + Missing %": round(missing_rate * 100, 1),
            "3rd 2D + 1st & 2nd %": round(first_second_rate * 100, 1),
            "Signal": vote,
        })
    decisive = {name: count for name, count in votes.items() if name != "Seimbang"}
    if not decisive:
        signal, support = "Seimbang", 0
    else:
        best = max(decisive.values())
        leaders = [name for name, count in decisive.items() if count == best]
        signal = leaders[0] if len(leaders) == 1 else "Seimbang"
        support = best if len(leaders) == 1 else 0
    return {
        "signal": signal, "support": support,
        "tested_states": len(evidence), "evidence": pd.DataFrame(evidence),
    }


@st.cache_data(show_spinner=False)
def build_first_third_extended_audit(history, first, second, third, lookback=100):
    """Audit digit, pair pelengkap dan gabungannya dengan kedudukan 2D."""
    frame = history.reset_index(drop=True) if history is not None else pd.DataFrame()
    start = max(0, len(frame) - int(lookback) - 1)
    stop = max(0, len(frame) - 1)
    second_positions = list(combinations(range(4), 2))

    def occurrence_pairs(first_no, third_no):
        pool = _pad4(first_no) + _pad4(third_no)
        return sorted({
            "".join(sorted((pool[left], pool[right])))
            for left, right in combinations(range(len(pool)), 2)
        })

    records = []
    for index in range(start, stop):
        source_first = _pad4(frame.iloc[index]["first"])
        source_second = _pad4(frame.iloc[index]["second"])
        source_third = _pad4(frame.iloc[index]["third"])
        pool = source_first + source_third
        suffixes = occurrence_pairs(source_first, source_third)
        target_keys = {
            _key4(frame.iloc[index + 1][column])
            for column in ("first", "second", "third")
        }
        hit_positions, hit_pairs, hit_digits, hit_joint = set(), set(), set(), set()
        for left, right in second_positions:
            position = f"{left + 1}+{right + 1}"
            duo = source_second[left] + source_second[right]
            for suffix in suffixes:
                if _key4(f"{duo}{suffix}") not in target_keys:
                    continue
                hit_positions.add(position)
                hit_pairs.add(suffix)
                hit_digits.update(suffix)
                hit_joint.add((position, suffix))
        records.append({
            "Available Digits": set(pool),
            "Available Pairs": set(suffixes),
            "Hit Positions": hit_positions,
            "Hit Pairs": hit_pairs,
            "Hit Digits": hit_digits,
            "Hit Joint": hit_joint,
        })

    current_digits = sorted(set(_pad4(first) + _pad4(third)))
    current_pairs = occurrence_pairs(first, third)
    windows = [10, 20, 30, 50, 100]

    engine_position_audit = build_2d_first_third_pair_engine(
        history, first, second, third, lookback=lookback
    )["audit"]
    best_position_hits = (
        int(engine_position_audit["Hit Draw"].max())
        if not engine_position_audit.empty else 0
    )
    selected_positions = engine_position_audit.loc[
        engine_position_audit["Hit Draw"].eq(best_position_hits), "Kedudukan 2D"
    ].astype(str).tolist()

    def recent_count(predicate, window):
        subset = records[-min(window, len(records)):] if records else []
        return sum(1 for record in subset if predicate(record))

    digit_rows = []
    for digit in current_digits:
        exposure = sum(digit in record["Available Digits"] for record in records)
        hit_all = sum(digit in record["Hit Digits"] for record in records)
        hit_selected = sum(
            any(
                position in selected_positions and digit in suffix
                for position, suffix in record["Hit Joint"]
            )
            for record in records
        )
        row = {
            "Digit": digit,
            "Exposure": exposure,
            "Hit Semua 2D": hit_all,
            "Rate Semua %": round(hit_all / exposure * 100, 1) if exposure else 0,
            "Hit Kedudukan Utama": hit_selected,
            "Rate Utama %": round(hit_selected / exposure * 100, 1) if exposure else 0,
        }
        for window in windows:
            row[f"Hit {window}"] = recent_count(
                lambda record, d=digit: d in record["Hit Digits"], window
            )
        digit_rows.append(row)
    digit_audit = pd.DataFrame(digit_rows).sort_values(
        ["Hit Kedudukan Utama", "Hit Semua 2D", "Rate Semua %", "Digit"],
        ascending=[False, False, False, True],
    ).reset_index(drop=True)

    pair_rows = []
    for suffix in current_pairs:
        exposure = sum(suffix in record["Available Pairs"] for record in records)
        hit_all = sum(suffix in record["Hit Pairs"] for record in records)
        hit_selected = sum(
            any(
                position in selected_positions and hit_suffix == suffix
                for position, hit_suffix in record["Hit Joint"]
            )
            for record in records
        )
        row = {
            "Pair Digit": suffix,
            "Exposure": exposure,
            "Hit Semua 2D": hit_all,
            "Rate Semua %": round(hit_all / exposure * 100, 1) if exposure else 0,
            "Hit Kedudukan Utama": hit_selected,
            "Rate Utama %": round(hit_selected / exposure * 100, 1) if exposure else 0,
        }
        for window in windows:
            row[f"Hit {window}"] = recent_count(
                lambda record, p=suffix: p in record["Hit Pairs"], window
            )
        pair_rows.append(row)
    pair_audit = pd.DataFrame(pair_rows).sort_values(
        ["Hit Kedudukan Utama", "Hit Semua 2D", "Rate Semua %", "Pair Digit"],
        ascending=[False, False, False, True],
    ).reset_index(drop=True)

    joint_rows = []
    current_second = _pad4(second)
    for left, right in second_positions:
        position = f"{left + 1}+{right + 1}"
        current_duo = current_second[left] + current_second[right]
        for suffix in current_pairs:
            exposure = sum(suffix in record["Available Pairs"] for record in records)
            hits = sum((position, suffix) in record["Hit Joint"] for record in records)
            joint_rows.append({
                "Kedudukan 2D": position,
                "Pair Digit": suffix,
                "Full No": f"{current_duo}{suffix}",
                "Exposure": exposure,
                "Hit Draw": hits,
                "Hit Rate %": round(hits / exposure * 100, 1) if exposure else 0,
            })
    joint_audit = pd.DataFrame(joint_rows).sort_values(
        ["Hit Draw", "Hit Rate %", "Kedudukan 2D", "Pair Digit"],
        ascending=[False, False, True, True],
    ).reset_index(drop=True)

    def leaders(frame, hit_column, label_column):
        if frame.empty:
            return []
        best = int(frame[hit_column].max())
        return frame.loc[frame[hit_column].eq(best), label_column].astype(str).tolist()

    return {
        "selected_positions": selected_positions,
        "digit_audit": digit_audit,
        "pair_audit": pair_audit,
        "joint_audit": joint_audit,
        "top_digit_all": leaders(digit_audit, "Hit Semua 2D", "Digit"),
        "top_digit_selected": leaders(digit_audit, "Hit Kedudukan Utama", "Digit"),
        "top_pair_all": leaders(pair_audit, "Hit Semua 2D", "Pair Digit"),
        "top_pair_selected": leaders(pair_audit, "Hit Kedudukan Utama", "Pair Digit"),
    }







def copy_button_clean(label, value, key_name):
    js_value = json.dumps(str(value))
    components.html(
        f"""
        <button onclick='navigator.clipboard.writeText({js_value}).then(() => {{
            const msg = document.getElementById("msg_{key_name}");
            msg.innerText = "Disalin";
            setTimeout(() => msg.innerText = "", 1600);
        }}).catch(() => {{
            const msg = document.getElementById("msg_{key_name}");
            msg.innerText = "Copy gagal. Sila salin manual dari kotak.";
        }});'
        style="border:0;border-radius:10px;background:#3157e5;color:white;padding:9px 15px;font-size:14px;font-weight:750;margin-right:8px;box-shadow:0 5px 14px rgba(49,87,229,.18);">
            {label}
        </button>
        <span id="msg_{key_name}" style="color:#15803d;font-size:14px;font-weight:700;margin-left:8px;"></span>
        """,
        height=48
    )

st.set_page_config(page_title="Rumah A Predictor", page_icon="🎯", layout="wide")

st.markdown('\n<style>\na[href^="#"] {\n    display: none !important;\n}\n.block-container {\n    padding-top: 1.2rem !important;\n}\nh1, h2, h3 {\n    letter-spacing: -0.02em;\n}\ndiv[data-testid="stRadio"] {\n    margin-top: 0.25rem;\n    margin-bottom: 1.25rem;\n}\n</style>\n', unsafe_allow_html=True)


st.markdown("""
<style>
.block-container {
    padding-top: 1.3rem;
    padding-bottom: 1rem;
}
h1, h2, h3 {
    margin-top: 0.45rem;
    margin-bottom: 0.45rem;
}
div[data-testid="stDataFrame"] {
    margin-bottom: 0.75rem;
}
.small-note {
    color: #666;
    font-size: 0.92rem;
}
.copy-box {
    border: 1px solid #e6e6e6;
    border-radius: 12px;
    padding: 12px 14px;
    background: #fffdf7;
    margin-top: 8px;
    margin-bottom: 12px;
    font-size: 1.05rem;
}
.pick-card {
    border: 1px solid #e6e6e6;
    border-radius: 14px;
    padding: 12px;
    text-align: center;
    background: #ffffff;
    margin-bottom: 8px;
}
.pick-no {
    font-size: 32px;
    font-weight: 850;
    letter-spacing: 2px;
}
</style>
""", unsafe_allow_html=True)


st.markdown(
    """
    <div class="rap-hero">
        <div class="rap-brand-mark">R</div>
        <div>
            <div class="rap-title">Rumah A Predictor</div>
            <div class="rap-subtitle">Number Pattern Analysis Engine</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

main_menu = "Home"  # Top menu removed for cleaner UI



if main_menu == "Analysis":
    st.subheader("📊 Analysis")
    st.caption("Analisis teknikal diletakkan di sini supaya Home lebih kemas.")

    try:
        ana_c1, ana_c2 = st.columns(2)
        with ana_c1:
            hot_window_analysis = st.selectbox("Hot Digit Window", [10, 30, 50, 100], index=1, key="analysis_hot_window")
            hot_df_analysis = hot_digit_analysis(st.session_state.history, window=hot_window_analysis)
            st.write(f"Hot Digits - {hot_window_analysis} draw terakhir")
            st.dataframe(hot_df_analysis, hide_index=True, use_container_width=True)

        with ana_c2:
            cold_window_analysis = st.selectbox("Cold Digit Window", [10, 30, 50, 100], index=3, key="analysis_cold_window")
            cold_df_analysis = cold_digit_analysis(st.session_state.history, window=cold_window_analysis)
            st.write(f"Cold Digits - {cold_window_analysis} draw terakhir")
            st.dataframe(cold_df_analysis, hide_index=True, use_container_width=True)

        st.info("Hybrid ranking, Score Breakdown dan audit penuh masih boleh dilihat selepas Generate di bahagian Advanced Audit.")
    except Exception:
        st.warning("Analisis belum dapat dipaparkan.")

    st.stop()

if main_menu == "History":
    st.subheader("📜 History")
    st.caption("Paparan 10 draw terakhir daripada data aplikasi.")
    try:
        hist_view = history.copy()
        hist_view["draw_no"] = hist_view["draw_no"].astype(str).str.zfill(6)
        hist_view["draw_date"] = hist_view["draw_date"].astype(str)
        hist_view["first"] = hist_view["first"].astype(str).str.zfill(4)
        hist_view["second"] = hist_view["second"].astype(str).str.zfill(4)
        hist_view["third"] = hist_view["third"].astype(str).str.zfill(4)
        hist_view = hist_view.sort_values("draw_no", ascending=False).head(10)
        hist_view = hist_view.rename(columns={
            "draw_no": "Draw No",
            "draw_date": "Draw Date",
            "first": "1st",
            "second": "2nd",
            "third": "3rd"
        })
        st.dataframe(hist_view[["Draw No", "Draw Date", "1st", "2nd", "3rd"]], hide_index=True, use_container_width=True)
    except Exception as e:
        st.warning("History belum dapat dipaparkan.")
    st.stop()

if main_menu == "Settings":
    st.subheader("⚙️ Settings")
    st.info("Versi ini menggunakan tetapan ringkas untuk APK WebView. Tetapan lanjutan boleh ditambah selepas APK pertama berjaya.")
    st.write("**App Name:** Rumah A Predictor")
    st.write("**Mode:** APK Preparation")
    st.write("**Data Source:** TotoHistoryAll.xlsx")
    st.write("**Auto-save GitHub:** Ikut status Streamlit Secrets")
    st.stop()

if main_menu == "About":
    st.subheader("ℹ️ About")
    st.markdown("""
**Rumah A Predictor** ialah aplikasi paparan analisis dan pemilihan nombor berasaskan data sejarah.

Fokus semasa:
- Paparan mudah untuk telefon
- Bridge V1 dan Bridge V2
- Bridge Pair Shortlist
- Bridge Dua Pair
- Carta 3D V2
- Backtest Bridge
- Sedia untuk dibungkus sebagai Android WebView APK

Nota: Aplikasi ini hanyalah alat analisis data dan tidak menjamin sebarang keputusan.
""")
    st.stop()




st.markdown("""
<style>
/* V20 mobile-ready UI */
.block-container {
    padding-top: 1.2rem;
    padding-bottom: 2rem;
}
[data-testid="stMetricValue"] {
    font-size: 1.8rem;
}
div[data-testid="stDataFrame"] {
    font-size: 0.92rem;
}
@media (max-width: 768px) {
    .block-container {
        padding-left: 0.7rem;
        padding-right: 0.7rem;
    }
    h1 {
        font-size: 2rem !important;
    }
    h2, h3 {
        font-size: 1.35rem !important;
    }
    div[data-testid="stDataFrame"] {
        font-size: 0.82rem;
    }
    button[kind="secondary"] {
        width: 100%;
    }
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
/* V31.37 — Clean Analysis Dashboard */
:root {
    --rap-ink: #172033;
    --rap-muted: #687386;
    --rap-line: #E2E7F0;
    --rap-surface: #FFFFFF;
    --rap-blue: #3157E5;
    --rap-violet: #7656D8;
    --rap-amber: #D98B18;
    --rap-teal: #0F9488;
}
[data-testid="stAppViewContainer"] {
    background:
        radial-gradient(circle at 86% 0%, rgba(49,87,229,.07), transparent 25rem),
        #F6F8FC;
}
.block-container {
    max-width: 1180px;
    padding-top: 1.55rem !important;
    padding-bottom: 4rem !important;
}
.rap-hero {
    display: flex;
    align-items: center;
    gap: 14px;
    padding: 8px 2px 17px;
    border-bottom: 1px solid var(--rap-line);
    margin-bottom: 15px;
}
.rap-brand-mark {
    width: 46px;
    height: 46px;
    display: flex;
    align-items: center;
    justify-content: center;
    border-radius: 14px;
    color: white;
    font-size: 22px;
    font-weight: 850;
    background: linear-gradient(145deg, #3157E5, #6A50D8);
    box-shadow: 0 9px 22px rgba(49,87,229,.22);
}
.rap-title {
    color: var(--rap-ink);
    font-size: 25px;
    line-height: 1.1;
    font-weight: 820;
    letter-spacing: -.035em;
}
.rap-subtitle {
    color: var(--rap-muted);
    font-size: 13px;
    margin-top: 5px;
    letter-spacing: .025em;
}
.rap-status-row {
    display: flex;
    gap: 8px;
    flex-wrap: wrap;
    margin: 3px 0 22px;
}
.rap-badge {
    display: inline-flex;
    align-items: center;
    gap: 7px;
    padding: 7px 11px;
    border-radius: 999px;
    color: #4A5568;
    background: rgba(255,255,255,.9);
    border: 1px solid var(--rap-line);
    font-size: 12px;
    font-weight: 700;
}
.rap-dot {
    width: 7px;
    height: 7px;
    border-radius: 50%;
    background: #22A06B;
    box-shadow: 0 0 0 3px rgba(34,160,107,.12);
}
.rap-section-kicker {
    color: #8490A3;
    font-size: 11px;
    font-weight: 800;
    letter-spacing: .14em;
    text-transform: uppercase;
    margin: 23px 0 8px;
}
.rap-panel-title {
    color: var(--rap-ink);
    font-size: 21px;
    font-weight: 800;
    letter-spacing: -.025em;
    margin: 2px 0 3px;
}
.engine-head {
    display: flex;
    align-items: center;
    gap: 11px;
    margin: 27px 0 7px;
    padding: 12px 15px;
    border-radius: 13px;
    border: 1px solid var(--rap-line);
    background: rgba(255,255,255,.9);
    font-size: 19px;
    font-weight: 800;
    letter-spacing: -.02em;
}
.engine-head::before {
    content: "";
    width: 5px;
    height: 25px;
    border-radius: 99px;
    background: var(--engine-color);
}
.engine-v1 { --engine-color: var(--rap-blue); }
.engine-v2 { --engine-color: var(--rap-violet); }
.engine-pair { --engine-color: var(--rap-amber); }
.engine-board { --engine-color: #3478a4; }
.engine-support { --engine-color: #D06C73; }
.engine-chart { --engine-color: var(--rap-teal); }
.engine-signal { --engine-color: #0B8F77; }
div[data-testid="stMetric"] {
    background: rgba(255,255,255,.94);
    border: 1px solid var(--rap-line);
    border-radius: 14px;
    padding: 14px 16px;
    box-shadow: 0 7px 22px rgba(31,45,74,.045);
}
[data-testid="stMetricLabel"] {
    color: var(--rap-muted);
    font-weight: 700;
}
[data-testid="stMetricValue"] {
    color: var(--rap-ink);
    font-weight: 800;
    letter-spacing: .02em;
}
[data-testid="stForm"] {
    background: rgba(255,255,255,.96);
    border: 1px solid var(--rap-line);
    border-radius: 16px;
    padding: 18px 20px 20px;
    box-shadow: 0 10px 28px rgba(31,45,74,.055);
}
[data-testid="stFormSubmitButton"] button {
    width: 100%;
    min-height: 44px;
    border-radius: 11px;
    border: 0;
    font-weight: 800;
    color: #FFFFFF !important;
    background: linear-gradient(100deg, #0F9F83, #087A73);
    box-shadow: 0 8px 19px rgba(8,122,115,.24);
}
[data-testid="stFormSubmitButton"] button p {
    color: #FFFFFF !important;
}
[data-testid="stFormSubmitButton"] button:hover {
    color: #FFFFFF !important;
    background: linear-gradient(100deg, #0B8F77, #066B66);
    box-shadow: 0 10px 23px rgba(8,122,115,.3);
}
[data-testid="stExpander"] {
    background: rgba(255,255,255,.9);
    border: 1px solid var(--rap-line);
    border-radius: 13px;
    box-shadow: 0 4px 15px rgba(31,45,74,.025);
    overflow: hidden;
}
div[data-testid="stDataFrame"] {
    border-radius: 12px;
    overflow: hidden;
}
button[kind="secondary"], button[kind="primary"] {
    border-radius: 10px !important;
    font-weight: 750 !important;
}
hr {
    border-color: var(--rap-line) !important;
}
@media (max-width: 768px) {
    .block-container {
        padding-left: .8rem !important;
        padding-right: .8rem !important;
    }
    .rap-title { font-size: 21px; }
    .rap-brand-mark { width: 42px; height: 42px; }
    .engine-head { font-size: 17px; }
}
</style>
""", unsafe_allow_html=True)

DATA_FILE = Path("TotoHistoryAll.xlsx")
GITHUB_OWNER = "wazley-hub"
GITHUB_REPO = "rumah-a-predictor-v9"
GITHUB_BRANCH = "main"
GITHUB_FILE_PATH = "TotoHistoryAll.xlsx"

def pad4(x):
    try:
        if pd.isna(x):
            return "0000"
    except Exception:
        pass
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.zfill(4)[-4:]

def get_pairs(nums):
    pairs = []
    for n in nums:
        pairs.extend([n[0:2], n[1:3], n[2:4]])
    return pairs

def max_repeat(n):
    return max(Counter(n).values())

def score_add(d, num, score, allow_triple_digits=None):
    if len(num) != 4:
        return
    allow_triple_digits = set(allow_triple_digits or [])
    mr = max_repeat(num)

    if mr <= 2:
        d[num] += score
        return

    # V31: triple terkawal.
    # Triple hanya dibenarkan jika digit yang berulang memang kuat dalam latest full result/top3 support.
    if mr == 3:
        counts = Counter(num)
        triple_digit = None
        for digit, cnt in counts.items():
            if cnt == 3:
                triple_digit = digit
                break
        if triple_digit in allow_triple_digits:
            d[num] += score * 0.82

def add_perm4(d, a, b, c, e, score, allow_triple_digits=None):
    combos = [
        (a,b,c,e,1.00), (a,b,e,c,0.96), (a,c,b,e,0.93),
        (a,c,e,b,0.90), (b,a,c,e,0.88), (c,a,b,e,0.86),
        (e,c,b,a,0.82),
    ]
    for x1,x2,x3,x4,m in combos:
        score_add(d, x1+x2+x3+x4, score*m, allow_triple_digits=allow_triple_digits)

@st.cache_data
def load_base_history():
    df = pd.read_excel(DATA_FILE)
    df = df.rename(columns={
        "DrawNo": "draw_no",
        "DrawDate": "draw_date",
        "1stPrizeNo": "first",
        "2ndPrizeNo": "second",
        "3rdPrizeNo": "third",
    })
    df = df[["draw_no", "draw_date", "first", "second", "third"]].dropna()
    # Pastikan semua kolum jadi teks supaya update rekod tidak gagal kerana dtype integer
    for c in ["draw_no", "draw_date", "first", "second", "third"]:
        df[c] = df[c].astype(str).str.strip()
    df["draw_no"] = df["draw_no"].str.zfill(6)
    for c in ["first", "second", "third"]:
        df[c] = df[c].apply(pad4)

    # Susun semula mengikut Draw No supaya latest betul dan tidak bergantung pada susunan baris/cached data.
    df["_draw_sort"] = pd.to_numeric(df["draw_no"], errors="coerce")
    df = df.sort_values("_draw_sort", ascending=True).drop(columns=["_draw_sort"]).reset_index(drop=True)
    return df

def to_original_excel(df):
    out = df.copy()
    out = out.rename(columns={
        "draw_no": "DrawNo",
        "draw_date": "DrawDate",
        "first": "1stPrizeNo",
        "second": "2ndPrizeNo",
        "third": "3rdPrizeNo",
    })
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="Sheet1")
    bio.seek(0)
    return bio


def get_github_token():
    try:
        return st.secrets["GITHUB_TOKEN"]
    except Exception:
        return ""

def github_headers():
    token = get_github_token()
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }

def update_github_excel(df):
    token = get_github_token()
    if not token:
        return False, "GITHUB_TOKEN belum diset dalam Streamlit Secrets."

    url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{GITHUB_FILE_PATH}"
    r = requests.get(url, headers=github_headers(), params={"ref": GITHUB_BRANCH}, timeout=30)
    if r.status_code != 200:
        return False, f"Gagal baca fail GitHub. Status {r.status_code}: {r.text[:300]}"

    sha = r.json().get("sha")
    excel_bytes = to_original_excel(df).getvalue()
    encoded = base64.b64encode(excel_bytes).decode("utf-8")
    payload = {
        "message": "Update TotoHistoryAll.xlsx from Streamlit V11",
        "content": encoded,
        "sha": sha,
        "branch": GITHUB_BRANCH,
    }
    r2 = requests.put(url, headers=github_headers(), json=payload, timeout=60)
    if r2.status_code not in (200, 201):
        return False, f"Gagal update GitHub. Status {r2.status_code}: {r2.text[:500]}"
    return True, "GitHub berjaya dikemaskini."


def get_latest_github_excel_bytes():
    token = get_github_token()
    if not token:
        return None, "GITHUB_TOKEN belum diset dalam Streamlit Secrets."

    url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{GITHUB_FILE_PATH}"
    r = requests.get(url, headers=github_headers(), params={"ref": GITHUB_BRANCH}, timeout=30)
    if r.status_code != 200:
        return None, f"Gagal baca fail GitHub. Status {r.status_code}: {r.text[:300]}"

    content = r.json().get("content", "")
    if not content:
        return None, "Fail GitHub tiada content."
    try:
        return base64.b64decode(content), "OK"
    except Exception as e:
        return None, f"Gagal decode fail GitHub: {e}"


def normalize_history_dataframe(df):
    """
    Normalize Excel history dataframe supaya format sama seperti load_base_history().
    """
    df = df.rename(columns={
        "DrawNo": "draw_no",
        "DrawDate": "draw_date",
        "1stPrizeNo": "first",
        "2ndPrizeNo": "second",
        "3rdPrizeNo": "third",
    })
    df = df[["draw_no", "draw_date", "first", "second", "third"]].dropna()

    for c in ["draw_no", "draw_date", "first", "second", "third"]:
        df[c] = df[c].astype(str).str.strip()

    df["draw_no"] = df["draw_no"].str.zfill(6)
    for c in ["first", "second", "third"]:
        df[c] = df[c].apply(pad4)

    df["_draw_sort"] = pd.to_numeric(df["draw_no"], errors="coerce")
    df = df.sort_values("_draw_sort", ascending=True).drop(columns=["_draw_sort"]).reset_index(drop=True)
    return df


@st.cache_data(ttl=60)
def load_active_history():
    """
    Load active history.
    Priority:
    1. GitHub TotoHistoryAll.xlsx
    2. Local TotoHistoryAll.xlsx
    """
    latest_bytes, latest_msg = get_latest_github_excel_bytes()
    if latest_bytes:
        try:
            df = pd.read_excel(BytesIO(latest_bytes))
            return normalize_history_dataframe(df), "GitHub"
        except Exception:
            pass

    return load_base_history().copy(), "Local"



@st.cache_data
@st.cache_data(show_spinner=False)
def build_audit(history):
    top3 = history[["first", "second", "third"]].values.tolist()
    firsts = history["first"].tolist()
    recent30, recent100, recent500, all_digit = Counter(), Counter(), Counter(), Counter()
    for nums in top3:
        all_digit.update("".join(nums))
    for nums in top3[-30:]:
        recent30.update("".join(nums))
    for nums in top3[-100:]:
        recent100.update("".join(nums))
    for nums in top3[-500:]:
        recent500.update("".join(nums))

    pair_occ, pair_inh = Counter(), Counter()
    pos_trans = {(pos, cur): Counter() for pos in range(4) for cur in "0123456789"}
    missing_next = Counter()

    for i in range(len(top3)-1):
        cur, nxt = top3[i], top3[i+1]
        cur_pairs, nxt_pairs = set(get_pairs(cur)), set(get_pairs(nxt))
        for p in cur_pairs:
            pair_occ[p] += 1
            if p in nxt_pairs:
                pair_inh[p] += 1

        cur_first, nxt_first = firsts[i], firsts[i+1]
        for pos in range(4):
            pos_trans[(pos, cur_first[pos])][nxt_first[pos]] += 1

        cur_digits, nxt_digits = set("".join(cur)), set("".join(nxt))
        for d in "0123456789":
            if d not in cur_digits and d in nxt_digits:
                missing_next[d] += 1

    pair_rate = {}
    for i in range(100):
        p = f"{i:02d}"
        pair_rate[p] = pair_inh[p] / pair_occ[p] if pair_occ[p] else 0

    return {
        "recent30": recent30, "recent100": recent100, "recent500": recent500,
        "all_digit": all_digit, "pair_rate": pair_rate, "pos_trans": pos_trans,
        "missing_next": missing_next,
    }


@st.cache_data(show_spinner=False)
def build_audit_snapshots_fast_v31_29(history, wanted_indices):
    """Satu pass sejarah untuk semua snapshot backtest; setara dengan build_audit(prefix)."""
    h = history.copy().reset_index(drop=True)
    wanted = {int(x) for x in wanted_indices}
    top3 = [[pad4(r[c]) for c in ("first", "second", "third")] for _, r in h.iterrows()]
    firsts = [nums[0] for nums in top3]
    row_digits = [Counter("".join(nums)) for nums in top3]
    all_digit, recent30, recent100, recent500 = Counter(), Counter(), Counter(), Counter()
    pair_occ, pair_inh, missing_next = Counter(), Counter(), Counter()
    pos_trans = {(pos, cur): Counter() for pos in range(4) for cur in "0123456789"}
    out = {}
    for idx, nums in enumerate(top3):
        dc = row_digits[idx]
        all_digit.update(dc); recent30.update(dc); recent100.update(dc); recent500.update(dc)
        if idx >= 30: recent30.subtract(row_digits[idx - 30]); recent30 += Counter()
        if idx >= 100: recent100.subtract(row_digits[idx - 100]); recent100 += Counter()
        if idx >= 500: recent500.subtract(row_digits[idx - 500]); recent500 += Counter()
        if idx > 0:
            cur, nxt = top3[idx - 1], nums
            cur_pairs, nxt_pairs = set(get_pairs(cur)), set(get_pairs(nxt))
            for p in cur_pairs:
                pair_occ[p] += 1
                if p in nxt_pairs: pair_inh[p] += 1
            for pos in range(4):
                pos_trans[(pos, firsts[idx - 1][pos])][firsts[idx][pos]] += 1
            cur_digits, nxt_digits = set("".join(cur)), set("".join(nxt))
            for d in "0123456789":
                if d not in cur_digits and d in nxt_digits: missing_next[d] += 1
        if idx in wanted:
            pair_rate = {
                f"{i:02d}": (pair_inh[f"{i:02d}"] / pair_occ[f"{i:02d}"] if pair_occ[f"{i:02d}"] else 0)
                for i in range(100)
            }
            out[idx] = {
                "recent30": recent30.copy(), "recent100": recent100.copy(),
                "recent500": recent500.copy(), "all_digit": all_digit.copy(),
                "pair_rate": pair_rate,
                "pos_trans": {k: v.copy() for k, v in pos_trans.items()},
                "missing_next": missing_next.copy(),
            }
    return out














































@st.cache_data(show_spinner=False)





































@st.cache_data(show_spinner=False)



def reset_audit_cache():
    build_audit.clear()

def reset_all_caches():
    build_audit.clear()
    load_base_history.clear()
    try:
        load_active_history.clear()
    except Exception:
        pass

base_history_now, history_source_now = load_active_history()

# Force sync: pastikan session_state ikut source aktif terbaru.
# Keutamaan: GitHub TotoHistoryAll.xlsx. Jika GitHub gagal, fallback local.
if (
    "history" not in st.session_state
    or len(st.session_state.history) != len(base_history_now)
    or str(st.session_state.history.iloc[-1]["draw_no"]).zfill(6) != str(base_history_now.iloc[-1]["draw_no"]).zfill(6)
):
    st.session_state.history = base_history_now.copy()

if "prediction_history" not in st.session_state:
    st.session_state.prediction_history = []

history = st.session_state.history
last = history.iloc[-1]

token_status = "Aktif" if get_github_token() else "Belum diset"
history_source_label = history_source_now if "history_source_now" in globals() else "Unknown"
status_dot = '<span class="rap-dot"></span>' if token_status == "Aktif" else ""
st.markdown(
    f"""
    <div class="rap-status-row">
        <span class="rap-badge">{status_dot} GitHub Sync: {token_status}</span>
        <span class="rap-badge">Data: Draw {str(last["draw_no"])}</span>
        <span class="rap-badge">Sumber: {history_source_label}</span>
    </div>
    """,
    unsafe_allow_html=True,
)


# -----------------------------
# V14: History Manager Lengkap
# -----------------------------

st.markdown('<div class="rap-panel-title">Keputusan Terbaru</div>', unsafe_allow_html=True)
try:
    latest = st.session_state.history.iloc[-1]
    latest_draw = str(latest["draw_no"])
    latest_date = str(latest["draw_date"])
    latest_first = pad4(latest["first"])
    latest_second = pad4(latest["second"])
    latest_third = pad4(latest["third"])

    lc1, lc2, lc3, lc4 = st.columns(4)
    lc1.metric("Draw No", latest_draw)
    lc2.metric("1st Prize", latest_first)
    lc3.metric("2nd Prize", latest_second)
    lc4.metric("3rd Prize", latest_third)
    st.caption(f"Tarikh keputusan: {latest_date}")
except Exception:
    st.warning("Keputusan terbaru belum dapat dipaparkan.")

st.markdown('<div class="rap-section-kicker">Tools & Data</div>', unsafe_allow_html=True)
with st.expander("📚 History Manager / Update Keputusan", expanded=False):
    st.subheader("History Manager")
    st.caption("Semua urusan sejarah keputusan dibuat di sini: cari, tambah/update, edit/padam dan download.")

    st.info("Panduan ringkas: gunakan bahagian Tambah / update untuk keputusan baru atau pembetulan. Gunakan Edit / padam hanya jika mahu ubah atau buang draw lama.")


    search_draw = st.text_input("Cari Draw No", value="", placeholder="Contoh: 614826")
    view_df = st.session_state.history.copy()

    view_df["draw_no"] = view_df["draw_no"].astype(str).str.zfill(6)

    if search_draw.strip():
        keyword = search_draw.strip().zfill(6)
        view_df = view_df[view_df["draw_no"] == keyword]
        st.caption(f"Keputusan carian untuk Draw No: {keyword}")
    else:
        view_df = view_df.sort_values("draw_no", ascending=False).head(10)
        st.caption("Paparan 10 draw terakhir")

    recent_view = view_df.copy().rename(columns={
        "draw_no": "Draw No",
        "draw_date": "Draw Date",
        "first": "1st",
        "second": "2nd",
        "third": "3rd",
    })
    st.dataframe(recent_view, hide_index=True, use_container_width=True)

    download_col1, download_col2 = st.columns(2)
    with download_col1:
        st.download_button(
            "Download Current App History",
            data=to_original_excel(st.session_state.history),
            file_name="TotoHistoryAll_current_app.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_current_history",
        )

    with download_col2:
        latest_bytes, latest_msg = get_latest_github_excel_bytes()
        if latest_bytes:
            pass  # Duplicate GitHub history download button removed
        else:
            st.info("Latest GitHub History belum boleh dimuat turun. Pastikan token aktif.")

    with st.expander("History Manager: Edit / padam draw", expanded=False):
        draw_options = st.session_state.history["draw_no"].astype(str).tolist()
        default_idx = len(draw_options) - 1 if draw_options else 0

        if search_draw.strip() and not view_df.empty:
            search_options = view_df["draw_no"].astype(str).tolist()
            selected_draw = st.selectbox(
                "Pilih Draw No untuk edit/padam",
                options=search_options,
                index=len(search_options)-1,
                key="edit_draw_select_search",
            )
        else:
            selected_draw = st.selectbox(
                "Pilih Draw No untuk edit/padam",
                options=draw_options,
                index=default_idx,
                key="edit_draw_select",
            )

        selected_rows = st.session_state.history[
            st.session_state.history["draw_no"].astype(str) == str(selected_draw)
        ]

        if not selected_rows.empty:
            selected_row = selected_rows.iloc[-1]

            action = st.radio(
                "Tindakan",
                ["Update rekod", "Padam rekod"],
                horizontal=True,
                key="history_action_radio",
            )

            if action == "Update rekod":
                with st.form("edit_existing_draw_form"):
                    c0, c1, c2, c3, c4 = st.columns(5)
                    edit_draw_no = c0.text_input("Draw No", value=str(selected_row["draw_no"]), key="edit_draw_no")
                    edit_date = c1.text_input("Draw Date", value=str(selected_row["draw_date"]), key="edit_draw_date")
                    edit_first = c2.text_input("1st", value=pad4(selected_row["first"]), max_chars=4, key="edit_first")
                    edit_second = c3.text_input("2nd", value=pad4(selected_row["second"]), max_chars=4, key="edit_second")
                    edit_third = c4.text_input("3rd", value=pad4(selected_row["third"]), max_chars=4, key="edit_third")
                    edit_auto_save = st.checkbox("Auto-save ke GitHub", value=True, key="edit_auto_save")
                    edit_clicked = st.form_submit_button("Update draw dipilih")

                if edit_clicked:
                    if not (edit_first and edit_second and edit_third):
                        st.error("Sila isi 1st, 2nd dan 3rd.")
                    else:
                        new_history = st.session_state.history.copy()
                        for col in ["draw_no", "draw_date", "first", "second", "third"]:
                            new_history[col] = new_history[col].astype(str)

                        match_idx = new_history.index[
                            new_history["draw_no"].astype(str) == str(selected_draw)
                        ].tolist()

                        if not match_idx:
                            st.error("Draw tidak dijumpai dalam history.")
                        else:
                            idx = match_idx[-1]
                            new_history.at[idx, "draw_no"] = str(edit_draw_no).strip()
                            new_history.at[idx, "draw_date"] = str(edit_date).strip()
                            new_history.at[idx, "first"] = pad4(edit_first)
                            new_history.at[idx, "second"] = pad4(edit_second)
                            new_history.at[idx, "third"] = pad4(edit_third)

                            st.session_state.history = new_history
                            build_audit.clear()

                            if edit_auto_save:
                                ok, msg = update_github_excel(new_history)
                                if ok:
                                    st.success(f"Draw {selected_draw} berjaya dikemaskini dan GitHub berjaya dikemaskini.")
                                    reset_all_caches()
                                else:
                                    st.warning(f"Draw {selected_draw} dikemaskini dalam sesi app, tetapi GitHub belum dikemaskini.")
                                    st.error(msg)
                            else:
                                st.success(f"Draw {selected_draw} dikemaskini dalam sesi app sahaja.")

                            st.rerun()

            else:
                st.warning(f"Anda akan memadam Draw No {selected_draw}. Tindakan ini tidak boleh dibatalkan selepas auto-save.")
                confirm_delete = st.checkbox("Saya sahkan mahu padam rekod ini", key="confirm_delete")
                delete_auto_save = st.checkbox("Auto-save ke GitHub", value=True, key="delete_auto_save")
                if st.button("Padam draw dipilih", disabled=not confirm_delete):
                    new_history = st.session_state.history.copy()
                    for col in ["draw_no", "draw_date", "first", "second", "third"]:
                        new_history[col] = new_history[col].astype(str)

                    match_idx = new_history.index[
                        new_history["draw_no"].astype(str) == str(selected_draw)
                    ].tolist()

                    if not match_idx:
                        st.error("Draw tidak dijumpai dalam history.")
                    else:
                        idx = match_idx[-1]
                        new_history = new_history.drop(index=idx).reset_index(drop=True)

                        st.session_state.history = new_history
                        build_audit.clear()

                        if delete_auto_save:
                            ok, msg = update_github_excel(new_history)
                            if ok:
                                st.success(f"Draw {selected_draw} berjaya dipadam dan GitHub berjaya dikemaskini.")
                                reset_all_caches()
                            else:
                                st.warning(f"Draw {selected_draw} dipadam dalam sesi app, tetapi GitHub belum dikemaskini.")
                                st.error(msg)
                        else:
                            st.success(f"Draw {selected_draw} dipadam dalam sesi app sahaja.")

                        st.rerun()

    st.divider()

    st.divider()


if False:
    pass
# Analysis / Hot & Cold Digits removed
    st.subheader("V17 Analysis")
    ana_c1, ana_c2 = st.columns(2)
    with ana_c1:
        hot_window = st.selectbox("Hot Digit Window", [10, 30, 50, 100], index=1)
        hot_df_preview = hot_digit_analysis(st.session_state.history, window=hot_window)
        st.write(f"Hot Digits - {hot_window} draw terakhir")
        st.dataframe(hot_df_preview, hide_index=True, use_container_width=True)
    with ana_c2:
        cold_window = st.selectbox("Cold Digit Window", [10, 30, 50, 100], index=3)
        cold_df_preview = cold_digit_analysis(st.session_state.history, window=cold_window)
        st.write(f"Cold Digits - {cold_window} draw terakhir")
        st.dataframe(cold_df_preview, hide_index=True, use_container_width=True)

    st.divider()

    with st.expander("History Manager: Tambah / update keputusan", expanded=True):
        with st.form("add_result_form"):
            c0, c1, c2, c3, c4 = st.columns(5)
            try:
                suggested_draw = str(int(last["draw_no"]) + 100)
            except Exception:
                suggested_draw = ""
            next_draw = c0.text_input("Draw No", value=suggested_draw)
            draw_date = c1.text_input("Draw Date", value="")
            new_first = c2.text_input("1st", max_chars=4)
            new_second = c3.text_input("2nd", max_chars=4)
            new_third = c4.text_input("3rd", max_chars=4)

            draw_exists = str(next_draw).strip() in set(st.session_state.history["draw_no"].astype(str))
            if draw_exists:
                st.warning(f"Draw No {next_draw} sudah wujud dalam history. Pilih sama ada mahu update rekod lama atau tambah baris baru.")
                save_mode = st.radio(
                    "Tindakan",
                    ["Update rekod sedia ada", "Tambah sebagai baris baru"],
                    horizontal=True,
                )
            else:
                save_mode = "Tambah sebagai baris baru"

            auto_save = st.checkbox("Auto-save ke GitHub", value=True)
            add_clicked = st.form_submit_button("Simpan keputusan")

        if add_clicked:
            if not (new_first and new_second and new_third):
                st.error("Sila isi 1st, 2nd dan 3rd.")
            else:
                new_row = {
                    "draw_no": str(next_draw).strip(),
                    "draw_date": str(draw_date).strip(),
                    "first": pad4(new_first),
                    "second": pad4(new_second),
                    "third": pad4(new_third),
                }

                new_history = st.session_state.history.copy()
                # Tukar semua kolum kepada object/string supaya pandas tidak reject update nilai teks
                for col in ["draw_no", "draw_date", "first", "second", "third"]:
                    new_history[col] = new_history[col].astype(str)
                match_idx = new_history.index[new_history["draw_no"].astype(str) == str(next_draw).strip()].tolist()

                if match_idx and save_mode == "Update rekod sedia ada":
                    idx = match_idx[-1]
                    # Update satu kolum demi satu kolum supaya stabil di Streamlit Cloud / pandas baru
                    new_history.at[idx, "draw_no"] = str(new_row["draw_no"])
                    new_history.at[idx, "draw_date"] = str(new_row["draw_date"])
                    new_history.at[idx, "first"] = str(new_row["first"])
                    new_history.at[idx, "second"] = str(new_row["second"])
                    new_history.at[idx, "third"] = str(new_row["third"])
                    action_msg = f"Draw {next_draw} dikemaskini."
                else:
                    new_history = pd.concat([new_history, pd.DataFrame([new_row])], ignore_index=True)
                    action_msg = f"Draw {next_draw} ditambah sebagai baris baru."

                st.session_state.history = new_history
                reset_audit_cache()

                if auto_save:
                    ok, msg = update_github_excel(new_history)
                    if ok:
                        st.success(action_msg + " GitHub berjaya dikemaskini.")
                        reset_all_caches()
                    else:
                        st.warning(action_msg + " Tetapi GitHub belum dikemaskini.")
                        st.error(msg)
                else:
                    st.success(action_msg + " Disimpan dalam sesi app sahaja.")
                st.rerun()

    st.download_button(
        "Download Updated TotoHistoryAll.xlsx",
        data=to_original_excel(st.session_state.history),
        file_name="TotoHistoryAll_updated.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.divider()

    last = st.session_state.history.iloc[-1]




def unordered_digit_key4(n):
    """Kunci semakan hit tanpa susunan; tidak digunakan untuk memilih calon."""
    try:
        return "".join(sorted(pad4(n)))
    except Exception:
        return ""


def pair_digit_key(pair):
    """Samakan pair terbalik, contohnya 13 dan 31."""
    return "".join(sorted(str(pair).zfill(2)[-2:]))


def keep_first_pair_orientation(pair_rows):
    """Kekalkan orientasi pair yang muncul dahulu sahaja."""
    kept = []
    seen = set()
    for row in pair_rows:
        key = pair_digit_key(row.get("Pair", ""))
        if key in seen:
            continue
        seen.add(key)
        kept.append(row)
    return kept








def build_bridge_model_v31_9(first, second, third):
    import pandas as pd
    nums=[pad4(first),pad4(second),pad4(third)]
    existing_digits=sorted(set("".join(nums)))
    missing_digits=sorted(set("0123456789")-set(existing_digits))
    pair_rows=[]; base_pairs=[]
    number_meta={}
    bridge_order=[]

    for label,no in zip(["1st","2nd","3rd"],nums):
        for ptype,pair in zip(["Front","Middle","Back"],[no[:2],no[1:3],no[2:4]]):
            base_pairs.append(pair)
            pair_rows.append({"Source":label,"No":no,"Pair Type":ptype,"Pair":pair})

    pair_rows = keep_first_pair_orientation(pair_rows)
    base_pairs = [row["Pair"] for row in pair_rows]

    for row in pair_rows:
        pair=row["Pair"]
        src=row["Source"]
        ptype=row["Pair Type"]
        for md in missing_digits:
            for ed in existing_digits:
                # V31.21: ikut tertib pair asal.
                # Contoh 82 + 7 + 6 = 8276, bukan canonical 2678 untuk paparan.
                display_no = f"{pair}{md}{ed}"
                if len(display_no)==4 and display_no.isdigit():
                    if display_no not in number_meta:
                        number_meta[display_no]={
                            "Display No":display_no,
                            "Formula List":[],
                            "Base Pairs":set(),
                            "Sources":set(),
                            "Pair Types":set(),
                            "Missing Digits":set(),
                            "Existing Digits":set(),
                        }
                        bridge_order.append(display_no)

                    formula=f"{pair}+{md}+{ed}"
                    number_meta[display_no]["Formula List"].append(formula)
                    number_meta[display_no]["Base Pairs"].add(pair)
                    number_meta[display_no]["Sources"].add(src)
                    number_meta[display_no]["Pair Types"].add(ptype)
                    number_meta[display_no]["Missing Digits"].add(md)
                    number_meta[display_no]["Existing Digits"].add(ed)

    rows=[]
    for order_idx, display_no in enumerate(bridge_order, start=1):
        meta = number_meta[display_no]
        rows.append({
            "No": meta["Display No"],
            "Order": order_idx,
            "Formula Support": len(set(meta["Formula List"])),
            "Source Support": len(meta["Sources"]),
            "Position Support": len(meta["Pair Types"]),
            "Base Pair Support": len(meta["Base Pairs"]),
            "Base Pairs": " / ".join(sorted(meta["Base Pairs"])),
            "Sources": " / ".join(sorted(meta["Sources"])),
            "Pair Types": " / ".join(sorted(meta["Pair Types"])),
            "Missing Digits": " / ".join(sorted(meta["Missing Digits"])),
            "Existing Digits": " / ".join(sorted(meta["Existing Digits"])),
            "Formula List": " / ".join(sorted(set(meta["Formula List"]))),
        })

    bridge_df=pd.DataFrame(rows)
    if not bridge_df.empty:
        bridge_df=bridge_df.sort_values(["Order"]).reset_index(drop=True)

    text="🧪 Rumah A Predictor - Bridge Model\n\n"
    text+="Base Pairs:\n"+" / ".join(base_pairs)
    text+="\n\nMissing Digits:\n"+" / ".join(missing_digits)
    text+="\n\nExisting Digits:\n"+" / ".join(existing_digits)
    nums_out=bridge_df["No"].astype(str).tolist() if not bridge_df.empty and "No" in bridge_df.columns else []
    text+=f"\n\nBridge Numbers (Total: {len(nums_out)}):\n"
    text += "\n".join([" / ".join(nums_out[i:i+10]) for i in range(0,len(nums_out),10)]) if nums_out else "Tiada output."
    return pd.DataFrame(pair_rows), bridge_df, text


def build_bridge_engine_v2_pair_double_digit(first, second, third):
    """Bridge V2: pair + 2 missing digits OR pair + 2 existing digits."""
    nums = [pad4(first), pad4(second), pad4(third)]
    existing_digits = sorted(set("".join(nums)))
    missing_digits = sorted(set("0123456789") - set(existing_digits))
    pair_rows, base_pairs = [], []
    for label, no in zip(["1st", "2nd", "3rd"], nums):
        for pair_type, pair in zip(["Front", "Middle", "Back"], [no[:2], no[1:3], no[2:4]]):
            pair_rows.append({"Source": label, "No": no, "Pair Type": pair_type, "Pair": pair})
            base_pairs.append(pair)
    pair_rows = keep_first_pair_orientation(pair_rows)
    base_pairs = [row["Pair"] for row in pair_rows]

    number_meta = {}
    def add_candidate(pair, d1, d2, mode, source, pair_type):
        display_no = f"{pair}{d1}{d2}"
        meta = number_meta.setdefault(display_no, {
            "No": display_no, "Modes": set(), "Base Pairs": set(),
            "Sources": set(), "Pair Types": set(), "Formula List": set(),
        })
        meta["Modes"].add(mode); meta["Base Pairs"].add(pair)
        meta["Sources"].add(source); meta["Pair Types"].add(pair_type)
        meta["Formula List"].add(f"{pair}+{d1}{d2}")

    for row in pair_rows:
        pair, source, pair_type = row["Pair"], row["Source"], row["Pair Type"]
        for digit_pool, mode in [(missing_digits, "2 Missing"), (existing_digits, "2 Existing")]:
            for d1 in digit_pool:
                for d2 in digit_pool:
                    if d1 != d2:
                        add_candidate(pair, d1, d2, mode, source, pair_type)

    rows = []
    for order, meta in enumerate(number_meta.values(), 1):
        rows.append({
            "No": meta["No"], "Order": order,
            "Mode": " / ".join(sorted(meta["Modes"])),
            "Formula Support": len(meta["Formula List"]), "Source Support": len(meta["Sources"]),
            "Position Support": len(meta["Pair Types"]), "Base Pair Support": len(meta["Base Pairs"]),
            "Base Pairs": " / ".join(sorted(meta["Base Pairs"])),
            "Sources": " / ".join(sorted(meta["Sources"])),
            "Pair Types": " / ".join(sorted(meta["Pair Types"])),
            "Formula List": " / ".join(sorted(meta["Formula List"])),
        })
    bridge_v2_df = pd.DataFrame(rows)
    text = "🧪 Rumah A Predictor - Bridge Engine V2\n\n"
    text += "Base Pairs:\n" + " / ".join(base_pairs)
    text += "\n\nMissing Digits:\n" + " / ".join(missing_digits)
    text += "\n\nExisting Digits:\n" + " / ".join(existing_digits)
    for mode in ("2 Missing", "2 Existing"):
        vals = bridge_v2_df[bridge_v2_df["Mode"].str.contains(mode, regex=False)]["No"].astype(str).tolist() if not bridge_v2_df.empty else []
        text += f"\n\n{mode} Numbers (Total: {len(vals)}):\n"
        text += "\n".join(" / ".join(vals[i:i+10]) for i in range(0, len(vals), 10)) if vals else "Tiada output."
    return pd.DataFrame(pair_rows), bridge_v2_df, text


def _ordered_top3_pairs(first, second, third):
    """Pair Top 3 unik; pasangan terbalik dikira sebagai pair yang sama."""
    rows = []
    for source, no in zip(("1st", "2nd", "3rd"), (pad4(first), pad4(second), pad4(third))):
        for pair_type, pair in zip(("Front", "Middle", "Back"), (no[:2], no[1:3], no[2:4])):
            rows.append({"Source": source, "Pair Type": pair_type, "Pair": pair})
    return keep_first_pair_orientation(rows)


@st.cache_data(show_spinner=False)
def build_bridge_pair_priority(history, first, second, third, lookback=100):
    """Rank pair daripada draw terkini; satu draw dikira sekali jika V1 atau V2 hit."""
    columns = [
        "Priority", "Source", "Pair Position", "Current Pair",
        "V1 Hit", "V2 Hit", "Total Support", "Hit Rate %", "Transitions",
    ]
    if history is None or history.empty or len(history) < 2:
        return pd.DataFrame(columns=columns)

    h = history.copy().reset_index(drop=True)
    if lookback and len(h) > int(lookback) + 1:
        h = h.tail(int(lookback) + 1).reset_index(drop=True)
    slots = [
        ("1st", "Front", 0, "first"),
        ("1st", "Middle", 1, "first"),
        ("1st", "Back", 2, "first"),
        ("2nd", "Front", 0, "second"),
        ("2nd", "Middle", 1, "second"),
        ("2nd", "Back", 2, "second"),
        ("3rd", "Front", 0, "third"),
        ("3rd", "Middle", 1, "third"),
        ("3rd", "Back", 2, "third"),
    ]
    v1_hits = Counter()
    v2_hits = Counter()
    combined_hits = Counter()
    transitions = len(h) - 1
    for idx in range(len(h) - 1):
        source_numbers = [pad4(h.iloc[idx][c]) for c in ("first", "second", "third")]
        existing_digits = sorted(set("".join(source_numbers)))
        missing_digits = sorted(set("0123456789") - set(existing_digits))
        target_digit_keys = {
            unordered_digit_key4(h.iloc[idx + 1][c]) for c in ("first", "second", "third")
        }
        for source, position, start, column in slots:
            pair = pad4(h.iloc[idx][column])[start:start + 2]
            bridge_v1_digit_keys = {
                unordered_digit_key4(f"{pair}{missing}{existing}")
                for missing in missing_digits
                for existing in existing_digits
            }
            bridge_v2_digit_keys = {
                unordered_digit_key4(f"{pair}{d1}{d2}")
                for pool in (missing_digits, existing_digits)
                for d1 in pool
                for d2 in pool
                if d1 != d2
            }
            v1_hit_now = bool(bridge_v1_digit_keys & target_digit_keys)
            v2_hit_now = bool(bridge_v2_digit_keys & target_digit_keys)
            if v1_hit_now:
                v1_hits[(source, position)] += 1
            if v2_hit_now:
                v2_hits[(source, position)] += 1
            if v1_hit_now or v2_hit_now:
                combined_hits[(source, position)] += 1

    current = {"first": pad4(first), "second": pad4(second), "third": pad4(third)}
    rows = []
    for original_order, (source, position, start, column) in enumerate(slots):
        v1_hit = int(v1_hits[(source, position)])
        v2_hit = int(v2_hits[(source, position)])
        combined_hit = int(combined_hits[(source, position)])
        rows.append({
            "Source": source,
            "Pair Position": position,
            "Current Pair": current[column][start:start + 2],
            "V1 Hit": v1_hit,
            "V2 Hit": v2_hit,
            "Total Support": combined_hit,
            "Hit Rate %": round((combined_hit / transitions) * 100, 1) if transitions else 0.0,
            "Transitions": transitions,
            "_Original Order": original_order,
        })

    ranked = pd.DataFrame(rows).sort_values(
        ["Total Support", "V1 Hit", "_Original Order"],
        ascending=[False, False, True],
        kind="stable",
    ).reset_index(drop=True)
    kept_indexes, seen_pair_keys = [], set()
    for index, row in ranked.iterrows():
        key = pair_digit_key(row["Current Pair"])
        if key in seen_pair_keys:
            continue
        seen_pair_keys.add(key)
        kept_indexes.append(index)
    ranked = ranked.loc[kept_indexes].reset_index(drop=True)
    ranked.insert(0, "Priority", range(1, len(ranked) + 1))
    return ranked.drop(columns=["_Original Order"])


def build_bridge_pair_priority_numbers(pair, pair_audit_row, first, second, third):
    """Keluarkan V1 dan V2 untuk satu pair sahaja; provenance route tidak dicampurkan."""
    columns = ["Pair", "No", "Route"]
    if not pair:
        return pd.DataFrame(columns=columns), ""

    nums = [pad4(first), pad4(second), pad4(third)]
    existing_digits = sorted(set("".join(nums)))
    missing_digits = sorted(set("0123456789") - set(existing_digits))
    pair = str(pair).zfill(2)[-2:]
    number_meta = {}

    def add_number(no, route):
        key = (route, no)
        if key not in number_meta:
            number_meta[key] = {"Pair": pair, "No": no, "Route": route}

    for missing in missing_digits:
        for existing in existing_digits:
            add_number(f"{pair}{missing}{existing}", "Bridge V1")
    for d1 in missing_digits:
        for d2 in missing_digits:
            if d1 != d2:
                add_number(f"{pair}{d1}{d2}", "Bridge V2 - 2 Missing")
    for d1 in existing_digits:
        for d2 in existing_digits:
            if d1 != d2:
                add_number(f"{pair}{d1}{d2}", "Bridge V2 - 2 Existing")

    rows = [
        {
            "Pair": meta["Pair"], "No": meta["No"], "Route": meta["Route"],
        }
        for meta in number_meta.values()
    ]
    number_df = pd.DataFrame(rows, columns=columns)
    text_lines = [
        "🧭 Rumah A Predictor - Bridge Pair Shortlist", "",
        f'Pair Pilihan: {pair}',
        f'Sumber Ranking: {pair_audit_row["Source"]} Prize - {pair_audit_row["Pair Position"]}',
        f'V1 Hit: {int(pair_audit_row["V1 Hit"])}',
        f'V2 Hit: {int(pair_audit_row["V2 Hit"])}',
        f'Total Support: {int(pair_audit_row["Total Support"])}',
    ]
    for route in ("Bridge V1", "Bridge V2 - 2 Missing", "Bridge V2 - 2 Existing"):
        route_values = number_df[number_df["Route"].str.contains(route, regex=False)]["No"].tolist()
        text_lines.extend(["", f"{route} (Pilihan Unik: {len(route_values)}):"])
        text_lines.extend(" / ".join(route_values[i:i + 10]) for i in range(0, len(route_values), 10))
    text_lines.extend(["", f"Jumlah Pilihan Unik Pair {pair}: {len(number_df)}"])
    return number_df, "\n".join(text_lines)






def build_second_pair_shortlist(pair, pair_numbers_df, first, second, third):
    """Tapis nombor asal yang mengekalkan generator pair dan current pair lain."""
    columns = ["Generator Pair", "No", "Bridge", "Pair Kedua"]
    if pair_numbers_df is None or pair_numbers_df.empty:
        return pd.DataFrame(columns=columns), ""

    current_rows = _ordered_top3_pairs(first, second, third)
    current_pairs = list(dict.fromkeys(str(row["Pair"]) for row in current_rows))
    other_pairs = [value for value in current_pairs if value != str(pair)]
    rows = []
    for _, row in pair_numbers_df.iterrows():
        number = pad4(row["No"])
        supporting_pairs = [value for value in other_pairs if value in number]
        if not supporting_pairs:
            continue
        rows.append({
            "Generator Pair": str(pair),
            "No": number,
            "Bridge": str(row["Route"]),
            "Pair Kedua": " / ".join(supporting_pairs),
        })

    shortlist_df = pd.DataFrame(rows, columns=columns)
    text_lines = [
        "🔗 Rumah A Predictor - Bridge Dua Pair", "",
        f"Generator Pair: {pair}",
        f"Jumlah Pilihan: {len(shortlist_df)}",
    ]
    for route in ("Bridge V1", "Bridge V2 - 2 Missing", "Bridge V2 - 2 Existing"):
        route_df = shortlist_df[shortlist_df["Bridge"] == route]
        if route_df.empty:
            continue
        text_lines.extend(["", f"{route} ({len(route_df)} Pilihan):"])
        for _, item in route_df.iterrows():
            text_lines.append(
                f'{item["No"]} | Pair Kedua {item["Pair Kedua"]}'
            )
    return shortlist_df, "\n".join(text_lines)






def build_chart_3d_signal_v31_39(first, second, third, bridge_v1_df=None, bridge_v2_df=None):
    """Carta ringan: Menegak/L + Bridge sahaja, tanpa imbasan bentuk Tetris 4D."""
    numbers = [pad4(first), pad4(second), pad4(third)]
    digit_sums = [sum(int(digit) for digit in number) for number in numbers]
    digit_roots = [0 if value == 0 else 1 + (value - 1) % 9 for value in digit_sums]
    total_sum = str(sum(digit_sums))
    root_sum = str(sum(digit_roots))
    cross_rows = [
        "".join(str(int(top_digit) + int(bottom_digit)) for bottom_digit in root_sum)
        for top_digit in total_sum
    ]
    final_row = str(sum(int(digit) for digit in total_sum)) + str(
        sum(int(digit) for digit in root_sum)
    )
    derived_rows = cross_rows + [final_row]
    chart_rows = [total_sum, root_sum] + derived_rows

    three_d_rows, seen = [], set()
    max_width = max(len(row) for row in derived_rows)
    for column in range(max_width):
        if all(column < len(row) for row in derived_rows):
            anchor = "".join(row[column] for row in derived_rows)
            key = ("Menegak", anchor)
            if len(anchor) == 3 and key not in seen:
                seen.add(key)
                three_d_rows.append({"Pilihan": "Menegak", "3D": anchor})
    for row_index in range(len(derived_rows) - 1):
        top_row, bottom_row = derived_rows[row_index], derived_rows[row_index + 1]
        for column in range(min(len(top_row), len(bottom_row)) - 1):
            choices = [
                ("L Kiri", top_row[column] + bottom_row[column] + bottom_row[column + 1]),
                ("L Kanan", top_row[column + 1] + bottom_row[column + 1] + bottom_row[column]),
            ]
            # Lengkapkan orientasi L atas pada baris campur-silang sahaja.
            # Blok 13 / 12 menghasilkan 113; baris jumlah akhir tidak diperluas.
            if row_index < len(cross_rows) - 1:
                upper_l = top_row[column] + bottom_row[column] + top_row[column + 1]
                if (
                    upper_l not in {anchor for _, anchor in choices}
                    and not any(existing_anchor == upper_l for _, existing_anchor in seen)
                ):
                    choices.append(("L Atas", upper_l))
            for label, anchor in choices:
                if any(
                    existing_label != "Menegak" and existing_anchor == anchor
                    for existing_label, existing_anchor in seen
                ):
                    continue
                key = (label, anchor)
                if key not in seen:
                    seen.add(key)
                    three_d_rows.append({"Pilihan": label, "3D": anchor})
    three_d_df = pd.DataFrame(three_d_rows, columns=["Pilihan", "3D"])

    def bridge_lookup(frame):
        numbers = []
        if frame is None or frame.empty:
            return numbers
        for _, row in frame.iterrows():
            number = pad4(row.get("No", ""))
            if number and number not in numbers:
                numbers.append(number)
        return numbers

    v1_numbers = bridge_lookup(bridge_v1_df)
    v2_numbers = bridge_lookup(bridge_v2_df)
    confirmed_rows = []
    for _, choice in three_d_df.iterrows():
        anchor = str(choice["3D"])
        for bridge_name, bridge_numbers in (
            ("V1", v1_numbers),
            ("V2", v2_numbers),
        ):
            for number in bridge_numbers:
                if Counter(anchor) - Counter(number):
                    continue
                confirmed_rows.append({
                    "Pilihan": str(choice["Pilihan"]),
                    "3D": anchor,
                    "No": number,
                    "Bridge": bridge_name,
                })
    confirmed_df = pd.DataFrame(
        confirmed_rows,
        columns=["Pilihan", "3D", "No", "Bridge"],
    )
    if not confirmed_df.empty:
        confirmed_df = (
            confirmed_df.drop_duplicates()
            .sort_values(["Pilihan", "3D", "Bridge", "No"])
            .reset_index(drop=True)
        )

    vertical_values = three_d_df[three_d_df["Pilihan"] == "Menegak"]["3D"].tolist()
    l_values = three_d_df[three_d_df["Pilihan"] != "Menegak"]["3D"].tolist()
    chart_text = (
        "🧩 Rumah A Predictor - Carta 3D V2\n\n"
        f"Top 3: {' / '.join(numbers)}\n"
        f"Jumlah Digit: {' / '.join(str(value) for value in digit_sums)}\n"
        f"Digital Root: {' / '.join(str(value) for value in digit_roots)}\n"
        f"Asas: {total_sum} / {root_sum}\n\n"
        + "\n".join(chart_rows)
        + f"\n\nPilihan Menegak: {' / '.join(vertical_values) or 'Tiada'}"
        + f"\nPilihan L: {' / '.join(l_values) or 'Tiada'}"
    )
    choice_lines = [
        "🎯 Rumah A Predictor - Pilihan Carta 3D + Bridge",
        "",
        f"Pilihan Menegak: {' / '.join(vertical_values) or 'Tiada'}",
        f"Pilihan L: {' / '.join(l_values) or 'Tiada'}",
        f"Jumlah 3D Carta + Bridge: {len(confirmed_df)}",
    ]
    if confirmed_df.empty:
        choice_lines.extend(["", "Tiada pilihan Carta 3D yang disahkan Bridge."])
    else:
        choice_lines.extend(["", "3D Carta + Bridge:"])
        for _, row in confirmed_df.iterrows():
            choice_lines.append(
                f'{row["Pilihan"]} {row["3D"]} | {row["Bridge"]} | {row["No"]}'
            )
    meta = {
        "Rows": chart_rows,
        "3D Choices": three_d_df,
        "3D Confirmed": confirmed_df,
    }
    return chart_text, "\n".join(choice_lines), meta


@st.cache_data(show_spinner=False)








@st.cache_data(show_spinner=False)
def run_backtest_bridge_dde_lite_v31_24_5(history_df, test_draws=30):
    import json
    import time
    t0 = time.perf_counter()
    if history_df is None or history_df.empty or len(history_df) < 2:
        return pd.DataFrame(), pd.DataFrame()
    h = history_df.copy().reset_index(drop=True)
    for col in ("first", "second", "third"):
        h[col] = h[col].apply(pad4)
    latest_idx = len(h) - 1
    count = max(1, min(int(test_draws), latest_idx + 1))
    start_idx = max(0, latest_idx - count + 1)
    cache_path = Path(".backtest_row_cache_v31_54_first_routes.json")
    cache = {}
    try:
        if cache_path.exists():
            payload = json.loads(cache_path.read_text(encoding="utf-8"))
            if payload.get("version") == "v31.54-first-routes":
                cache = payload.get("rows", {})
    except Exception:
        cache = {}
    rows = []
    for idx in range(start_idx, latest_idx + 1):
        source = h.iloc[idx]
        first, second, third = (pad4(source[c]) for c in ("first", "second", "third"))
        if idx + 1 < len(h):
            nxt = h.iloc[idx + 1]
            next_sig = "|".join([str(nxt.get("draw_no", ""))] + [pad4(nxt[c]) for c in ("first", "second", "third")])
        else:
            nxt, next_sig = None, "PENDING"
        key = "|".join([str(source.get("draw_no", idx)), first, second, third, next_sig])
        if key in cache:
            rows.append(cache[key])
            continue
        _, v1_df, _ = build_bridge_model_v31_9(first, second, third)
        _, v2_df, _ = build_bridge_engine_v2_pair_double_digit(first, second, third)
        v1_list = v1_df["No"].astype(str).tolist() if not v1_df.empty else []
        v2_list = v2_df["No"].astype(str).tolist() if not v2_df.empty else []
        v1_digit_keys = {unordered_digit_key4(x) for x in v1_list}
        v2_digit_keys = {unordered_digit_key4(x) for x in v2_list}
        v2_missing = {
            unordered_digit_key4(x)
            for x in v2_df[v2_df["Mode"].str.contains("2 Missing", regex=False)]["No"].astype(str)
        } if not v2_df.empty else set()
        v2_existing = {
            unordered_digit_key4(x)
            for x in v2_df[v2_df["Mode"].str.contains("2 Existing", regex=False)]["No"].astype(str)
        } if not v2_df.empty else set()
        if nxt is None:
            next_draw, next_result = "", "Belum ada next draw"
            actual_nums, actual_digit_keys = [], []
            status = "PENDING"
        else:
            actual_nums = [pad4(nxt[c]) for c in ("first", "second", "third")]
            actual_digit_keys = [unordered_digit_key4(x) for x in actual_nums]
            next_draw, next_result, status = str(nxt.get("draw_no", "")), " / ".join(actual_nums), "DONE"
        v1_hits = [n for n, key in zip(actual_nums, actual_digit_keys) if key in v1_digit_keys]
        v2_hits = [n for n, key in zip(actual_nums, actual_digit_keys) if key in v2_digit_keys]
        missing_hits = [n for n, key in zip(actual_nums, actual_digit_keys) if key in v2_missing]
        existing_hits = [n for n, key in zip(actual_nums, actual_digit_keys) if key in v2_existing]
        union_hits = list(dict.fromkeys(v1_hits + v2_hits))

        # 1st 2D + missing + digit daripada 2nd atau 3rd.
        source_top3 = [first, second, third]
        carry_missing = sorted(set("0123456789") - set("".join(source_top3)))
        first_missing_keys = {"Digit 2nd": set(), "Digit 3rd": set()}
        first_missing_positions = {"Digit 2nd": [], "Digit 3rd": []}
        first_missing_routes = {"Digit 2nd": [], "Digit 3rd": []}
        for left, right in combinations(range(4), 2):
            position = f"{left + 1}+{right + 1}"
            duo = first[left] + first[right]
            for source_name, source_number in (
                ("Digit 2nd", second), ("Digit 3rd", third)
            ):
                for digit_position, digit in enumerate(source_number, start=1):
                    for missing_digit in carry_missing:
                        generated = f"{duo}{missing_digit}{digit}"
                        generated_key = unordered_digit_key4(generated)
                        first_missing_keys[source_name].add(generated_key)
                        for target_number, target_key in zip(actual_nums, actual_digit_keys):
                            if generated_key != target_key:
                                continue
                            first_missing_positions[source_name].append(position)
                            first_missing_routes[source_name].append(
                                f"{position} x {source_name}-{digit_position} | "
                                f"{generated} -> {target_number}"
                            )
        first_missing_hits = {
            source_name: list(dict.fromkeys(
                number for number, target_key in zip(actual_nums, actual_digit_keys)
                if target_key in keys
            )) for source_name, keys in first_missing_keys.items()
        }
        first_missing_bridge_hits = {
            source_name: list(dict.fromkeys(
                number for number, target_key in zip(actual_nums, actual_digit_keys)
                if target_key in keys and target_key in v1_digit_keys
            )) for source_name, keys in first_missing_keys.items()
        }

        # 1st 2D + dua digit yang benar-benar muncul dalam 2nd dan 3rd.
        second_third_suffixes = _occurrence_pairs_two_prizes(second, third)
        first_second_third_keys, first_second_third_routes = set(), []
        first_second_third_positions = []
        for left, right in combinations(range(4), 2):
            position = f"{left + 1}+{right + 1}"
            duo = first[left] + first[right]
            for suffix in second_third_suffixes:
                generated = f"{duo}{suffix}"
                generated_key = unordered_digit_key4(generated)
                first_second_third_keys.add(generated_key)
                for target_number, target_key in zip(actual_nums, actual_digit_keys):
                    if generated_key != target_key:
                        continue
                    first_second_third_positions.append(position)
                    first_second_third_routes.append(
                        f"{position} | {duo}+{suffix}={generated} -> {target_number}"
                    )
        first_second_third_hits = list(dict.fromkeys(
            number for number, target_key in zip(actual_nums, actual_digit_keys)
            if target_key in first_second_third_keys
        ))
        first_second_third_bridge_hits = list(dict.fromkeys(
            number for number, target_key in zip(actual_nums, actual_digit_keys)
            if target_key in first_second_third_keys and target_key in v2_digit_keys
        ))

        # Engine berasingan: 2D daripada 2nd + missing + satu digit 1st.
        # Semua 6 kedudukan 2D dan semua 4 kedudukan digit 1st diuji.
        carry_routes = []
        carry_candidate_keys = set()
        carry_hit_numbers = []
        carry_hit_second_positions = []
        carry_hit_first_positions = []
        for left, right in combinations(range(4), 2):
            second_position = f"{left + 1}+{right + 1}"
            duo = second[left] + second[right]
            for first_position in range(4):
                first_digit = first[first_position]
                for missing_digit in carry_missing:
                    generated = f"{duo}{missing_digit}{first_digit}"
                    generated_key = unordered_digit_key4(generated)
                    carry_candidate_keys.add(generated_key)
                    for target_number, target_key in zip(actual_nums, actual_digit_keys):
                        if generated_key != target_key:
                            continue
                        carry_hit_numbers.append(target_number)
                        carry_hit_second_positions.append(second_position)
                        carry_hit_first_positions.append(str(first_position + 1))
                        carry_routes.append(
                            f"{second_position} x 1st-{first_position + 1} | "
                            f"{duo}+{missing_digit}+{first_digit} | "
                            f"{generated} -> {target_number}"
                        )
        carry_hit_numbers = list(dict.fromkeys(carry_hit_numbers))
        carry_hit_second_positions = list(dict.fromkeys(carry_hit_second_positions))
        carry_hit_first_positions = list(dict.fromkeys(carry_hit_first_positions))
        carry_routes = list(dict.fromkeys(carry_routes))

        # Engine berasingan: 2D daripada 2nd + dua digit 1st/3rd mengikut
        # kemunculan sebenar. Double dibenarkan hanya jika benar-benar muncul.
        first_third_pool = first + third
        first_third_suffixes = sorted({
            "".join(sorted((first_third_pool[left], first_third_pool[right])))
            for left, right in combinations(range(len(first_third_pool)), 2)
        })
        ft_candidate_keys = set()
        ft_hit_numbers, ft_hit_positions, ft_hit_duos, ft_hit_routes = [], [], [], []
        for left, right in combinations(range(4), 2):
            ft_position = f"{left + 1}+{right + 1}"
            ft_duo = second[left] + second[right]
            for ft_suffix in first_third_suffixes:
                ft_generated = f"{ft_duo}{ft_suffix}"
                ft_key = unordered_digit_key4(ft_generated)
                ft_candidate_keys.add(ft_key)
                for target_number, target_key in zip(actual_nums, actual_digit_keys):
                    if ft_key != target_key:
                        continue
                    ft_hit_numbers.append(target_number)
                    ft_hit_positions.append(ft_position)
                    ft_hit_duos.append(ft_duo)
                    ft_hit_routes.append(
                        f"{ft_position} | {ft_duo}+{ft_suffix}="
                        f"{ft_generated} -> {target_number}"
                    )
        ft_hit_numbers = list(dict.fromkeys(ft_hit_numbers))
        ft_hit_positions = list(dict.fromkeys(ft_hit_positions))
        ft_hit_duos = list(dict.fromkeys(ft_hit_duos))
        ft_hit_routes = list(dict.fromkeys(ft_hit_routes))

        # Enjin berasingan: 2D daripada 3rd + missing + satu digit 1st.
        third_missing_keys, third_missing_routes = set(), []
        third_missing_positions, third_missing_first_positions = [], []
        for left, right in combinations(range(4), 2):
            position = f"{left + 1}+{right + 1}"
            duo = third[left] + third[right]
            for first_position, first_digit in enumerate(first, start=1):
                for missing_digit in carry_missing:
                    generated = f"{duo}{missing_digit}{first_digit}"
                    generated_key = unordered_digit_key4(generated)
                    third_missing_keys.add(generated_key)
                    for target_number, target_key in zip(actual_nums, actual_digit_keys):
                        if generated_key != target_key:
                            continue
                        third_missing_positions.append(position)
                        third_missing_first_positions.append(str(first_position))
                        third_missing_routes.append(
                            f"{position} x 1st-{first_position} | "
                            f"{duo}+{missing_digit}+{first_digit} | "
                            f"{generated} -> {target_number}"
                        )
        third_missing_hits = [
            number for number, target_key in zip(actual_nums, actual_digit_keys)
            if target_key in third_missing_keys
        ]
        third_missing_bridge_hits = [
            number for number, target_key in zip(actual_nums, actual_digit_keys)
            if target_key in third_missing_keys and target_key in v1_digit_keys
        ]
        third_missing_hits = list(dict.fromkeys(third_missing_hits))
        third_missing_bridge_hits = list(dict.fromkeys(third_missing_bridge_hits))
        third_missing_positions = list(dict.fromkeys(third_missing_positions))
        third_missing_first_positions = list(dict.fromkeys(third_missing_first_positions))
        third_missing_routes = list(dict.fromkeys(third_missing_routes))

        # Enjin berasingan: 2D daripada 3rd + dua digit 1st/2nd.
        first_second_suffixes = _occurrence_pairs_two_prizes(first, second)
        third_first_second_keys, third_first_second_routes = set(), []
        third_first_second_positions, third_first_second_duos = [], []
        for left, right in combinations(range(4), 2):
            position = f"{left + 1}+{right + 1}"
            duo = third[left] + third[right]
            for suffix in first_second_suffixes:
                generated = f"{duo}{suffix}"
                generated_key = unordered_digit_key4(generated)
                third_first_second_keys.add(generated_key)
                for target_number, target_key in zip(actual_nums, actual_digit_keys):
                    if generated_key != target_key:
                        continue
                    third_first_second_positions.append(position)
                    third_first_second_duos.append(duo)
                    third_first_second_routes.append(
                        f"{position} | {duo}+{suffix}={generated} -> {target_number}"
                    )
        third_first_second_hits = [
            number for number, target_key in zip(actual_nums, actual_digit_keys)
            if target_key in third_first_second_keys
        ]
        third_first_second_bridge_hits = [
            number for number, target_key in zip(actual_nums, actual_digit_keys)
            if target_key in third_first_second_keys and target_key in v2_digit_keys
        ]
        third_first_second_hits = list(dict.fromkeys(third_first_second_hits))
        third_first_second_bridge_hits = list(dict.fromkeys(third_first_second_bridge_hits))
        third_first_second_positions = list(dict.fromkeys(third_first_second_positions))
        third_first_second_duos = list(dict.fromkeys(third_first_second_duos))
        third_first_second_routes = list(dict.fromkeys(third_first_second_routes))

        def hit_state(values):
            return "PENDING" if status == "PENDING" else ("YES" if values else "NO")
        row = {
            "Source Draw": str(source.get("draw_no", idx)),
            "Source Result": f"{first} / {second} / {third}",
            "Next Draw": next_draw, "Next Result": next_result,
            "Bridge Count": len(v1_list), "Bridge List": " / ".join(v1_list),
            "Bridge Hit": hit_state(v1_hits), "Bridge Hit Number": " / ".join(v1_hits),
            "Bridge V2 Count": len(v2_list), "Bridge V2 List": " / ".join(v2_list),
            "Bridge V2 Hit": hit_state(v2_hits), "Bridge V2 Hit Number": " / ".join(v2_hits),
            "Bridge V2 2-Missing Hit": hit_state(missing_hits),
            "Bridge V2 2-Missing Hit Number": " / ".join(missing_hits),
            "Bridge V2 2-Existing Hit": hit_state(existing_hits),
            "Bridge V2 2-Existing Hit Number": " / ".join(existing_hits),
            "Hit": hit_state(union_hits), "Hit Number": " / ".join(union_hits),
            "1st2D+Missing2nd Candidate Count": len(first_missing_keys["Digit 2nd"]),
            "1st2D+Missing2nd Hit": hit_state(first_missing_hits["Digit 2nd"]),
            "1st2D+Missing2nd Hit Number": " / ".join(first_missing_hits["Digit 2nd"]),
            "1st2D+Missing2nd Bridge Hit": hit_state(first_missing_bridge_hits["Digit 2nd"]),
            "1st2D+Missing2nd Bridge Hit Number": " / ".join(first_missing_bridge_hits["Digit 2nd"]),
            "1st2D+Missing2nd Hit Positions": " / ".join(dict.fromkeys(first_missing_positions["Digit 2nd"])),
            "1st2D+Missing2nd Hit Routes": " || ".join(dict.fromkeys(first_missing_routes["Digit 2nd"])),
            "1st2D+Missing3rd Candidate Count": len(first_missing_keys["Digit 3rd"]),
            "1st2D+Missing3rd Hit": hit_state(first_missing_hits["Digit 3rd"]),
            "1st2D+Missing3rd Hit Number": " / ".join(first_missing_hits["Digit 3rd"]),
            "1st2D+Missing3rd Bridge Hit": hit_state(first_missing_bridge_hits["Digit 3rd"]),
            "1st2D+Missing3rd Bridge Hit Number": " / ".join(first_missing_bridge_hits["Digit 3rd"]),
            "1st2D+Missing3rd Hit Positions": " / ".join(dict.fromkeys(first_missing_positions["Digit 3rd"])),
            "1st2D+Missing3rd Hit Routes": " || ".join(dict.fromkeys(first_missing_routes["Digit 3rd"])),
            "1st2D+2nd3rd Candidate Count": len(first_second_third_keys),
            "1st2D+2nd3rd Hit": hit_state(first_second_third_hits),
            "1st2D+2nd3rd Hit Number": " / ".join(first_second_third_hits),
            "1st2D+2nd3rd Bridge Hit": hit_state(first_second_third_bridge_hits),
            "1st2D+2nd3rd Bridge Hit Number": " / ".join(first_second_third_bridge_hits),
            "1st2D+2nd3rd Hit Positions": " / ".join(dict.fromkeys(first_second_third_positions)),
            "1st2D+2nd3rd Hit Routes": " || ".join(dict.fromkeys(first_second_third_routes)),
            "2D+Missing Candidate Count": len(carry_candidate_keys),
            "2D+Missing Hit": hit_state(carry_hit_numbers),
            "2D+Missing Hit Number": " / ".join(carry_hit_numbers),
            "2D+Missing Hit 2D Positions": " / ".join(carry_hit_second_positions),
            "2D+Missing Hit 1st Positions": " / ".join(carry_hit_first_positions),
            "2D+Missing Hit Routes": " || ".join(carry_routes),
            "2D+1st3rd Candidate Count": len(ft_candidate_keys),
            "2D+1st3rd Hit": hit_state(ft_hit_numbers),
            "2D+1st3rd Hit Number": " / ".join(ft_hit_numbers),
            "2D+1st3rd Hit Positions": " / ".join(ft_hit_positions),
            "2D+1st3rd Hit 2D": " / ".join(ft_hit_duos),
            "2D+1st3rd Hit Routes": " || ".join(ft_hit_routes),
            "3rd2D+Missing Candidate Count": len(third_missing_keys),
            "3rd2D+Missing Hit": hit_state(third_missing_hits),
            "3rd2D+Missing Hit Number": " / ".join(third_missing_hits),
            "3rd2D+Missing Bridge Hit": hit_state(third_missing_bridge_hits),
            "3rd2D+Missing Bridge Hit Number": " / ".join(third_missing_bridge_hits),
            "3rd2D+Missing Hit Positions": " / ".join(third_missing_positions),
            "3rd2D+Missing Hit 1st Positions": " / ".join(third_missing_first_positions),
            "3rd2D+Missing Hit Routes": " || ".join(third_missing_routes),
            "3rd2D+1st2nd Candidate Count": len(third_first_second_keys),
            "3rd2D+1st2nd Hit": hit_state(third_first_second_hits),
            "3rd2D+1st2nd Hit Number": " / ".join(third_first_second_hits),
            "3rd2D+1st2nd Bridge Hit": hit_state(third_first_second_bridge_hits),
            "3rd2D+1st2nd Bridge Hit Number": " / ".join(third_first_second_bridge_hits),
            "3rd2D+1st2nd Hit Positions": " / ".join(third_first_second_positions),
            "3rd2D+1st2nd Hit 2D": " / ".join(third_first_second_duos),
            "3rd2D+1st2nd Hit Routes": " || ".join(third_first_second_routes),
        }
        rows.append(row)
        cache[key] = row
    try:
        cache_path.write_text(json.dumps({"version": "v31.54-first-routes", "rows": cache}, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass
    detail = pd.DataFrame(rows)
    valid = detail[detail.get("Hit", pd.Series(dtype=str)).astype(str).isin(["YES", "NO"])]
    total = len(valid)
    v1_yes = int(valid.get("Bridge Hit", pd.Series(dtype=str)).eq("YES").sum())
    v2_yes = int(valid.get("Bridge V2 Hit", pd.Series(dtype=str)).eq("YES").sum())
    union_yes = int(valid.get("Hit", pd.Series(dtype=str)).eq("YES").sum())
    miss_yes = int(valid.get("Bridge V2 2-Missing Hit", pd.Series(dtype=str)).eq("YES").sum())
    exist_yes = int(valid.get("Bridge V2 2-Existing Hit", pd.Series(dtype=str)).eq("YES").sum())
    summary = pd.DataFrame([
        {"Metric": "Tested source draws", "Value": total},
        {"Metric": "Pending latest draw", "Value": int(detail.get("Hit", pd.Series(dtype=str)).eq("PENDING").sum())},
        {"Metric": "Bridge V1 YES", "Value": v1_yes},
        {"Metric": "Bridge V1 Hit Rate %", "Value": round(v1_yes / total * 100, 1) if total else 0},
        {"Metric": "Bridge V2 YES", "Value": v2_yes},
        {"Metric": "Bridge V2 Hit Rate %", "Value": round(v2_yes / total * 100, 1) if total else 0},
        {"Metric": "V2 2-Missing YES", "Value": miss_yes},
        {"Metric": "V2 2-Existing YES", "Value": exist_yes},
        {"Metric": "Bridge V1 atau V2 Hit", "Value": union_yes},
        {"Metric": "Total Unique Hit Rate %", "Value": round(union_yes / total * 100, 1) if total else 0},
        {"Metric": "Elapsed Seconds", "Value": round(time.perf_counter() - t0, 3)},
    ])
    return summary, detail

def _first_existing_backtest_column(df, names):
    for name in names:
        if name in df.columns:
            return name
    return None


def build_clean_backtest_quick_review(detail_df):
    """Paparan ringkas keputusan draw serta hit Bridge V1 dan V2."""
    q = pd.DataFrame(index=detail_df.index)
    for target, choices in {
        "Source Draw": ["Source Draw"],
        "Source Result": ["Source Result"],
        "Next Draw": ["Next Draw"],
        "Next Result": ["Next Result"],
        "Bridge Hit No": ["Bridge Hit Number", "Bridge Hit No"],
        "Bridge V2 Hit No": ["Bridge V2 Hit Number", "Bridge V2 Hit No"],
        "1st 2nd+3rd Hit No": ["1st2D+2nd3rd Bridge Hit Number"],
        "2nd Missing Hit No": ["2D+Missing Hit Number"],
        "2nd 1st+3rd Hit No": ["2D+1st3rd Hit Number"],
        "3rd Missing Hit No": ["3rd2D+Missing Bridge Hit Number"],
        "3rd 1st+2nd Hit No": ["3rd2D+1st2nd Bridge Hit Number"],
    }.items():
        source = _first_existing_backtest_column(detail_df, choices)
        q[target] = detail_df[source].fillna("").astype(str) if source else ""
    first_missing_columns = [
        column for column in (
            "1st2D+Missing2nd Bridge Hit Number",
            "1st2D+Missing3rd Bridge Hit Number",
        ) if column in detail_df.columns
    ]
    if first_missing_columns:
        q.insert(6, "1st Missing Hit No", detail_df[first_missing_columns].apply(
            lambda row: " / ".join(dict.fromkeys(
                number.strip()
                for value in row.fillna("").astype(str)
                for number in value.split("/")
                if number.strip()
            )), axis=1,
        ))
    else:
        q.insert(6, "1st Missing Hit No", "")
    return q.reset_index(drop=True)


def build_clean_backtest_summary(detail_df):
    """Summary mesra pengguna untuk Bridge V1 dan V2."""
    hit_status = detail_df.get("Hit", pd.Series("", index=detail_df.index)).astype(str)
    valid_mask = hit_status.isin(["YES", "NO"])
    pending_mask = hit_status.eq("PENDING")
    valid = detail_df.loc[valid_mask].copy()
    total_draws = len(detail_df)
    completed = len(valid)
    pending = int(pending_mask.sum())

    bridge_hits = 0
    if "Bridge Hit" in valid.columns:
        bridge_hits = int(valid["Bridge Hit"].astype(str).eq("YES").sum())
    else:
        bridge_col = _first_existing_backtest_column(valid, ["Bridge Hit Number", "Bridge Hit No"])
        if bridge_col:
            bridge_hits = int(valid[bridge_col].fillna("").astype(str).str.strip().ne("").sum())

    bridge_v2_hits = int(valid.get("Bridge V2 Hit", pd.Series("", index=valid.index)).astype(str).eq("YES").sum())
    bridge_union_hits = int((
        valid.get("Bridge Hit", pd.Series("", index=valid.index)).astype(str).eq("YES")
        | valid.get("Bridge V2 Hit", pd.Series("", index=valid.index)).astype(str).eq("YES")
    ).sum())
    rows = [
        {"Metric": "Jumlah Draw", "Value": total_draws},
        {"Metric": "Draw Selesai", "Value": completed},
        {"Metric": "Draw Pending", "Value": pending},
        {"Metric": "Bridge Hit", "Value": bridge_hits},
        {"Metric": "Bridge Hit Rate %", "Value": round((bridge_hits / completed) * 100, 1) if completed else 0},
        {"Metric": "Bridge V2 Hit", "Value": bridge_v2_hits},
        {"Metric": "Bridge V2 Hit Rate %", "Value": round((bridge_v2_hits / completed) * 100, 1) if completed else 0},
        {"Metric": "Bridge V1 atau V2 Hit", "Value": bridge_union_hits},
        {"Metric": "Total Unique Hit Rate %", "Value": round((bridge_union_hits / completed) * 100, 1) if completed else 0},
    ]
    summary = pd.DataFrame(rows)
    summary["Value"] = summary["Value"].astype(str)
    return summary


def build_2d_missing_backtest_summary(detail_df):
    """Ringkasan engine 2D + missing tanpa dicampur dengan Bridge."""
    status = detail_df.get(
        "2D+Missing Hit", pd.Series("", index=detail_df.index)
    ).astype(str)
    valid = detail_df.loc[status.isin(["YES", "NO"])].copy()
    completed = len(valid)
    total_hits = int(valid.get(
        "2D+Missing Hit", pd.Series("", index=valid.index)
    ).astype(str).eq("YES").sum())
    rows = [
        {"Kategori": "Keseluruhan", "Laluan": "Mana-mana laluan", "Hit Draw": total_hits,
         "Draw Diuji": completed,
         "Hit Rate %": round(total_hits / completed * 100, 1) if completed else 0},
    ]

    def count_token(column, token):
        if column not in valid.columns:
            return 0
        return int(valid[column].fillna("").astype(str).apply(
            lambda value: token in [part.strip() for part in value.split("/")]
        ).sum())

    for left, right in combinations(range(4), 2):
        label = f"{left + 1}+{right + 1}"
        hits = count_token("2D+Missing Hit 2D Positions", label)
        rows.append({
            "Kategori": "Kedudukan 2D", "Laluan": label,
            "Hit Draw": hits, "Draw Diuji": completed,
            "Hit Rate %": round(hits / completed * 100, 1) if completed else 0,
        })
    for position in range(1, 5):
        hits = count_token("2D+Missing Hit 1st Positions", str(position))
        rows.append({
            "Kategori": "Kedudukan Digit 1st", "Laluan": str(position),
            "Hit Draw": hits, "Draw Diuji": completed,
            "Hit Rate %": round(hits / completed * 100, 1) if completed else 0,
        })
    return pd.DataFrame(rows)


def build_2d_missing_backtest_detail(detail_df):
    columns = [
        "Source Draw", "Source Result", "Next Draw", "Next Result",
        "2D+Missing Candidate Count", "2D+Missing Hit",
        "2D+Missing Hit Number", "2D+Missing Hit 2D Positions",
        "2D+Missing Hit 1st Positions", "2D+Missing Hit Routes",
    ]
    return detail_df.reindex(columns=columns).copy()


def build_first_third_backtest_summary(detail_df):
    """Ringkasan 2D 2nd + dua digit 1st/3rd secara berasingan."""
    status = detail_df.get(
        "2D+1st3rd Hit", pd.Series("", index=detail_df.index)
    ).astype(str)
    valid = detail_df.loc[status.isin(["YES", "NO"])].copy()
    completed = len(valid)
    total_hits = int(valid.get(
        "2D+1st3rd Hit", pd.Series("", index=valid.index)
    ).astype(str).eq("YES").sum())
    rows = [{
        "Kategori": "Keseluruhan", "Laluan": "Mana-mana laluan",
        "Hit Draw": total_hits, "Draw Diuji": completed,
        "Hit Rate %": round(total_hits / completed * 100, 1) if completed else 0,
    }]
    for left, right in combinations(range(4), 2):
        label = f"{left + 1}+{right + 1}"
        hits = int(valid.get(
            "2D+1st3rd Hit Positions", pd.Series("", index=valid.index)
        ).fillna("").astype(str).apply(
            lambda value: label in [part.strip() for part in value.split("/")]
        ).sum())
        rows.append({
            "Kategori": "Kedudukan 2D", "Laluan": label,
            "Hit Draw": hits, "Draw Diuji": completed,
            "Hit Rate %": round(hits / completed * 100, 1) if completed else 0,
        })
    return pd.DataFrame(rows)


def build_first_third_backtest_detail(detail_df):
    columns = [
        "Source Draw", "Source Result", "Next Draw", "Next Result",
        "2D+1st3rd Candidate Count", "2D+1st3rd Hit",
        "2D+1st3rd Hit Number", "2D+1st3rd Hit Positions",
        "2D+1st3rd Hit 2D", "2D+1st3rd Hit Routes",
    ]
    return detail_df.reindex(columns=columns).copy()


def build_third_route_backtest_summary(detail_df, prefix, label):
    """Ringkasan satu laluan 1st/3rd; tidak digabung dengan laluan lain."""
    status_column = f"{prefix} Hit"
    bridge_column = f"{prefix} Bridge Hit"
    status = detail_df.get(status_column, pd.Series("", index=detail_df.index)).astype(str)
    valid = detail_df.loc[status.isin(["YES", "NO"])].copy()
    completed = len(valid)
    raw_hits = int(valid.get(
        status_column, pd.Series("", index=valid.index)
    ).astype(str).eq("YES").sum())
    bridge_hits = int(valid.get(
        bridge_column, pd.Series("", index=valid.index)
    ).astype(str).eq("YES").sum())
    rows = [
        {"Kategori": "Keseluruhan", "Laluan": label,
         "Hit Draw": raw_hits, "Draw Diuji": completed,
         "Hit Rate %": round(raw_hits / completed * 100, 1) if completed else 0},
        {"Kategori": "Selepas Bridge", "Laluan": label,
         "Hit Draw": bridge_hits, "Draw Diuji": completed,
         "Hit Rate %": round(bridge_hits / completed * 100, 1) if completed else 0},
    ]
    position_column = f"{prefix} Hit Positions"
    prize_name = "1st" if str(prefix).startswith("1st") else "3rd"
    for left, right in combinations(range(4), 2):
        position = f"{left + 1}+{right + 1}"
        hits = int(valid.get(
            position_column, pd.Series("", index=valid.index)
        ).fillna("").astype(str).apply(
            lambda value: position in [part.strip() for part in value.split("/")]
        ).sum())
        rows.append({
            "Kategori": f"Kedudukan {prize_name} 2D", "Laluan": position,
            "Hit Draw": hits, "Draw Diuji": completed,
            "Hit Rate %": round(hits / completed * 100, 1) if completed else 0,
        })
    return pd.DataFrame(rows)


def build_third_route_backtest_detail(detail_df, prefix):
    columns = [
        "Source Draw", "Source Result", "Next Draw", "Next Result",
        f"{prefix} Candidate Count", f"{prefix} Hit", f"{prefix} Hit Number",
        f"{prefix} Bridge Hit", f"{prefix} Bridge Hit Number",
        f"{prefix} Hit Positions", f"{prefix} Hit Routes",
    ]
    return detail_df.reindex(columns=columns).copy()


def build_first_missing_backtest_summary(detail_df):
    """Satu summary untuk laluan Missing 1st; sumber digit kekal boleh diaudit."""
    frames = []
    for prefix, source in (
        ("1st2D+Missing2nd", "Digit 2nd"),
        ("1st2D+Missing3rd", "Digit 3rd"),
    ):
        frame = build_third_route_backtest_summary(
            detail_df, prefix, f"1st 2D + Missing + {source}"
        )
        frame.insert(1, "Sumber Digit", source)
        frames.append(frame)
    return pd.concat(frames, ignore_index=True)


def build_first_missing_backtest_detail(detail_df):
    """Satu detail untuk kedua-dua sumber digit laluan Missing 1st."""
    frames = []
    for prefix, source in (
        ("1st2D+Missing2nd", "Digit 2nd"),
        ("1st2D+Missing3rd", "Digit 3rd"),
    ):
        frame = build_third_route_backtest_detail(detail_df, prefix)
        frame = frame.rename(columns={
            column: column.replace(prefix, "1st2D+Missing")
            for column in frame.columns if prefix in column
        })
        frame.insert(4, "Sumber Digit", source)
        frames.append(frame)
    return pd.concat(frames, ignore_index=True)


def simple_backtest_excel_bytes(summary_df, detail_df):
    from io import BytesIO
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    quick_df = build_clean_backtest_quick_review(detail_df)
    clean_summary_df = build_clean_backtest_summary(detail_df)
    first_missing_summary_df = build_first_missing_backtest_summary(detail_df)
    first_missing_detail_df = build_first_missing_backtest_detail(detail_df)
    first_second_third_summary_df = build_third_route_backtest_summary(
        detail_df, "1st2D+2nd3rd", "1st 2D + Digit 2nd & 3rd"
    )
    first_second_third_detail_df = build_third_route_backtest_detail(
        detail_df, "1st2D+2nd3rd"
    )
    carry_summary_df = build_2d_missing_backtest_summary(detail_df)
    carry_detail_df = build_2d_missing_backtest_detail(detail_df)
    first_third_summary_df = build_first_third_backtest_summary(detail_df)
    first_third_detail_df = build_first_third_backtest_detail(detail_df)
    third_missing_summary_df = build_third_route_backtest_summary(
        detail_df, "3rd2D+Missing", "3rd 2D + Missing + Digit 1st"
    )
    third_missing_detail_df = build_third_route_backtest_detail(
        detail_df, "3rd2D+Missing"
    )
    third_first_second_summary_df = build_third_route_backtest_summary(
        detail_df, "3rd2D+1st2nd", "3rd 2D + Digit 1st & 2nd"
    )
    third_first_second_detail_df = build_third_route_backtest_detail(
        detail_df, "3rd2D+1st2nd"
    )

    # Enjin lama tidak lagi dipaparkan dalam Detail fail muat turun.
    obsolete_prefixes = ("Bridge V2 Selection", "Bridge V2 Top", "Bridge V3", "BDE ")
    clean_detail_df = detail_df.drop(
        columns=[c for c in detail_df.columns if str(c).startswith(obsolete_prefixes)],
        errors="ignore",
    )

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        quick_df.to_excel(writer, sheet_name="Quick Review", index=False)
        clean_summary_df.to_excel(writer, sheet_name="Summary", index=False)
        clean_detail_df.to_excel(writer, sheet_name="Detail", index=False)
        first_missing_summary_df.to_excel(
            writer, sheet_name="1st Missing Summary", index=False
        )
        first_missing_detail_df.to_excel(
            writer, sheet_name="1st Missing Detail", index=False
        )
        first_second_third_summary_df.to_excel(
            writer, sheet_name="1st 2nd3rd Summary", index=False
        )
        first_second_third_detail_df.to_excel(
            writer, sheet_name="1st 2nd3rd Detail", index=False
        )
        carry_summary_df.to_excel(writer, sheet_name="2nd Missing Summary", index=False)
        carry_detail_df.to_excel(writer, sheet_name="2nd Missing Detail", index=False)
        first_third_summary_df.to_excel(writer, sheet_name="2nd 1st3rd Summary", index=False)
        first_third_detail_df.to_excel(writer, sheet_name="2nd 1st3rd Detail", index=False)
        third_missing_summary_df.to_excel(writer, sheet_name="3rd Missing Summary", index=False)
        third_missing_detail_df.to_excel(writer, sheet_name="3rd Missing Detail", index=False)
        third_first_second_summary_df.to_excel(writer, sheet_name="3rd 1st2nd Summary", index=False)
        third_first_second_detail_df.to_excel(writer, sheet_name="3rd 1st2nd Detail", index=False)

        wb = writer.book
        navy = "17365D"
        pale_green = "EAF7EE"
        pale_blue = "EEF4FF"
        pale_gold = "FFF4D6"
        light_border = Side(style="thin", color="E5E7EB")

        quick_ws = wb["Quick Review"]
        quick_ws.freeze_panes = "A2"
        quick_ws.sheet_view.showGridLines = False
        quick_ws.sheet_view.zoomScale = 65
        quick_ws.auto_filter.ref = quick_ws.dimensions
        for cell in quick_ws[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        quick_ws.row_dimensions[1].height = 42
        for row in quick_ws.iter_rows(min_row=2, max_row=quick_ws.max_row):
            for cell in row:
                cell.border = Border(bottom=light_border)
                cell.alignment = Alignment(vertical="center")
                cell.number_format = "@"
            # Bridge V1/V2
            for cell in row[4:6]:
                cell.fill = PatternFill("solid", fgColor=pale_green)
                cell.font = Font(color="166534", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Dua laluan berasaskan 1st Prize
            for cell in row[6:8]:
                cell.fill = PatternFill("solid", fgColor="F3E8FF")
                cell.font = Font(color="6B21A8", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Dua laluan berasaskan 2nd Prize
            for cell in row[8:10]:
                cell.fill = PatternFill("solid", fgColor=pale_blue)
                cell.font = Font(color="1E3A8A", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Dua laluan berasaskan 3rd Prize
            for cell in row[10:12]:
                cell.fill = PatternFill("solid", fgColor=pale_gold)
                cell.font = Font(color="92400E", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for col, width in {
            "A": 10, "B": 20, "C": 10, "D": 20,
            "E": 13, "F": 13, "G": 15, "H": 15,
            "I": 15, "J": 15, "K": 15, "L": 15,
        }.items():
            quick_ws.column_dimensions[col].width = width

        summary_ws = wb["Summary"]
        summary_ws.sheet_view.showGridLines = False
        summary_ws.freeze_panes = "A2"
        summary_ws.auto_filter.ref = summary_ws.dimensions
        for cell in summary_ws[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_no in range(2, summary_ws.max_row + 1):
            summary_ws.cell(row_no, 1).border = Border(bottom=light_border)
            summary_ws.cell(row_no, 2).border = Border(bottom=light_border)
            summary_ws.cell(row_no, 2).font = Font(color=navy, bold=True)
            summary_ws.cell(row_no, 2).alignment = Alignment(horizontal="center")
            if row_no in (5, 6):
                fill = pale_green
            elif row_no in (7, 8):
                fill = pale_blue
            else:
                fill = "F8FAFC"
            summary_ws.cell(row_no, 1).fill = PatternFill("solid", fgColor=fill)
            summary_ws.cell(row_no, 2).fill = PatternFill("solid", fgColor=fill)
        summary_ws.column_dimensions["A"].width = 26
        summary_ws.column_dimensions["B"].width = 18

        detail_ws = wb["Detail"]
        detail_ws.freeze_panes = "A2"
        detail_ws.sheet_view.showGridLines = False
        detail_ws.sheet_view.zoomScale = 60
        detail_ws.auto_filter.ref = detail_ws.dimensions
        for cell in detail_ws[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        detail_ws.row_dimensions[1].height = 44
        for column_cells in detail_ws.columns:
            letter = column_cells[0].column_letter
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            detail_ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 24)

        for sheet_name in (
            "1st Missing Summary", "1st Missing Detail",
            "1st 2nd3rd Summary", "1st 2nd3rd Detail",
            "2nd Missing Summary", "2nd Missing Detail",
            "2nd 1st3rd Summary", "2nd 1st3rd Detail",
            "3rd Missing Summary", "3rd Missing Detail",
            "3rd 1st2nd Summary", "3rd 1st2nd Detail",
        ):
            ws = wb[sheet_name]
            ws.freeze_panes = "A2"
            ws.sheet_view.showGridLines = False
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor=navy)
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            ws.row_dimensions[1].height = 38
            ws.sheet_view.zoomScale = 75
            for column_cells in ws.columns:
                letter = column_cells[0].column_letter
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 28)

    output.seek(0)
    return output.getvalue()

# -----------------------------
# V31.6: Simple Backtest
# -----------------------------
st.markdown('<div class="rap-section-kicker">Backtest</div>', unsafe_allow_html=True)
with st.expander("🧪 Backtest Bridge V1 + V2", expanded=False):
    st.caption("Fast Backtest: keputusan draw lama dibaca daripada cache; hanya draw baharu atau berubah dikira semula.")
    bt_col1, bt_col2 = st.columns(2)
    with bt_col1:
        bt_draws = st.selectbox("Jumlah source draw untuk test", [10, 20, 30, 50, 100, 200, 300, 500], index=2, key="simple_bt_draws_v31_6")
    with bt_col2:
        st.write("")
        st.write("")
        run_bt = st.button("Run Backtest Turbo Lite", key="run_backtest_turbo_v31_7")

    if run_bt:
        with st.spinner("Backtest Bridge V1 + V2 sedang berjalan..."):
            bt_summary, bt_detail = run_backtest_bridge_dde_lite_v31_24_5(
                st.session_state.history, test_draws=bt_draws
            )

        if bt_detail.empty:
            st.warning("Backtest tidak menghasilkan data.")
        else:
            clean_bt_summary = build_clean_backtest_summary(bt_detail)
            st.subheader("Summary Bridge V1 + V2")
            st.dataframe(clean_bt_summary, hide_index=True, use_container_width=True)

            st.subheader("Detail")
            st.dataframe(bt_detail, hide_index=True, use_container_width=True)

            st.subheader("1st 2D + Missing — Digit 2nd vs Digit 3rd")
            st.dataframe(
                build_first_missing_backtest_summary(bt_detail),
                hide_index=True, use_container_width=True,
            )
            with st.expander("Lihat detail hit 1st 2D + Missing", expanded=False):
                st.dataframe(
                    build_first_missing_backtest_detail(bt_detail),
                    hide_index=True, use_container_width=True,
                )
            st.subheader("1st 2D + Digit 2nd & 3rd")
            st.dataframe(
                build_third_route_backtest_summary(
                    bt_detail, "1st2D+2nd3rd",
                    "1st 2D + Digit 2nd & 3rd",
                ), hide_index=True, use_container_width=True,
            )

            st.subheader("2nd 2D + Missing + Digit 1st")
            st.dataframe(
                build_2d_missing_backtest_summary(bt_detail),
                hide_index=True,
                use_container_width=True,
            )
            with st.expander("Lihat detail hit 2D + Missing", expanded=False):
                st.dataframe(
                    build_2d_missing_backtest_detail(bt_detail),
                    hide_index=True,
                    use_container_width=True,
                )

            st.subheader("2nd 2D + Digit 1st & 3rd")
            st.dataframe(
                build_first_third_backtest_summary(bt_detail),
                hide_index=True,
                use_container_width=True,
            )
            with st.expander("Lihat detail hit 2D + Digit 1st & 3rd", expanded=False):
                st.dataframe(
                    build_first_third_backtest_detail(bt_detail),
                    hide_index=True,
                    use_container_width=True,
                )

            st.subheader("3rd 2D + Missing + Digit 1st")
            st.dataframe(
                build_third_route_backtest_summary(
                    bt_detail, "3rd2D+Missing", "3rd 2D + Missing + Digit 1st"
                ), hide_index=True, use_container_width=True,
            )
            with st.expander("Lihat detail hit 3rd 2D + Missing", expanded=False):
                st.dataframe(
                    build_third_route_backtest_detail(bt_detail, "3rd2D+Missing"),
                    hide_index=True, use_container_width=True,
                )

            st.subheader("3rd 2D + Digit 1st & 2nd")
            st.dataframe(
                build_third_route_backtest_summary(
                    bt_detail, "3rd2D+1st2nd", "3rd 2D + Digit 1st & 2nd"
                ), hide_index=True, use_container_width=True,
            )
            with st.expander("Lihat detail hit 3rd 2D + Digit 1st & 2nd", expanded=False):
                st.dataframe(
                    build_third_route_backtest_detail(bt_detail, "3rd2D+1st2nd"),
                    hide_index=True, use_container_width=True,
                )

            bt_bytes = simple_backtest_excel_bytes(bt_summary, bt_detail)
            st.download_button(
                "Download Backtest Turbo Excel",
                data=bt_bytes,
                file_name="Rumah_A_Predictor_Backtest_Clean_Review_V31_23_3.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_backtest_turbo_v31_7"
            )


with st.form("predict_form"):
    st.markdown('<div class="rap-panel-title">Generate Analisis</div>', unsafe_allow_html=True)
    st.caption(
        "Keputusan terbaru telah diisi secara automatik. Tekan Generate untuk "
        "analisis Bridge berdasarkan nombor formula asal."
    )
    c1, c2, c3 = st.columns(3)
    first = c1.text_input("1st Prize", value=last["first"], max_chars=4)
    second = c2.text_input("2nd Prize", value=last["second"], max_chars=4)
    third = c3.text_input("3rd Prize", value=last["third"], max_chars=4)
    submitted = st.form_submit_button("Generate")

if submitted:
    st.success("Analisis Bridge berjaya dijana.")

    # -----------------------------
    # Bridge V1
    # -----------------------------
    st.markdown('<div class="engine-head engine-v1">Bridge V1</div>', unsafe_allow_html=True)
    st.caption(
        "Pair depan/tengah/belakang + 1 missing digit + 1 existing digit. "
        "Nombor formula asal dikekalkan."
    )

    bridge_df = pd.DataFrame()
    bridge_pair_df = pd.DataFrame()
    try:
        bridge_pair_df, bridge_df, bridge_text = build_bridge_model_v31_9(first, second, third)
        if bridge_df.empty:
            st.info("Bridge Model belum menghasilkan output.")
        else:
            st.caption(f"Jumlah Pilihan Bridge: {len(bridge_df)}")
            copy_button_clean("📋 Copy Bridge V1", bridge_text, "bridge_model_v31_9")
            with st.expander("Lihat Detail Bridge V1", expanded=False):
                st.markdown("**Base Pair**")
                st.dataframe(bridge_pair_df, hide_index=True, use_container_width=True)
                st.markdown("**Senarai Bridge**")
                st.dataframe(bridge_df, hide_index=True, use_container_width=True)
    except Exception as e:
        st.warning(f"Bridge Model belum dapat dipaparkan: {e}")

    # -----------------------------
    # Bridge Engine V2 - Pair + 2D Missing / Pair + 2D Existing
    # -----------------------------
    st.markdown('<div class="engine-head engine-v2">Bridge V2</div>', unsafe_allow_html=True)
    st.caption("Base pair + 2 digit missing, atau base pair + 2 digit existing. Digit pasangan mestilah berbeza.")
    bridge_v2_df = pd.DataFrame()
    try:
        bridge_v2_pair_df, bridge_v2_df, bridge_v2_text = build_bridge_engine_v2_pair_double_digit(first, second, third)
        if bridge_v2_df.empty:
            st.info("Bridge V2 belum menghasilkan output.")
        else:
            v2_missing_count = int(bridge_v2_df["Mode"].str.contains("2 Missing", regex=False).sum())
            v2_existing_count = int(bridge_v2_df["Mode"].str.contains("2 Existing", regex=False).sum())
            st.caption(
                f"Jumlah pilihan unik: {len(bridge_v2_df)} | "
                f"2 Missing: {v2_missing_count} | 2 Existing: {v2_existing_count}"
            )
            copy_button_clean("📋 Copy Bridge V2", bridge_v2_text, "bridge_engine_v2_double_digit")
            with st.expander("Lihat Detail Bridge V2", expanded=False):
                st.dataframe(bridge_v2_df, hide_index=True, use_container_width=True)
    except Exception as e:
        st.warning(f"Bridge Engine V2 belum dapat dipaparkan: {e}")

    # -----------------------------
    # Route Signal: satu pilihan ringkas berdasarkan tiga hadiah.
    # -----------------------------
    st.markdown(
        '<div class="engine-head engine-support">Route Signal</div>',
        unsafe_allow_html=True,
    )
    try:
        first_route_signal = build_1st_route_signal(
            st.session_state.history, first, second, third, lookback=100
        )
        route_signal = build_2d_route_signal(
            st.session_state.history, first, second, third, lookback=100
        )
        third_route_signal = build_3rd_route_signal(
            st.session_state.history, first, second, third, lookback=100
        )
        first_route_version = {
            "1st 2D + Missing": "V1",
            "1st 2D + 2nd & 3rd": "V2",
        }.get(first_route_signal["signal"])
        second_route_version = {
            "2D + Missing": "V1",
            "2D + 1st & 3rd": "V2",
        }.get(route_signal["signal"])
        third_route_version = {
            "3rd 2D + Missing": "V1",
            "3rd 2D + 1st & 2nd": "V2",
        }.get(third_route_signal["signal"])
        route_versions = [
            version for version in (
                first_route_version, second_route_version, third_route_version
            ) if version
        ]
        route_version_counts = Counter(route_versions)
        if route_version_counts:
            selected_version, selected_count = route_version_counts.most_common(1)[0]
            match_route_text = (
                f"Triple Match {selected_version}" if selected_count == 3
                else f"Double Match {selected_version}" if selected_count == 2
                else f"Single Match {selected_version}"
            )
        else:
            match_route_text = "Seimbang"
        st.markdown(f"**Laluan Pilihan:** {match_route_text}")
        with st.expander("Lihat asas Route Signal", expanded=False):
            st.markdown(f"1st: {first_route_signal['signal']}")
            st.markdown(f"2nd: {route_signal['signal']}")
            st.markdown(f"3rd: {third_route_signal['signal']}")

        v1_route_lookup = {
            _key4(number): _pad4(number)
            for number in bridge_df.get("No", pd.Series(dtype=str)).astype(str)
        } if bridge_df is not None and not bridge_df.empty else {}
        v2_route_lookup = {
            _key4(number): _pad4(number)
            for number in bridge_v2_df.get("No", pd.Series(dtype=str)).astype(str)
        } if bridge_v2_df is not None and not bridge_v2_df.empty else {}

        def bridge_route_numbers(families, bridge="both"):
            numbers = []
            seen = set()
            for family in families:
                if family in seen:
                    continue
                if bridge == "V1":
                    bridge_number = v1_route_lookup.get(family)
                elif bridge == "V2":
                    bridge_number = v2_route_lookup.get(family)
                else:
                    bridge_number = v1_route_lookup.get(family) or v2_route_lookup.get(family)
                if not bridge_number:
                    continue
                seen.add(family)
                numbers.append(bridge_number)
            return numbers

        first_missing_route = build_1st_missing_digit_engine(
            st.session_state.history, first, second, third,
            bridge_df, bridge_v2_df, lookback=100,
        )
        first_pair_route = build_1st_second_third_pair_engine(
            st.session_state.history, first, second, third, lookback=100,
        )
        first_missing_families = [
            _key4(number) for number in first_missing_route["selected"].get(
                "No Terhasil", pd.Series(dtype=str)
            ).astype(str)
        ]
        first_pair_families = [
            _key4(number) for number in first_pair_route["selected"].get(
                "No Terhasil", pd.Series(dtype=str)
            ).astype(str)
        ]
        first_missing_numbers = bridge_route_numbers(first_missing_families, "V1")
        first_pair_numbers = bridge_route_numbers(first_pair_families, "V2")
        first_missing_focus, first_missing_coverage = (
            first_missing_numbers[:5], first_missing_numbers[:10]
        )
        first_pair_focus, first_pair_coverage = (
            first_pair_numbers[:5], first_pair_numbers[:10]
        )

        missing_route = build_2d_missing_first_digit_engine(
            st.session_state.history, first, second, third,
            bridge_df, bridge_v2_df, lookback=100,
        )
        current_second = _pad4(second)
        current_first = _pad4(first)
        missing_families = []
        for _, audit_row in missing_route["joint_audit"].iterrows():
            left, right = [
                int(value) - 1
                for value in str(audit_row["Kedudukan 2D"]).split("+")
            ]
            duo = current_second[left] + current_second[right]
            first_digit = current_first[
                int(audit_row["Kedudukan Digit 1st"]) - 1
            ]
            for missing in missing_route["missing"]:
                missing_families.append(
                    _key4(f"{duo}{missing}{first_digit}")
                )
        missing_numbers = bridge_route_numbers(missing_families)
        missing_focus = missing_numbers[:5]
        missing_coverage = missing_numbers[:10]

        first_third_extended = build_first_third_extended_audit(
            st.session_state.history, first, second, third, lookback=100
        )
        first_third_audit = first_third_extended["joint_audit"]
        first_third_families = [
            _key4(number)
            for number in first_third_audit.get(
                "Full No", pd.Series(dtype=str)
            ).astype(str)
        ]
        first_third_numbers = bridge_route_numbers(first_third_families)
        first_third_focus = first_third_numbers[:5]
        first_third_coverage = first_third_numbers[:10]

        third_missing_route = build_3rd_missing_first_digit_engine(
            st.session_state.history, first, second, third,
            bridge_df, bridge_v2_df, lookback=100,
        )
        current_third = _pad4(third)
        third_missing_families = []
        for _, audit_row in third_missing_route["joint_audit"].iterrows():
            left, right = [
                int(value) - 1
                for value in str(audit_row["Kedudukan 3rd 2D"]).split("+")
            ]
            duo = current_third[left] + current_third[right]
            first_digit = current_first[
                int(audit_row["Kedudukan Digit 1st"]) - 1
            ]
            for missing in third_missing_route["missing"]:
                third_missing_families.append(_key4(f"{duo}{missing}{first_digit}"))
        third_missing_numbers = bridge_route_numbers(third_missing_families, "V1")
        third_missing_focus = third_missing_numbers[:5]
        third_missing_coverage = third_missing_numbers[:10]

        third_first_second_route = build_3rd_first_second_pair_engine(
            st.session_state.history, first, second, third, lookback=100,
        )
        third_first_second_families = [
            _key4(number)
            for number in third_first_second_route["selected"].get(
                "No Terhasil", pd.Series(dtype=str)
            ).astype(str)
        ]
        third_first_second_numbers = bridge_route_numbers(
            third_first_second_families, "V2"
        )
        third_first_second_focus = third_first_second_numbers[:5]
        third_first_second_coverage = third_first_second_numbers[:10]

        # Butiran laluan lama disimpan sebagai rujukan dalaman dan tidak dirender.
        _hidden_route_details = """
        st.markdown("**2D + Missing**")
        st.markdown(
            f"Focus 5: {' / '.join(missing_focus) or 'Tiada'}"
        )
        st.markdown(
            f"Coverage 10: {' / '.join(missing_coverage) or 'Tiada'}"
        )
        missing_copy = (
            "Rumah A Predictor - Route Signal 2D + Missing\n\n"
            f"Focus 5: {' / '.join(missing_focus) or 'Tiada'}\n"
            f"Coverage 10: {' / '.join(missing_coverage) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy 2D + Missing Signal",
            missing_copy,
            "copy_route_missing",
        )

        st.markdown("**2D + 1st & 3rd**")
        st.markdown(
            f"Focus 5: {' / '.join(first_third_focus) or 'Tiada'}"
        )
        st.markdown(
            f"Coverage 10: {' / '.join(first_third_coverage) or 'Tiada'}"
        )
        first_third_copy = (
            "Rumah A Predictor - Route Signal 2D + 1st & 3rd\n\n"
            f"Focus 5: {' / '.join(first_third_focus) or 'Tiada'}\n"
            f"Coverage 10: {' / '.join(first_third_coverage) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy 2D + 1st & 3rd Signal",
            first_third_copy,
            "copy_route_first_third",
        )

        st.markdown("**3rd 2D + Missing**")
        st.markdown(f"Focus 5: {' / '.join(third_missing_focus) or 'Tiada'}")
        st.markdown(f"Coverage 10: {' / '.join(third_missing_coverage) or 'Tiada'}")
        third_missing_copy = (
            "Rumah A Predictor - Route Signal 3rd 2D + Missing\n\n"
            f"Focus 5: {' / '.join(third_missing_focus) or 'Tiada'}\n"
            f"Coverage 10: {' / '.join(third_missing_coverage) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy 3rd 2D + Missing Signal", third_missing_copy,
            "copy_route_third_missing",
        )

        st.markdown("**3rd 2D + 1st & 2nd**")
        st.markdown(f"Focus 5: {' / '.join(third_first_second_focus) or 'Tiada'}")
        st.markdown(f"Coverage 10: {' / '.join(third_first_second_coverage) or 'Tiada'}")
        third_first_second_copy = (
            "Rumah A Predictor - Route Signal 3rd 2D + 1st & 2nd\n\n"
            f"Focus 5: {' / '.join(third_first_second_focus) or 'Tiada'}\n"
            f"Coverage 10: {' / '.join(third_first_second_coverage) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy 3rd 2D + 1st & 2nd Signal", third_first_second_copy,
            "copy_route_third_first_second",
        )

        """
        # Match diletakkan terus selepas Route Signal.
        st.markdown(
            '<div class="engine-head engine-support">Triple Match</div>',
            unsafe_allow_html=True,
        )

        def _route_family_set(frame):
            if frame is None or frame.empty or "No Terhasil" not in frame.columns:
                return set()
            return {
                _key4(number)
                for number in frame["No Terhasil"].dropna().astype(str)
            }

        def _ordered_route_numbers(families, lookup):
            wanted = set(families)
            return [
                number for family, number in lookup.items()
                if family in wanted
            ]

        first_missing_all_families = _route_family_set(first_missing_route["all"])
        first_pair_all_families = _route_family_set(first_pair_route["all"])
        second_missing_families = _route_family_set(missing_route["all"])
        third_missing_all_families = _route_family_set(third_missing_route["all"])
        first_third_all_engine = build_2d_first_third_pair_engine(
            st.session_state.history, first, second, third, lookback=100
        )
        second_first_third_families = _route_family_set(
            first_third_all_engine["all"]
        )
        third_first_second_all_families = _route_family_set(
            third_first_second_route["all"]
        )

        v1_bridge_families = set(v1_route_lookup)
        v1_first = v1_bridge_families & first_missing_all_families
        v1_second = v1_bridge_families & second_missing_families
        v1_third = v1_bridge_families & third_missing_all_families
        v1_support = Counter(
            family for route_set in (v1_first, v1_second, v1_third)
            for family in route_set
        )
        v1_triple = {family for family, count in v1_support.items() if count == 3}
        v1_double_only = {family for family, count in v1_support.items() if count == 2}

        v2_bridge_families = set(v2_route_lookup)
        v2_first = v2_bridge_families & first_pair_all_families
        v2_second = v2_bridge_families & second_first_third_families
        v2_third = v2_bridge_families & third_first_second_all_families
        v2_support = Counter(
            family for route_set in (v2_first, v2_second, v2_third)
            for family in route_set
        )
        v2_triple = {family for family, count in v2_support.items() if count == 3}
        v2_double_only = {family for family, count in v2_support.items() if count == 2}

        v1_triple_numbers = _ordered_route_numbers(v1_triple, v1_route_lookup)
        v2_triple_numbers = _ordered_route_numbers(v2_triple, v2_route_lookup)
        v1_double_numbers = _ordered_route_numbers(v1_double_only, v1_route_lookup)
        v2_double_numbers = _ordered_route_numbers(v2_double_only, v2_route_lookup)

        def _double_route_overlap(route_numbers, double_families):
            return list(dict.fromkeys(
                _pad4(number)
                for number in route_numbers
                if _key4(number) in double_families
            ))

        v1_double_focus = _double_route_overlap(
            first_missing_focus + missing_focus + third_missing_focus,
            v1_double_only,
        )
        v1_double_coverage = _double_route_overlap(
            first_missing_coverage + missing_coverage + third_missing_coverage,
            v1_double_only,
        )
        v2_double_focus = _double_route_overlap(
            first_pair_focus + first_third_focus + third_first_second_focus,
            v2_double_only,
        )
        v2_double_coverage = _double_route_overlap(
            first_pair_coverage + first_third_coverage + third_first_second_coverage,
            v2_double_only,
        )
        v1_route_focus_both = set.intersection(*[
            {_key4(number) for number in values}
            for values in (first_missing_focus, missing_focus, third_missing_focus)
        ])
        v1_route_coverage_both = set.intersection(*[
            {_key4(number) for number in values}
            for values in (
                first_missing_coverage, missing_coverage, third_missing_coverage
            )
        ])
        v2_route_focus_both = set.intersection(*[
            {_key4(number) for number in values}
            for values in (first_pair_focus, first_third_focus, third_first_second_focus)
        ])
        v2_route_coverage_both = set.intersection(*[
            {_key4(number) for number in values}
            for values in (
                first_pair_coverage, first_third_coverage,
                third_first_second_coverage,
            )
        ])
        # Route Signal untuk Triple kekal berasingan mengikut versi Bridge.
        # Ini sama dengan paparan Double Match dan mengelakkan V1/V2 bercampur.
        v1_triple_focus = _ordered_route_numbers(
            v1_triple & v1_route_focus_both, v1_route_lookup
        )
        v1_triple_coverage = _ordered_route_numbers(
            v1_triple & v1_route_coverage_both, v1_route_lookup
        )
        v2_triple_focus = _ordered_route_numbers(
            v2_triple & v2_route_focus_both, v2_route_lookup
        )
        v2_triple_coverage = _ordered_route_numbers(
            v2_triple & v2_route_coverage_both, v2_route_lookup
        )

        if match_route_text == "Triple Match V1":
            route_choice_numbers = _ordered_route_numbers(
                v1_triple & v1_route_coverage_both, v1_route_lookup
            )
        elif match_route_text == "Triple Match V2":
            route_choice_numbers = _ordered_route_numbers(
                v2_triple & v2_route_coverage_both, v2_route_lookup
            )
        elif "Double Match V1" in match_route_text and "Double Match V2" not in match_route_text:
            route_choice_numbers = v1_double_coverage
        elif "Double Match V2" in match_route_text and "Double Match V1" not in match_route_text:
            route_choice_numbers = v2_double_coverage
        else:
            route_choice_numbers = list(dict.fromkeys(
                v1_double_coverage + v2_double_coverage
            ))

        st.markdown("**Triple Match V1**")
        st.markdown(f"**Jumlah Pilihan:** {len(v1_triple_numbers)}")
        st.markdown(f"{' / '.join(v1_triple_numbers) or 'Tiada'}")
        st.markdown(
            f"**Route Focus ({len(v1_triple_focus)}):** "
            f"{' / '.join(v1_triple_focus) or 'Tiada'}"
        )
        st.markdown(
            f"**Route Coverage ({len(v1_triple_coverage)}):** "
            f"{' / '.join(v1_triple_coverage) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy Triple Match V1",
            "Rumah A Predictor - Triple Match V1\n\n"
            "Bridge V1 + 1st Missing + 2nd Missing + 3rd Missing\n"
            f"Jumlah Pilihan: {len(v1_triple_numbers)}\n"
            f"{' / '.join(v1_triple_numbers) or 'Tiada'}\n\n"
            f"Route Focus ({len(v1_triple_focus)}): "
            f"{' / '.join(v1_triple_focus) or 'Tiada'}\n"
            f"Route Coverage ({len(v1_triple_coverage)}): "
            f"{' / '.join(v1_triple_coverage) or 'Tiada'}",
            "copy_triple_match_v1_top",
        )

        st.markdown("**Triple Match V2**")
        st.markdown(f"**Jumlah Pilihan:** {len(v2_triple_numbers)}")
        st.markdown(f"{' / '.join(v2_triple_numbers) or 'Tiada'}")
        st.markdown(
            f"**Route Focus ({len(v2_triple_focus)}):** "
            f"{' / '.join(v2_triple_focus) or 'Tiada'}"
        )
        st.markdown(
            f"**Route Coverage ({len(v2_triple_coverage)}):** "
            f"{' / '.join(v2_triple_coverage) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy Triple Match V2",
            "Rumah A Predictor - Triple Match V2\n\n"
            "Bridge V2 + 1st 2nd & 3rd + 2nd 1st & 3rd + 3rd 1st & 2nd\n"
            f"Jumlah Pilihan: {len(v2_triple_numbers)}\n"
            f"{' / '.join(v2_triple_numbers) or 'Tiada'}\n\n"
            f"Route Focus ({len(v2_triple_focus)}): "
            f"{' / '.join(v2_triple_focus) or 'Tiada'}\n"
            f"Route Coverage ({len(v2_triple_coverage)}): "
            f"{' / '.join(v2_triple_coverage) or 'Tiada'}",
            "copy_triple_match_v2_top",
        )

        st.markdown("**Double Match V1**")
        st.markdown(f"**Jumlah Pilihan:** {len(v1_double_numbers)}")
        st.markdown(f"{' / '.join(v1_double_numbers) or 'Tiada'}")
        st.markdown(
            f"**Route Focus ({len(v1_double_focus)}):** "
            f"{' / '.join(v1_double_focus) or 'Tiada'}"
        )
        st.markdown(
            f"**Route Coverage ({len(v1_double_coverage)}):** "
            f"{' / '.join(v1_double_coverage) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy Double Match V1",
            "Rumah A Predictor - Double Match V1\n\n"
            f"Jumlah Pilihan: {len(v1_double_numbers)}\n"
            f"{' / '.join(v1_double_numbers) or 'Tiada'}\n\n"
            f"Route Focus ({len(v1_double_focus)}): "
            f"{' / '.join(v1_double_focus) or 'Tiada'}\n"
            f"Route Coverage ({len(v1_double_coverage)}): "
            f"{' / '.join(v1_double_coverage) or 'Tiada'}",
            "copy_double_match_v1_top",
        )

        st.markdown("**Double Match V2**")
        st.markdown(f"**Jumlah Pilihan:** {len(v2_double_numbers)}")
        st.markdown(f"{' / '.join(v2_double_numbers) or 'Tiada'}")
        st.markdown(
            f"**Route Focus ({len(v2_double_focus)}):** "
            f"{' / '.join(v2_double_focus) or 'Tiada'}"
        )
        st.markdown(
            f"**Route Coverage ({len(v2_double_coverage)}):** "
            f"{' / '.join(v2_double_coverage) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy Double Match V2",
            "Rumah A Predictor - Double Match V2\n\n"
            f"Jumlah Pilihan: {len(v2_double_numbers)}\n"
            f"{' / '.join(v2_double_numbers) or 'Tiada'}\n\n"
            f"Route Focus ({len(v2_double_focus)}): "
            f"{' / '.join(v2_double_focus) or 'Tiada'}\n"
            f"Route Coverage ({len(v2_double_coverage)}): "
            f"{' / '.join(v2_double_coverage) or 'Tiada'}",
            "copy_double_match_v2_top",
        )
    except Exception as e:
        st.warning(f"Route Signal belum dapat dipaparkan: {e}")

    # -----------------------------
    # Top 2D Engine - 1st Prize
    # -----------------------------
    st.markdown(
        '<div class="engine-head engine-support">1st 2D Engine</div>',
        unsafe_allow_html=True,
    )
    try:
        first_missing_engine = first_missing_route
        source_text = " / ".join(first_missing_engine["selected_sources"]) or "Tiada"
        position_text = " / ".join(first_missing_engine["selected_positions"]) or "Tiada"
        digit_text = " / ".join(dict.fromkeys(
            f"{item['Sumber Digit']} posisi {item['Kedudukan Digit']}"
            for item in first_missing_engine["selected_digits"]
        )) or "Tiada"
        first_v1_numbers = list(dict.fromkeys(
            number for number in first_missing_engine["selected"].get(
                "Bridge V1", pd.Series(dtype=str)
            ).astype(str) if number.strip()
        ))
        st.markdown("**1st 2D + Missing**")
        st.markdown(
            f"1st 2D: {position_text} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Sumber digit: {source_text}"
        )
        st.markdown(f"**Pilihan Bridge V1 ({len(first_v1_numbers)}):** "
                    f"{' / '.join(first_v1_numbers) or 'Tiada'}")
        copy_button_clean(
            "📋 Copy 1st 2D + Missing",
            "Rumah A Predictor - 1st 2D + Missing\n\n"
            f"1st 2D: {position_text}\nSumber digit: {source_text}\n"
            f"Kedudukan digit: {digit_text}\n"
            f"Jumlah Pilihan Bridge V1: {len(first_v1_numbers)}\n"
            f"{' / '.join(first_v1_numbers) or 'Tiada'}",
            "copy_first_2d_missing",
        )
        first_missing_pair_df = first_missing_engine["all"].copy()
        if not first_missing_pair_df.empty:
            first_missing_pair_df = first_missing_pair_df[
                first_missing_pair_df["Bridge V1"].fillna("").astype(str).str.strip().ne("")
            ].copy()
            for pair_index, pair_value in enumerate(
                dict.fromkeys(first_missing_pair_df["1st 2D"].astype(str)), start=1
            ):
                pair_df = first_missing_pair_df[
                    first_missing_pair_df["1st 2D"].astype(str).eq(pair_value)
                ].copy()
                pair_numbers = list(dict.fromkeys(
                    pair_df["Bridge V1"].astype(str).tolist()
                ))
                pair_positions = " / ".join(dict.fromkeys(
                    pair_df["Kedudukan 1st 2D"].astype(str).tolist()
                ))
                with st.expander(
                    f"Pair {pair_value} — {pair_positions} "
                    f"({len(pair_numbers)} pilihan Bridge)", expanded=False,
                ):
                    st.markdown(f"**Pilihan Bridge V1:** {' / '.join(pair_numbers)}")
                    copy_button_clean(
                        f"📋 Copy Pair {pair_value}",
                        "Rumah A Predictor - 1st 2D + Missing — Bridge V1\n\n"
                        f"Pair: {pair_value}\nKedudukan: {pair_positions}\n"
                        f"Pilihan (Total: {len(pair_numbers)}):\n"
                        f"{' / '.join(pair_numbers)}",
                        f"copy_first_missing_pair_{pair_index}_{pair_value}",
                    )
        with st.expander("Lihat audit 1st 2D + Missing", expanded=False):
            st.markdown("**Digit 2nd berbanding Digit 3rd**")
            st.dataframe(
                first_missing_engine["source_audit"], hide_index=True,
                use_container_width=True,
            )
            st.markdown("**Gabungan kedudukan**")
            st.dataframe(
                first_missing_engine["joint_audit"], hide_index=True,
                use_container_width=True,
            )

        first_pair_engine = first_pair_route
        first_pair_positions = " / ".join(
            first_pair_engine["selected_positions"]
        ) or "Tiada"
        v2_lookup = {
            _key4(number): _pad4(number)
            for number in bridge_v2_df.get("No", pd.Series(dtype=str)).astype(str)
        } if bridge_v2_df is not None and not bridge_v2_df.empty else {}
        first_v2_numbers = list(dict.fromkeys(
            v2_lookup.get(_key4(number), "")
            for number in first_pair_engine["selected"].get(
                "No Terhasil", pd.Series(dtype=str)
            ).astype(str)
            if v2_lookup.get(_key4(number), "")
        ))
        st.markdown("**1st 2D + Digit 2nd & 3rd**")
        st.markdown(f"1st 2D: {first_pair_positions}")
        st.markdown(f"**Pilihan Bridge V2 ({len(first_v2_numbers)}):** "
                    f"{' / '.join(first_v2_numbers) or 'Tiada'}")
        copy_button_clean(
            "📋 Copy 1st 2D + 2nd & 3rd",
            "Rumah A Predictor - 1st 2D + Digit 2nd & 3rd\n\n"
            f"1st 2D: {first_pair_positions}\n"
            f"Jumlah Pilihan Bridge V2: {len(first_v2_numbers)}\n"
            f"{' / '.join(first_v2_numbers) or 'Tiada'}",
            "copy_first_2d_second_third",
        )
        first_pair_all_df = first_pair_engine["all"].copy()
        if not first_pair_all_df.empty:
            first_pair_all_df["Bridge V2"] = first_pair_all_df["No Terhasil"].map(
                lambda number: v2_lookup.get(_key4(number), "")
            )
            first_pair_all_df = first_pair_all_df[
                first_pair_all_df["Bridge V2"].astype(str).str.strip().ne("")
            ].copy()
            for pair_index, pair_value in enumerate(
                dict.fromkeys(first_pair_all_df["1st 2D"].astype(str)), start=1
            ):
                pair_df = first_pair_all_df[
                    first_pair_all_df["1st 2D"].astype(str).eq(pair_value)
                ].copy()
                pair_numbers = list(dict.fromkeys(pair_df["Bridge V2"].astype(str)))
                pair_positions = " / ".join(dict.fromkeys(
                    pair_df["Kedudukan 1st 2D"].astype(str).tolist()
                ))
                with st.expander(
                    f"Pair {pair_value} — {pair_positions} "
                    f"({len(pair_numbers)} pilihan Bridge)", expanded=False,
                ):
                    st.markdown(f"**Pilihan Bridge V2:** {' / '.join(pair_numbers)}")
                    copy_button_clean(
                        f"📋 Copy Pair {pair_value}",
                        "Rumah A Predictor - 1st 2D + 2nd & 3rd — Bridge V2\n\n"
                        f"Pair: {pair_value}\nKedudukan: {pair_positions}\n"
                        f"Pilihan (Total: {len(pair_numbers)}):\n"
                        f"{' / '.join(pair_numbers)}",
                        f"copy_first_pair_route_{pair_index}_{pair_value}",
                    )
        with st.expander("Lihat audit 1st 2D + Digit 2nd & 3rd", expanded=False):
            st.dataframe(
                first_pair_engine["audit"], hide_index=True,
                use_container_width=True,
            )
    except Exception as e:
        st.warning(f"Top 2D Engine belum dapat dipaparkan: {e}")

    # -----------------------------
    # 2D + Missing + Digit 1st
    # -----------------------------
    st.markdown(
        '<div class="engine-head engine-support">2nd 2D Engine</div>',
        unsafe_allow_html=True,
    )
    st.markdown("**2nd 2D + Missing + Digit 1st**")
    try:
        route_engine = missing_route
        all_route_df = route_engine["all"]
        selected_second_positions = set(route_engine["selected_second"])
        selected_route_df = all_route_df[
            all_route_df["Kedudukan 2D"].apply(
                lambda value: any(
                    position in [part.strip() for part in str(value).split("/")]
                    for position in selected_second_positions
                )
            )
        ].copy() if not all_route_df.empty else all_route_df.copy()
        selected_numbers = (
            list(dict.fromkeys(selected_route_df["No Terhasil"].astype(str).tolist()))
            if not selected_route_df.empty else []
        )
        selected_first_positions = {1, 2, 3, 4}
        all_pair_view = all_route_df[
            all_route_df["Kedudukan Digit 1st"].isin(selected_first_positions)
        ].copy() if not all_route_df.empty else all_route_df.copy()
        all_numbers = (
            list(dict.fromkeys(all_pair_view["No Terhasil"].astype(str).tolist()))
            if not all_pair_view.empty else []
        )
        selected_second_text = " / ".join(route_engine["selected_second"]) or "Tiada"
        selected_first_text = "1 / 2 / 3 / 4"
        missing_text = " / ".join(route_engine["missing"]) or "Tiada"

        # Padankan setiap laluan audit dengan nombor draw semasa supaya
        # keputusan audit boleh terus dibaca tanpa mengira secara manual.
        current_second = _pad4(second)
        current_first = _pad4(first)
        joint_view = route_engine["joint_audit"].copy()
        joint_current_numbers = []
        for _, audit_row in joint_view.iterrows():
            position_parts = [
                int(value) - 1
                for value in str(audit_row["Kedudukan 2D"]).split("+")
            ]
            duo = "".join(current_second[position] for position in position_parts)
            first_digit = current_first[
                int(audit_row["Kedudukan Digit 1st"]) - 1
            ]
            numbers = [
                f"{duo}{missing}{first_digit}"
                for missing in route_engine["missing"]
            ]
            joint_current_numbers.append(" / ".join(numbers) or "Tiada")
        joint_view["Pilihan Semasa"] = joint_current_numbers

        best_joint_hits = (
            int(joint_view["Hit Draw"].max()) if not joint_view.empty else 0
        )
        top_joint_view = joint_view[
            joint_view["Hit Draw"].eq(best_joint_hits)
        ].copy()
        top_joint_numbers = list(dict.fromkeys(
            number
            for value in top_joint_view.get(
                "Pilihan Semasa", pd.Series(dtype=str)
            ).astype(str)
            for number in value.split(" / ")
            if number and number != "Tiada"
        ))

        st.markdown(
            f"**2D utama:** {selected_second_text} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"**Missing:** {missing_text}"
        )
        st.markdown(
            f"**Pilihan utama:** {' / '.join(top_joint_numbers) or 'Tiada'}"
        )

        route_text = (
            "Rumah A Predictor - 2D + Missing + Digit 1st\n\n"
            f"2nd Prize: {_pad4(second)}\n"
            f"1st Prize: {_pad4(first)}\n"
            "Kedudukan 2D: Semua pair\n"
            f"Kedudukan digit 1st: {selected_first_text}\n"
            f"Missing: {missing_text}\n\n"
            f"Pilihan (Total: {len(all_numbers)}):\n"
            f"{' / '.join(all_numbers) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy Semua 2D + Missing",
            route_text,
            "copy_2d_missing_first_digit",
        )

        # Semua pair unik daripada enam kedudukan 2nd Prize. Pair yang sama
        # digabungkan, tetapi asal kedudukannya masih dipaparkan.
        # Senarai mentah semua pair tidak dipaparkan. Hasilnya masih dikira dan
        # tersedia melalui Bridge Selection yang lebih ringkas di bawah.
        if False and not all_pair_view.empty:
            position_priority = {
                str(row["Kedudukan 2D"]): priority
                for priority, (_, row) in enumerate(
                    route_engine["second_audit"].iterrows(), start=1
                )
            }
            pair_order = []
            for pair_value in all_pair_view["2D"].astype(str):
                if pair_value not in pair_order:
                    pair_order.append(pair_value)
            pair_order.sort(key=lambda pair_value: min(
                min(
                    position_priority.get(part.strip(), 999)
                    for part in str(position).split("/")
                )
                for position in all_pair_view.loc[
                    all_pair_view["2D"].astype(str).eq(pair_value),
                    "Kedudukan 2D",
                ].astype(str).unique()
            ))
            st.markdown("**Semua pair 2nd Prize**")
            for pair_index, pair_value in enumerate(pair_order, start=1):
                pair_df = all_pair_view[
                    all_pair_view["2D"].astype(str).eq(pair_value)
                ].copy()
                pair_positions = list(dict.fromkeys(
                    pair_df["Kedudukan 2D"].astype(str).tolist()
                ))
                pair_numbers = list(dict.fromkeys(
                    pair_df["No Terhasil"].astype(str).tolist()
                ))
                position_text = " / ".join(pair_positions)
                with st.expander(
                    f"Pair {pair_value} — kedudukan {position_text} "
                    f"({len(pair_numbers)} pilihan)",
                    expanded=False,
                ):
                    st.markdown(
                        f"**Pilihan:** {' / '.join(pair_numbers) or 'Tiada'}"
                    )
                    pair_text = (
                        "Rumah A Predictor - 2D + Missing + Digit 1st\n\n"
                        f"2nd Prize: {_pad4(second)}\n"
                        f"Pair: {pair_value}\n"
                        f"Kedudukan 2D: {position_text}\n"
                        f"Kedudukan digit 1st: {selected_first_text}\n"
                        f"Missing: {missing_text}\n\n"
                        f"Pilihan (Total: {len(pair_numbers)}):\n"
                        f"{' / '.join(pair_numbers) or 'Tiada'}"
                    )
                    copy_button_clean(
                        f"📋 Copy Pair {pair_value}",
                        pair_text,
                        f"copy_2d_missing_pair_{pair_index}_{pair_value}",
                    )
                    st.dataframe(
                        pair_df,
                        hide_index=True,
                        use_container_width=True,
                    )
        with st.expander(
            "Lihat audit gabungan 2D + digit 1st", expanded=False
        ):
            st.dataframe(
                joint_view, hide_index=True, use_container_width=True,
            )
    except Exception as e:
        st.warning(f"Engine 2D + Missing + Digit 1st belum dapat dipaparkan: {e}")

    # -----------------------------
    # Selection 2D + Missing yang benar-benar terdapat dalam Bridge
    # -----------------------------
    st.markdown("**2nd 2D + Missing — Bridge V1**")
    try:
        bridge_selection_engine = route_engine
        selected_first_positions = {1, 2, 3, 4}
        source_rows = bridge_selection_engine["all"]
        source_rows = source_rows[
            source_rows["Kedudukan Digit 1st"].isin(selected_first_positions)
        ].copy() if not source_rows.empty else source_rows.copy()

        filtered_rows = []
        for _, source_row in source_rows.iterrows():
            for bridge_name in ("Bridge V1", "Bridge V2"):
                bridge_number = str(source_row.get(bridge_name, "")).strip()
                if not bridge_number or bridge_number.lower() == "nan":
                    continue
                filtered_rows.append({
                    "Pair": str(source_row["2D"]),
                    "Kedudukan 2D": str(source_row["Kedudukan 2D"]),
                    "No Formula": str(source_row["No Terhasil"]),
                    "Bridge": bridge_name.replace("Bridge ", ""),
                    "No Pilihan": _pad4(bridge_number),
                })
        bridge_selection_df = pd.DataFrame(
            filtered_rows,
            columns=["Pair", "Kedudukan 2D", "No Formula", "Bridge", "No Pilihan"],
        )
        if not bridge_selection_df.empty:
            bridge_selection_df["Family Key"] = bridge_selection_df["No Pilihan"].map(_key4)
            bridge_selection_df = bridge_selection_df.drop_duplicates(
                subset=["Pair", "Kedudukan 2D", "Family Key", "Bridge"]
            ).reset_index(drop=True)
            unique_selection_df = bridge_selection_df.drop_duplicates(
                subset=["Family Key"]
            ).reset_index(drop=True)
            selection_numbers = unique_selection_df["No Pilihan"].astype(str).tolist()
        else:
            unique_selection_df = bridge_selection_df.copy()
            selection_numbers = []

        st.markdown(f"**Jumlah Pilihan:** {len(selection_numbers)}")
        st.markdown(f"**Pilihan Bridge:** {' / '.join(selection_numbers) or 'Tiada'}")
        selection_copy_text = (
            "Rumah A Predictor - 2D Missing Bridge Selection\n\n"
            f"Jumlah Pilihan: {len(selection_numbers)}\n"
            f"{' / '.join(selection_numbers) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy 2D Missing Bridge Selection",
            selection_copy_text,
            "copy_2d_missing_bridge_selection",
        )

        if not bridge_selection_df.empty:
            position_priority = {
                str(row["Kedudukan 2D"]): priority
                for priority, (_, row) in enumerate(
                    bridge_selection_engine["second_audit"].iterrows(), start=1
                )
            }
            pair_order = list(dict.fromkeys(
                bridge_selection_df["Pair"].astype(str).tolist()
            ))
            pair_order.sort(key=lambda pair_value: min(
                min(
                    position_priority.get(part.strip(), 999)
                    for part in str(position).split("/")
                )
                for position in bridge_selection_df.loc[
                    bridge_selection_df["Pair"].astype(str).eq(pair_value),
                    "Kedudukan 2D",
                ].astype(str).unique()
            ))
            for pair_index, pair_value in enumerate(pair_order, start=1):
                pair_df = bridge_selection_df[
                    bridge_selection_df["Pair"].astype(str).eq(pair_value)
                ].copy()
                pair_unique = pair_df.drop_duplicates(subset=["Family Key"])
                pair_numbers = pair_unique["No Pilihan"].astype(str).tolist()
                pair_positions = " / ".join(dict.fromkeys(
                    pair_df["Kedudukan 2D"].astype(str).tolist()
                ))
                with st.expander(
                    f"Pair {pair_value} — {pair_positions} "
                    f"({len(pair_numbers)} pilihan Bridge)",
                    expanded=False,
                ):
                    st.markdown(
                        f"**Pilihan Bridge:** {' / '.join(pair_numbers) or 'Tiada'}"
                    )
                    pair_copy_text = (
                        "Rumah A Predictor - 2D Missing Bridge Selection\n\n"
                        f"Pair: {pair_value}\n"
                        f"Kedudukan: {pair_positions}\n"
                        f"Pilihan (Total: {len(pair_numbers)}):\n"
                        f"{' / '.join(pair_numbers) or 'Tiada'}"
                    )
                    copy_button_clean(
                        f"📋 Copy Pair {pair_value}",
                        pair_copy_text,
                        f"copy_2d_missing_bridge_pair_{pair_index}_{pair_value}",
                    )
                    st.dataframe(
                        pair_df.drop(columns=["Family Key"], errors="ignore"),
                        hide_index=True,
                        use_container_width=True,
                    )
    except Exception as e:
        st.warning(f"2D Missing Bridge Selection belum dapat dipaparkan: {e}")

    # -----------------------------
    # 2D 2nd + dua digit daripada 1st dan 3rd (ikut kemunculan sebenar)
    # -----------------------------
    st.markdown("**2nd 2D + Digit 1st & 3rd**")
    try:
        first_third_engine = build_2d_first_third_pair_engine(
            st.session_state.history, first, second, third, lookback=100
        )
        top_positions = first_third_engine["selected_positions"]
        top_position_text = " / ".join(top_positions) or "Tiada"
        top_df = first_third_engine["selected"]
        all_df = first_third_engine["all"]
        top_numbers = list(dict.fromkeys(
            top_df["No Terhasil"].astype(str).tolist()
        )) if not top_df.empty else []
        all_numbers = list(dict.fromkeys(
            all_df["No Terhasil"].astype(str).tolist()
        )) if not all_df.empty else []
        suffix_text = " / ".join(first_third_engine["suffixes"]) or "Tiada"
        extended_audit = first_third_extended
        top_digit_all_text = " / ".join(
            extended_audit["top_digit_all"]
        ) or "Tiada"
        top_digit_selected_text = " / ".join(
            extended_audit["top_digit_selected"]
        ) or "Tiada"
        top_pair_all_text = " / ".join(
            extended_audit["top_pair_all"]
        ) or "Tiada"
        top_pair_selected_text = " / ".join(
            extended_audit["top_pair_selected"]
        ) or "Tiada"

        st.markdown(
            f"**2D utama:** {top_position_text} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"**Pair digit utama:** {top_pair_selected_text}"
        )
        all_copy_text = (
            "Rumah A Predictor - 2D + Digit 1st & 3rd\n\n"
            f"2nd Prize: {_pad4(second)}\n"
            f"1st Prize: {_pad4(first)}\n"
            f"3rd Prize: {_pad4(third)}\n"
            f"Pair digit 1st + 3rd: {suffix_text}\n\n"
            f"Semua Pilihan (Total: {len(all_numbers)}):\n"
            f"{' / '.join(all_numbers) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy Semua 2D + Digit 1st & 3rd",
            all_copy_text,
            "copy_all_2d_first_third_pair",
        )

        # Senarai mentah semua pair tidak dipaparkan. Pair yang benar-benar
        # terdapat dalam Bridge kekal tersedia di bahagian Selection.
        if False and not all_df.empty:
            position_priority = {
                str(row["Kedudukan 2D"]): priority
                for priority, (_, row) in enumerate(
                    first_third_engine["audit"].iterrows(), start=1
                )
            }
            pair_order = list(dict.fromkeys(all_df["2D"].astype(str).tolist()))
            pair_order.sort(key=lambda pair_value: min(
                min(
                    position_priority.get(part.strip(), 999)
                    for part in str(position).split("/")
                )
                for position in all_df.loc[
                    all_df["2D"].astype(str).eq(pair_value), "Kedudukan 2D"
                ].astype(str).unique()
            ))
            for pair_index, pair_value in enumerate(pair_order, start=1):
                pair_df = all_df[all_df["2D"].astype(str).eq(pair_value)].copy()
                pair_positions = " / ".join(dict.fromkeys(
                    pair_df["Kedudukan 2D"].astype(str).tolist()
                ))
                pair_numbers = list(dict.fromkeys(
                    pair_df["No Terhasil"].astype(str).tolist()
                ))
                with st.expander(
                    f"Pair {pair_value} — {pair_positions} "
                    f"({len(pair_numbers)} pilihan)",
                    expanded=False,
                ):
                    st.markdown(f"**Pilihan:** {' / '.join(pair_numbers)}")
                    pair_copy_text = (
                        "Rumah A Predictor - 2D + Digit 1st & 3rd\n\n"
                        f"Pair: {pair_value}\n"
                        f"Kedudukan: {pair_positions}\n"
                        f"Pilihan (Total: {len(pair_numbers)}):\n"
                        f"{' / '.join(pair_numbers)}"
                    )
                    copy_button_clean(
                        f"📋 Copy Pair {pair_value}",
                        pair_copy_text,
                        f"copy_2d_first_third_pair_{pair_index}_{pair_value}",
                    )
                    st.dataframe(pair_df, hide_index=True, use_container_width=True)
        with st.expander("Lihat audit gabungan 2D + pair digit", expanded=False):
            st.dataframe(
                extended_audit["joint_audit"], hide_index=True,
                use_container_width=True,
            )
    except Exception as e:
        st.warning(f"2D + Digit 1st & 3rd belum dapat dipaparkan: {e}")

    # -----------------------------
    # Selection: hasil 2D + Digit 1st/3rd yang terdapat dalam Bridge
    # -----------------------------
    st.markdown("**2nd 2D + Digit 1st & 3rd — Bridge V2**")
    try:
        ft_selection_engine = first_third_engine
        ft_source_df = ft_selection_engine["all"].copy()
        v1_lookup = {
            _key4(number): _pad4(number)
            for number in bridge_df.get("No", pd.Series(dtype=str)).astype(str)
        } if bridge_df is not None and not bridge_df.empty else {}
        v2_lookup = {
            _key4(number): _pad4(number)
            for number in bridge_v2_df.get("No", pd.Series(dtype=str)).astype(str)
        } if bridge_v2_df is not None and not bridge_v2_df.empty else {}

        ft_filtered_rows = []
        for _, source_row in ft_source_df.iterrows():
            formula_number = str(source_row["No Terhasil"])
            family_key = _key4(formula_number)
            for bridge_name, lookup in (("V1", v1_lookup), ("V2", v2_lookup)):
                bridge_number = lookup.get(family_key, "")
                if not bridge_number:
                    continue
                ft_filtered_rows.append({
                    "Pair": str(source_row["2D"]),
                    "Kedudukan 2D": str(source_row["Kedudukan 2D"]),
                    "Pair 1st+3rd": str(source_row["Pair 1st+3rd"]),
                    "No Formula": formula_number,
                    "Bridge": bridge_name,
                    "No Pilihan": bridge_number,
                    "Family Key": family_key,
                })
        ft_bridge_df = pd.DataFrame(ft_filtered_rows)
        if not ft_bridge_df.empty:
            ft_bridge_df = ft_bridge_df.drop_duplicates(
                subset=["Pair", "Kedudukan 2D", "Family Key", "Bridge"]
            ).reset_index(drop=True)
            ft_unique_df = ft_bridge_df.drop_duplicates(
                subset=["Family Key"]
            ).reset_index(drop=True)
            ft_selection_numbers = ft_unique_df["No Pilihan"].astype(str).tolist()
        else:
            ft_unique_df = ft_bridge_df.copy()
            ft_selection_numbers = []

        st.markdown(f"**Jumlah Pilihan:** {len(ft_selection_numbers)}")
        st.markdown(
            f"**Pilihan Bridge:** {' / '.join(ft_selection_numbers) or 'Tiada'}"
        )
        ft_selection_text = (
            "Rumah A Predictor - 2D 1st & 3rd Bridge Selection\n\n"
            f"Jumlah Pilihan: {len(ft_selection_numbers)}\n"
            f"{' / '.join(ft_selection_numbers) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy 2D 1st & 3rd Bridge Selection",
            ft_selection_text,
            "copy_2d_first_third_bridge_selection",
        )

        if not ft_bridge_df.empty:
            position_priority = {
                str(row["Kedudukan 2D"]): priority
                for priority, (_, row) in enumerate(
                    ft_selection_engine["audit"].iterrows(), start=1
                )
            }
            pair_order = list(dict.fromkeys(
                ft_bridge_df["Pair"].astype(str).tolist()
            ))
            pair_order.sort(key=lambda pair_value: min(
                min(
                    position_priority.get(part.strip(), 999)
                    for part in str(position).split("/")
                )
                for position in ft_bridge_df.loc[
                    ft_bridge_df["Pair"].astype(str).eq(pair_value),
                    "Kedudukan 2D",
                ].astype(str).unique()
            ))
            for pair_index, pair_value in enumerate(pair_order, start=1):
                pair_df = ft_bridge_df[
                    ft_bridge_df["Pair"].astype(str).eq(pair_value)
                ].copy()
                pair_unique = pair_df.drop_duplicates(subset=["Family Key"])
                pair_numbers = pair_unique["No Pilihan"].astype(str).tolist()
                pair_positions = " / ".join(dict.fromkeys(
                    pair_df["Kedudukan 2D"].astype(str).tolist()
                ))
                with st.expander(
                    f"Pair {pair_value} — {pair_positions} "
                    f"({len(pair_numbers)} pilihan Bridge)",
                    expanded=False,
                ):
                    st.markdown(
                        f"**Pilihan Bridge:** {' / '.join(pair_numbers) or 'Tiada'}"
                    )
                    pair_text = (
                        "Rumah A Predictor - 2D 1st & 3rd Bridge Selection\n\n"
                        f"Pair: {pair_value}\n"
                        f"Kedudukan: {pair_positions}\n"
                        f"Pilihan (Total: {len(pair_numbers)}):\n"
                        f"{' / '.join(pair_numbers) or 'Tiada'}"
                    )
                    copy_button_clean(
                        f"📋 Copy Pair {pair_value}",
                        pair_text,
                        f"copy_2d_first_third_bridge_pair_{pair_index}_{pair_value}",
                    )
                    st.dataframe(
                        pair_df.drop(columns=["Family Key"], errors="ignore"),
                        hide_index=True,
                        use_container_width=True,
                    )
    except Exception as e:
        st.warning(f"2D 1st & 3rd Bridge Selection belum dapat dipaparkan: {e}")

    # -----------------------------
    # -----------------------------
    # 3rd 2D engines
    # -----------------------------
    st.markdown(
        '<div class="engine-head engine-support">3rd 2D Engine</div>',
        unsafe_allow_html=True,
    )
    # 3rd 2D + Missing + Digit 1st (laluan bebas)
    st.markdown("**3rd 2D + Missing + Digit 1st**")
    try:
        third_missing_engine = third_missing_route
        third_missing_selected = third_missing_engine["selected"]
        third_missing_all = third_missing_engine["all"]
        third_missing_positions_text = " / ".join(
            third_missing_engine["selected_positions"]
        ) or "Tiada"
        third_missing_text = " / ".join(third_missing_engine["missing"]) or "Tiada"
        third_missing_raw_numbers = list(dict.fromkeys(
            third_missing_selected.get("No Terhasil", pd.Series(dtype=str)).astype(str).tolist()
        ))
        st.markdown(
            f"**3rd 2D utama:** {third_missing_positions_text} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"**Missing:** {third_missing_text}"
        )
        third_missing_raw_copy = (
            "Rumah A Predictor - 3rd 2D + Missing + Digit 1st\n\n"
            f"3rd Prize: {_pad4(third)}\n"
            f"Kedudukan 3rd 2D: {third_missing_positions_text}\n"
            f"Missing: {third_missing_text}\n\n"
            f"Pilihan (Total: {len(third_missing_raw_numbers)}):\n"
            f"{' / '.join(third_missing_raw_numbers) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy 3rd 2D + Missing", third_missing_raw_copy,
            "copy_third_missing_all",
        )
        with st.expander("Lihat audit 3rd 2D + Missing", expanded=False):
            st.dataframe(
                third_missing_engine["joint_audit"], hide_index=True,
                use_container_width=True,
            )
    except Exception as e:
        st.warning(f"3rd 2D + Missing belum dapat dipaparkan: {e}")

    st.markdown("**3rd 2D + Missing — Bridge V1**")
    try:
        third_missing_bridge_rows = []
        for _, source_row in third_missing_all.iterrows():
            bridge_number = str(source_row.get("Bridge V1", "")).strip()
            if not bridge_number or bridge_number.lower() == "nan":
                continue
            third_missing_bridge_rows.append({
                "3rd 2D": str(source_row["2D"]),
                "Kedudukan": str(source_row["Kedudukan 3rd 2D"]),
                "No Formula": str(source_row["No Terhasil"]),
                "Bridge": "V1", "No Pilihan": _pad4(bridge_number),
                "Key": _key4(bridge_number),
            })
        third_missing_bridge_df = pd.DataFrame(third_missing_bridge_rows)
        if not third_missing_bridge_df.empty:
            third_missing_bridge_df = third_missing_bridge_df.drop_duplicates(
                subset=["Key"]
            ).reset_index(drop=True)
            third_missing_bridge_numbers = third_missing_bridge_df["No Pilihan"].astype(str).tolist()
        else:
            third_missing_bridge_numbers = []
        st.markdown(f"**Jumlah Pilihan:** {len(third_missing_bridge_numbers)}")
        st.markdown(
            f"**Pilihan Bridge V1:** {' / '.join(third_missing_bridge_numbers) or 'Tiada'}"
        )
        third_missing_bridge_copy = (
            "Rumah A Predictor - 3rd Missing Bridge Selection\n\n"
            f"Jumlah Pilihan: {len(third_missing_bridge_numbers)}\n"
            f"{' / '.join(third_missing_bridge_numbers) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy 3rd Missing Bridge Selection", third_missing_bridge_copy,
            "copy_third_missing_bridge",
        )
        if not third_missing_bridge_df.empty:
            for pair_index, pair_value in enumerate(
                dict.fromkeys(third_missing_bridge_df["3rd 2D"].astype(str)), start=1
            ):
                pair_df = third_missing_bridge_df[
                    third_missing_bridge_df["3rd 2D"].astype(str).eq(pair_value)
                ].copy()
                pair_numbers = list(dict.fromkeys(pair_df["No Pilihan"].astype(str)))
                pair_positions = " / ".join(dict.fromkeys(
                    pair_df["Kedudukan"].astype(str).tolist()
                ))
                with st.expander(
                    f"Pair {pair_value} — {pair_positions} "
                    f"({len(pair_numbers)} pilihan Bridge)", expanded=False,
                ):
                    st.markdown(f"**Pilihan Bridge V1:** {' / '.join(pair_numbers)}")
                    copy_button_clean(
                        f"📋 Copy Pair {pair_value}",
                        "Rumah A Predictor - 3rd 2D + Missing — Bridge V1\n\n"
                        f"Pair: {pair_value}\nKedudukan: {pair_positions}\n"
                        f"Pilihan (Total: {len(pair_numbers)}):\n"
                        f"{' / '.join(pair_numbers)}",
                        f"copy_third_missing_pair_{pair_index}_{pair_value}",
                    )
        with st.expander("Lihat detail 3rd Missing Bridge Selection", expanded=False):
            st.dataframe(
                third_missing_bridge_df.drop(columns=["Key"], errors="ignore"),
                hide_index=True, use_container_width=True,
            )
    except Exception as e:
        st.warning(f"3rd Missing Bridge Selection belum dapat dipaparkan: {e}")

    # -----------------------------
    # 3rd 2D + Digit 1st & 2nd (laluan bebas)
    # -----------------------------
    st.markdown("**3rd 2D + Digit 1st & 2nd**")
    try:
        third_first_second_engine = third_first_second_route
        third_first_second_selected = third_first_second_engine["selected"]
        third_first_second_all = third_first_second_engine["all"]
        third_first_second_positions_text = " / ".join(
            third_first_second_engine["selected_positions"]
        ) or "Tiada"
        third_first_second_raw_numbers = list(dict.fromkeys(
            third_first_second_selected.get(
                "No Terhasil", pd.Series(dtype=str)
            ).astype(str).tolist()
        ))
        st.markdown(f"**3rd 2D utama:** {third_first_second_positions_text}")
        third_first_second_raw_copy = (
            "Rumah A Predictor - 3rd 2D + Digit 1st & 2nd\n\n"
            f"3rd Prize: {_pad4(third)}\n"
            f"Kedudukan 3rd 2D: {third_first_second_positions_text}\n\n"
            f"Pilihan (Total: {len(third_first_second_raw_numbers)}):\n"
            f"{' / '.join(third_first_second_raw_numbers) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy 3rd 2D + 1st & 2nd", third_first_second_raw_copy,
            "copy_third_first_second_all",
        )
        with st.expander("Lihat audit 3rd 2D + 1st & 2nd", expanded=False):
            st.dataframe(
                third_first_second_engine["audit"], hide_index=True,
                use_container_width=True,
            )
    except Exception as e:
        st.warning(f"3rd 2D + 1st & 2nd belum dapat dipaparkan: {e}")

    st.markdown("**3rd 2D + Digit 1st & 2nd — Bridge V2**")
    try:
        third_first_second_bridge_rows = []
        for _, source_row in third_first_second_all.iterrows():
            formula = str(source_row["No Terhasil"])
            key = _key4(formula)
            bridge_number = v2_route_lookup.get(key, "")
            if not bridge_number:
                continue
            third_first_second_bridge_rows.append({
                "3rd 2D": str(source_row["2D"]),
                "Kedudukan": str(source_row["Kedudukan 3rd 2D"]),
                "Pair 1st+2nd": str(source_row["Pair 1st+2nd"]),
                "No Formula": formula, "Bridge": "V2",
                "No Pilihan": bridge_number, "Key": key,
            })
        third_first_second_bridge_df = pd.DataFrame(third_first_second_bridge_rows)
        if not third_first_second_bridge_df.empty:
            third_first_second_bridge_df = third_first_second_bridge_df.drop_duplicates(
                subset=["Key"]
            ).reset_index(drop=True)
            third_first_second_bridge_numbers = third_first_second_bridge_df[
                "No Pilihan"
            ].astype(str).tolist()
        else:
            third_first_second_bridge_numbers = []
        st.markdown(f"**Jumlah Pilihan:** {len(third_first_second_bridge_numbers)}")
        st.markdown(
            f"**Pilihan Bridge V2:** {' / '.join(third_first_second_bridge_numbers) or 'Tiada'}"
        )
        third_first_second_bridge_copy = (
            "Rumah A Predictor - 3rd 1st & 2nd Bridge Selection\n\n"
            f"Jumlah Pilihan: {len(third_first_second_bridge_numbers)}\n"
            f"{' / '.join(third_first_second_bridge_numbers) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy 3rd 1st & 2nd Bridge Selection",
            third_first_second_bridge_copy, "copy_third_first_second_bridge",
        )
        if not third_first_second_bridge_df.empty:
            for pair_index, pair_value in enumerate(
                dict.fromkeys(third_first_second_bridge_df["3rd 2D"].astype(str)), start=1
            ):
                pair_df = third_first_second_bridge_df[
                    third_first_second_bridge_df["3rd 2D"].astype(str).eq(pair_value)
                ].copy()
                pair_numbers = list(dict.fromkeys(pair_df["No Pilihan"].astype(str)))
                pair_positions = " / ".join(dict.fromkeys(
                    pair_df["Kedudukan"].astype(str).tolist()
                ))
                with st.expander(
                    f"Pair {pair_value} — {pair_positions} "
                    f"({len(pair_numbers)} pilihan Bridge)", expanded=False,
                ):
                    st.markdown(f"**Pilihan Bridge V2:** {' / '.join(pair_numbers)}")
                    copy_button_clean(
                        f"📋 Copy Pair {pair_value}",
                        "Rumah A Predictor - 3rd 2D + 1st & 2nd — Bridge V2\n\n"
                        f"Pair: {pair_value}\nKedudukan: {pair_positions}\n"
                        f"Pilihan (Total: {len(pair_numbers)}):\n"
                        f"{' / '.join(pair_numbers)}",
                        f"copy_third_pair_route_{pair_index}_{pair_value}",
                    )
        with st.expander("Lihat detail 3rd 1st & 2nd Bridge Selection", expanded=False):
            st.dataframe(
                third_first_second_bridge_df.drop(columns=["Key"], errors="ignore"),
                hide_index=True, use_container_width=True,
            )
    except Exception as e:
        st.warning(f"3rd 1st & 2nd Bridge Selection belum dapat dipaparkan: {e}")

    _old_triple_match_hidden = """
    # -----------------------------
    # Triple Match: persetujuan tiga laluan bebas, V1 dan V2 kekal berasingan
    # -----------------------------
    st.markdown(
        '<div class="engine-head engine-support">Triple Match</div>',
        unsafe_allow_html=True,
    )
    try:
        def _family_set(frame, bridge_name=None):
            if frame is None or frame.empty:
                return set()
            view = frame
            if bridge_name is not None and "Bridge" in view.columns:
                view = view[view["Bridge"].astype(str).eq(bridge_name)]
            if "Family Key" in view.columns:
                return set(view["Family Key"].dropna().astype(str))
            if "Key" in view.columns:
                return set(view["Key"].dropna().astype(str))
            if "No Pilihan" in view.columns:
                return {
                    _key4(number)
                    for number in view["No Pilihan"].dropna().astype(str)
                }
            return set()

        def _ordered_bridge_numbers(families, lookup):
            wanted = set(families)
            return [
                number for family, number in lookup.items()
                if family in wanted
            ]

        # V1 = Bridge V1 + 2nd Missing + 3rd Missing.
        v1_second_families = _family_set(bridge_selection_df, "V1")
        v1_third_families = _family_set(third_missing_bridge_df, "V1")
        v1_triple_families = v1_second_families & v1_third_families
        v1_double_only_families = (
            v1_second_families ^ v1_third_families
        )
        v1_triple_numbers = _ordered_bridge_numbers(
            v1_triple_families, v1_route_lookup
        )
        v1_double_only_numbers = _ordered_bridge_numbers(
            v1_double_only_families, v1_route_lookup
        )

        # V2 = Bridge V2 + 2nd 1st&3rd + 3rd 1st&2nd.
        v2_second_families = _family_set(ft_bridge_df, "V2")
        v2_third_families = _family_set(
            third_first_second_bridge_df, "V2"
        )
        v2_triple_families = v2_second_families & v2_third_families
        v2_double_only_families = (
            v2_second_families ^ v2_third_families
        )
        v2_triple_numbers = _ordered_bridge_numbers(
            v2_triple_families, v2_route_lookup
        )
        v2_double_only_numbers = _ordered_bridge_numbers(
            v2_double_only_families, v2_route_lookup
        )

        st.markdown("**Triple Match V1**")
        st.markdown(
            f"{' / '.join(v1_triple_numbers) or 'Tiada'}"
        )
        v1_triple_copy = (
            "Rumah A Predictor - Triple Match V1\n\n"
            "Bridge V1 + 2nd Missing + 3rd Missing\n"
            f"Jumlah Pilihan: {len(v1_triple_numbers)}\n"
            f"{' / '.join(v1_triple_numbers) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy Triple Match V1", v1_triple_copy,
            "copy_triple_match_v1",
        )

        st.markdown("**Triple Match V2**")
        st.markdown(
            f"{' / '.join(v2_triple_numbers) or 'Tiada'}"
        )
        v2_triple_copy = (
            "Rumah A Predictor - Triple Match V2\n\n"
            "Bridge V2 + 2nd 1st & 3rd + 3rd 1st & 2nd\n"
            f"Jumlah Pilihan: {len(v2_triple_numbers)}\n"
            f"{' / '.join(v2_triple_numbers) or 'Tiada'}"
        )
        copy_button_clean(
            "📋 Copy Triple Match V2", v2_triple_copy,
            "copy_triple_match_v2",
        )

        with st.expander("Lihat Double Match (2 daripada 3)", expanded=False):
            st.markdown(
                f"**Double Match V1:** "
                f"{' / '.join(v1_double_only_numbers) or 'Tiada'}"
            )
            v1_double_copy = (
                "Rumah A Predictor - Double Match V1\n\n"
                f"Jumlah Pilihan: {len(v1_double_only_numbers)}\n"
                f"{' / '.join(v1_double_only_numbers) or 'Tiada'}"
            )
            copy_button_clean(
                "📋 Copy Double Match V1", v1_double_copy,
                "copy_double_match_v1",
            )
            st.markdown(
                f"**Double Match V2:** "
                f"{' / '.join(v2_double_only_numbers) or 'Tiada'}"
            )
            v2_double_copy = (
                "Rumah A Predictor - Double Match V2\n\n"
                f"Jumlah Pilihan: {len(v2_double_only_numbers)}\n"
                f"{' / '.join(v2_double_only_numbers) or 'Tiada'}"
            )
            copy_button_clean(
                "📋 Copy Double Match V2", v2_double_copy,
                "copy_double_match_v2",
            )
    except Exception as e:
        st.warning(f"Triple Match belum dapat dipaparkan: {e}")

    """
    _legacy_hidden_ui = """Paparan legacy disimpan dalam kod tetapi tidak dijalankan.

    # Selection Engine V1
    # -----------------------------
    st.markdown('<div class="engine-head engine-support">Selection Engine</div>', unsafe_allow_html=True)
    try:
        selection = build_selection_engine(
            st.session_state.history, first, second, third, lookback=100
        )
        selection_numbers = selection.get("combined", [])
        st.markdown(f'**Pilihan 10:** {" / ".join(selection_numbers) or "Tiada"}')
        selection_text = (
            "Rumah A Predictor - Selection Engine\n\n"
            f'Pilihan 10:\n{" / ".join(selection_numbers) or "Tiada"}'
        )
        copy_button_clean(
            "\U0001F4CB Copy Selection",
            selection_text,
            "copy_selection_engine_v1",
        )
    except Exception as e:
        st.warning(f"Selection Engine belum dapat dipaparkan: {e}")

    # -----------------------------
    # Bridge Pair Priority - pair carry-forward daripada Top 3
    # -----------------------------
    st.markdown('<div class="engine-head engine-pair">Bridge Pair Shortlist</div>', unsafe_allow_html=True)
    st.caption(
        "Pair disusun berdasarkan satu hit gabungan V1/V2 bagi 100 draw terkini. "
        "Jika V1 dan V2 sama-sama hit dalam satu draw, ia tetap dikira sekali. Buka pair yang dikehendaki; "
        "nombor dan butang Copy bagi pair itu sahaja tersedia di dalamnya."
    )
    try:
        pair_priority_df = build_bridge_pair_priority(
            st.session_state.history, first, second, third
        )
        if pair_priority_df.empty:
            st.info("Data sejarah belum mencukupi untuk Bridge Pair Shortlist.")
        else:
            ranking_text = " / ".join(
                f'#{int(row["Priority"])} {row["Current Pair"]}'
                for _, row in pair_priority_df.iterrows()
            )
            st.markdown(f"**Ranking Pair:** {ranking_text}")

            shown_pairs = set()
            for _, audit_row in pair_priority_df.iterrows():
                pair = str(audit_row["Current Pair"]).zfill(2)[-2:]
                if pair in shown_pairs:
                    continue
                shown_pairs.add(pair)
                pair_numbers_df, pair_copy_text = build_bridge_pair_priority_numbers(
                    pair, audit_row, first, second, third
                )
                sources = pair_priority_df[pair_priority_df["Current Pair"].astype(str).str.zfill(2) == pair]
                source_text = " / ".join(
                    f'{row["Source"]} {row["Pair Position"]}' for _, row in sources.iterrows()
                )
                label = (
                    f'#{int(audit_row["Priority"])} Pair {pair} | '
                    f'Hit {int(audit_row["Total Support"])}/{int(audit_row["Transitions"])}'
                )
                with st.expander(label, expanded=False):
                    st.caption(
                        f'Sumber semasa: {source_text} | '
                        f'Hit Gabungan: {int(audit_row["Total Support"])} '
                        f'({float(audit_row["Hit Rate %"]):.1f}%) | '
                        f'V1 Hit: {int(audit_row["V1 Hit"])} | '
                        f'V2 Hit: {int(audit_row["V2 Hit"])}'
                    )
                    copy_button_clean(
                        f"📋 Copy Pair {pair}",
                        pair_copy_text,
                        f"bridge_pair_shortlist_{pair}_v31_35_5",
                    )
                    v1_rows = pair_numbers_df[pair_numbers_df["Route"] == "Bridge V1"]
                    v2_rows = pair_numbers_df[pair_numbers_df["Route"].str.startswith("Bridge V2")]
                    st.markdown(f"**Bridge V1 — {len(v1_rows)} pilihan unik**")
                    st.dataframe(v1_rows, hide_index=True, use_container_width=True)
                    st.markdown(f"**Bridge V2 — {len(v2_rows)} pilihan unik**")
                    st.dataframe(v2_rows, hide_index=True, use_container_width=True)

            with st.expander("Lihat audit pair 100 draw terkini", expanded=False):
                st.dataframe(pair_priority_df, hide_index=True, use_container_width=True)
    except Exception as e:
        st.warning(f"Bridge Pair Shortlist belum dapat dipaparkan: {e}")

    # -----------------------------
    # Bridge Dua Pair - blok tambahan, tidak mengubah shortlist asal
    # -----------------------------
    st.markdown('<div class="engine-head engine-support">Bridge Dua Pair</div>', unsafe_allow_html=True)
    st.caption(
        "Pilihan daripada generator pair yang turut mengandungi sekurang-kurangnya "
        "satu pair lain daripada keputusan semasa. Shortlist asal di atas tidak berubah."
    )
    try:
        second_pair_rank_df = build_bridge_pair_priority(
            st.session_state.history, first, second, third
        )
        shown_second_pairs = set()
        for _, audit_row in second_pair_rank_df.iterrows():
            pair = str(audit_row["Current Pair"]).zfill(2)[-2:]
            if pair in shown_second_pairs:
                continue
            shown_second_pairs.add(pair)
            pair_numbers_df, _ = build_bridge_pair_priority_numbers(
                pair, audit_row, first, second, third
            )
            second_pair_df, second_pair_text = build_second_pair_shortlist(
                pair, pair_numbers_df, first, second, third
            )
            with st.expander(
                f'#{int(audit_row["Priority"])} Pair {pair} — {len(second_pair_df)} pilihan dua pair',
                expanded=False,
            ):
                copy_button_clean(
                    f"📋 Copy Pair Kedua {pair}",
                    second_pair_text,
                    f"second_pair_family_{pair}_v31_35_6",
                )
                if second_pair_df.empty:
                    st.info("Tiada pilihan dua pair untuk pair ini.")
                else:
                    st.dataframe(second_pair_df, hide_index=True, use_container_width=True)
    except Exception as e:
        st.warning(f"Bridge Dua Pair belum dapat dipaparkan: {e}")

    # -----------------------------
    """

    # Carta 3D V2 - Menegak/L sahaja untuk Historical Signal Engine
    # -----------------------------
    st.markdown('<div class="engine-head engine-chart">Carta 3D V2</div>', unsafe_allow_html=True)
    st.caption(
        "Jumlah digit dan campur silang untuk pilihan 3D Menegak/L. "
        "Pilihan ini menjadi input dalaman Historical Signal Engine."
    )
    try:
        chart_v2_text, _, chart_v2_meta = build_chart_3d_signal_v31_39(
            first, second, third, bridge_df, bridge_v2_df
        )
        st.code("\n".join(chart_v2_meta.get("Rows", [])), language=None)
        chart_3d_df = chart_v2_meta.get("3D Choices", pd.DataFrame())
        chart_3d_confirmed_df = chart_v2_meta.get("3D Confirmed", pd.DataFrame())
        vertical_values = chart_3d_df[chart_3d_df["Pilihan"] == "Menegak"]["3D"].tolist() if not chart_3d_df.empty else []
        l_values = chart_3d_df[chart_3d_df["Pilihan"] != "Menegak"]["3D"].tolist() if not chart_3d_df.empty else []
        st.markdown(
            f'**Pilihan Menegak:** {" / ".join(vertical_values) or "Tiada"}  \n'
            f'**Pilihan L:** {" / ".join(l_values) or "Tiada"}  \n'
            f'**Carta 3D + Bridge:** {len(chart_3d_confirmed_df)}'
        )
        copy_button_clean(
            "📋 Copy Carta 3D V2",
            chart_v2_text,
            "copy_chart_3d_v2_v31_39",
        )

    except Exception as e:
        st.warning(f"Carta 3D V2 belum dapat dipaparkan: {e}")
